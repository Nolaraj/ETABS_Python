import openpyxl as op
import sys
from openpyxl.styles import Font, Fill, Color, PatternFill  # Connect styles for text
from openpyxl.utils import get_column_letter, column_index_from_string
import comtypes.client
import main
import os
import openpyxl
Reporting_Dict = main.Reporting_Dict
Bearing_Capacity = main.Bearing_Capacity
Plan_Area = main.Plan_Area
EQ_cases = ["EQx SLS", "EQy SLS", "EQx ULS", "EQy ULS", "EQx", "EQy"]

def model_initializer():
    global SapModel
    EtabsObject = None

    myETABSObject = comtypes.client.GetActiveObject("CSI.ETABS.API.ETABSObject")

    SapModel = myETABSObject.SapModel
    ret = SapModel.SetPresentUnits(6)



def GetTableForDisplayArray():
    #Output Writing Section
    def tables_writer(sheet_name, titles, outputs, nth_iteration=1, k=0, special_case=0, h_index = 0, heading= ""):
        global wb
        global ws
        if h_index == 0:
                wb = op.Workbook()
        if special_case == 0:
            try:
                ws = wb[sheet_name]
            except:
                ws = wb.create_sheet("Sheet_1")
                ws.title = sheet_name
                ws.merge_cells('B2:M3')
                ws.cell(row=2, column=2).value = heading
                ws_h2 = ws['B2']
                ws_h2.font = Font(size=23, underline='single', color='FFBB00', bold=True, italic=False)

            row = ws.max_row
            presentUnits = SapModel.GetPresentUnits()
            ws.merge_cells('A4:C4')
            ws.cell(row=4, column=1).value = f'Present Unit = {main.units[presentUnits-1]}'
            # Titles Writer
            for i in range(0, len(titles)):
                cell = ws.cell(row=5, column=i + 1)
                cell.value = titles[i]
                ws.cell(row=5, column=i + 1).fill = PatternFill(start_color='32CD32',
                                                                 end_color='32CD32', fill_type="solid")

            # Results Writer
            for h in range(0, len(outputs)):
                for column in range(1, len(titles) + 1):
                    try:
                        cell_value = float(outputs[h][column - 1])
                    except:
                        cell_value = str(outputs[h][column - 1])
                    if cell_value == "None":
                        cell_value = ''
                    cell = ws.cell(row=row + h + 3, column=column)
                    cell.value = cell_value


            ws.freeze_panes = ws['A6']

            #Page Setup
            page_setup = ws.page_setup
            page_setup.orientation = 'landscape'
            page_setup.paperSize = 8  ##8 - A3, 9 -A4, 10- A5,


            wb.save(filename=main.result_file)
            if h_index == (len(main.Table_Keys) - 1):
                wb.close()


    #OutPut Extractor Section
    ret = SapModel.SelectObj.Group("ALL")
    for h in range(0, len(main.Table_Keys)):
        sheet_name  = main.Table_Keys[h]
        TableKey = sheet_name
        FieldKeyList = []
        GroupName = "ALL"
        TableVersion = 0
        FieldsKeysIncluded = []
        NumberRecords = 0
        TableData = []
        [FieldKeyList, TableVersion, FieldsKeysIncluded, NumberRecords, TableData,
         ret] = SapModel.DatabaseTables.GetTableForDisplayArray(TableKey,
                                                                FieldKeyList, GroupName, TableVersion,
                                                                FieldsKeysIncluded, NumberRecords, TableData)
        titles = [x for x in FieldsKeysIncluded]
        title_length = len(FieldsKeysIncluded)

        outputs = []
        data = []
        for i in range(1, len(TableData) + 1):
            data.append(TableData[i - 1])
            if i % title_length == 0:
                outputs.append(data)
                data = []

        heading = sheet_name
        while len(sheet_name) > 30:
            sheet_name = " ".join(sheet_name.split(" ")[0: - 1])

        tables_writer(sheet_name,titles, outputs, h_index=h, heading = heading)
        print(f'{main.Table_Keys[h]} has successfully been written to document')


def Excel_Modifier(Building_Dimension):
    # try:
        wb = op.load_workbook(main.result_file)


        for sheetnames in main.result_sheets:

            ws = wb[sheetnames]
            dims = {}
            for i, row in enumerate(ws.rows):
                print(i)
                if i > 3:
                    print("Entered", i)
                    for cell in row:
                        if cell.value:
                            dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
            for col, value in dims.items():
                ws.column_dimensions[get_column_letter(col)].width = value

            def uniquelist(list1):
                list2 = []
                for items in list1:
                    if items not in list2:
                        list2.append(items)
                return list2

            def column_identifier(key):
                for c in range(1, max_column + 1):
                    if str(ws.cell(row=5, column=c).value) == key:
                        return c

            if sheetnames == "Centers Of Mass And Rigidity":
                titles_extra = ["Ex", "Ey", "Ex %", "Ey %"]
                max_row = ws.max_row
                max_column  = ws.max_column

                ws.cell(row=4, column=13).value = "Length"
                ws.cell(row=4, column=14).value = f"{Building_Dimension[0]} m"
                ws.cell(row=4, column=15).value = "Width"
                ws.cell(row=4, column=16).value = f"{Building_Dimension[1]} m"


                for i in range(len(titles_extra)):
                    ws.cell(row=5, column=13+i).value = titles_extra[i]
                    ws.cell(row=5, column=i + 13).fill = PatternFill(start_color='32CD32',
                                                                    end_color='32CD32', fill_type="solid")

                for i in range(6, max_row + 1):
                    ws.cell(row=i, column=13).fill = PatternFill(start_color='00FFFF',end_color='00FFFF', fill_type="solid")
                    ws.cell(row=i, column=14).fill = PatternFill(start_color='00FFFF',end_color='00FFFF', fill_type="solid")
                    ws.cell(row=i, column=15).fill = PatternFill(start_color='00FFFF',end_color='00FFFF', fill_type="solid")
                    ws.cell(row=i, column=16).fill = PatternFill(start_color='00FFFF',end_color='00FFFF', fill_type="solid")
                    ws.cell(row=i, column=13).value = float(ws.cell(row=i, column=9).value) - float(ws.cell(row=i, column=11).value)
                    ws.cell(row=i, column=14).value = float(ws.cell(row=i, column=10).value) - float(ws.cell(row=i, column=12).value)
                    ws.cell(row=i, column=15).value = float(ws.cell(row=i, column=13).value) *100 / float(ws.cell(row=4, column=14).value[:-2])
                    ws.cell(row=i, column=16).value = float(ws.cell(row=i, column=14).value) *100 / float(ws.cell(row=4, column=16).value[:-2])

            if sheetnames == "Modal Participating Mass":
                ws1 = wb["Modal Periods And Frequencies"]

                max_row = ws.max_row
                max_column  = ws.max_column


                for i in range(6, max_row):
                    sum_Ux = float(ws.cell(row=i, column=7).value)
                    sum_Uy = float(ws.cell(row=i, column=8).value)

                    if sum_Ux > 0.9 and sum_Uy > 0.9:
                        for j in range(1, max_column+1):
                            ws.cell(row=i, column=j).fill = PatternFill(start_color='e5d2e0', end_color='e5d2e0',
                                                                         fill_type="solid")
                            ws.cell(row=8, column=j).fill = PatternFill(start_color='e5d2e0', end_color='e5d2e0',
                                                                         fill_type="solid")
                            ws.cell(row=max_row, column=j).fill = PatternFill(start_color='e5d2e0', end_color='e5d2e0',
                                                                         fill_type="solid")
                        for k in range(1, 7):
                            ws1.cell(row=i, column=k).fill = PatternFill(start_color='e5d2e0', end_color='e5d2e0',
                                                                         fill_type="solid")
                        break

            if sheetnames =="Diaphragm Max Over Avg Drifts":

                max_row = ws.max_row
                for m in range(6, max_row + 1):
                    if ws.cell(row=m, column=1).value is None:
                        max_row = m-1

                        break
                max_column = ws.max_column
                row_indexes = []
                ratio_column = column_identifier("Ratio")
                diaphragm_column = column_identifier("Item")
                row = 0

                for i in range(6, max_row+1):
                    ratio = 0.0
                    ratio1 = 0.0
                    ratio2 = 0.0




                    story_name = ws.cell(row=i, column=1).value
                    output_case = ws.cell(row=i, column=2).value
                    diaphragm = ws.cell(row=i, column=diaphragm_column).value

                    for j in range(6, max_row+1):
                        if Reporting_Dict["max_avg_eq"]:
                            if ((ws.cell(row=j, column=1).value == story_name and
                                 ws.cell(row=j, column=2).value == output_case) and
                                    ws.cell(row=j, column=diaphragm_column).value == diaphragm):
                                if output_case in EQ_cases:
                                    ratio2 = float(ws.cell(row=j, column=ratio_column).value)
                                    if ratio2 > ratio1:
                                        ratio = ratio2
                                        row = j
                                    ratio1 = ratio2
                        if Reporting_Dict["Max_Avg_All"]:

                            if ((ws.cell(row=j, column=1).value == story_name and
                                 ws.cell(row=j, column=2).value == output_case) and
                                    ws.cell(row=j, column=diaphragm_column).value == diaphragm):
                                ratio2 = float(ws.cell(row=j, column= ratio_column ).value)
                                if ratio2 > ratio1:
                                    ratio = ratio2
                                    row = j
                                ratio1 = ratio2

                    row_indexes.append(row)
                row_indexes = uniquelist(row_indexes)

                for i in range(6, 6+len(row_indexes)):
                    input_row = i - 6
                    for j in range(1, max_column + 1):
                        ws.cell(row=i, column=j).value = ws.cell(row=row_indexes[input_row], column=j).value


                for i in range(6+len(row_indexes), max_row+1):
                    for j in range(1, max_column + 1):
                        ws.cell(row=i, column=j).value = ""

            if sheetnames =="Story Drifts":
                max_row = ws.max_row
                for m in range(6, max_row + 1):
                    if ws.cell(row=m, column=1).value is None:
                        max_row = m - 1
                        break
                max_column = ws.max_column
                row_indexes = []

                drift_column = column_identifier("Drift")
                direction_column = column_identifier("Direction")
                row = 0

                for i in range(6, max_row+1):
                    ratio = 0.0
                    ratio1 = 0.0
                    ratio2 = 0.0
                    story_name = ws.cell(row=i, column=1).value
                    output_case = ws.cell(row=i, column=2).value
                    direction = ws.cell(row=i, column=direction_column).value
                    for j in range(6, max_row):
                        if Reporting_Dict["drifts_eq"]:
                            if ((ws.cell(row=j, column=1).value == story_name and
                                 ws.cell(row=j, column=2).value == output_case) and
                                    ws.cell(row=j, column=direction_column).value == direction):
                                if ws.cell(row=j, column=2).value in EQ_cases:
                                    ratio2 = float(ws.cell(row=j, column=drift_column).value)
                                    if ratio2 > ratio1:
                                        ratio = ratio2
                                        row = j
                                    ratio1 = ratio2
                        if Reporting_Dict["drifts_all"]:
                            if ((ws.cell(row=j, column=1).value == story_name and
                                 ws.cell(row=j, column=2).value == output_case) and
                                    ws.cell(row=j, column=direction_column).value == direction):
                                ratio2 = float(ws.cell(row=j, column=drift_column).value)
                                if ratio2 > ratio1:
                                    ratio = ratio2
                                    row = j
                                ratio1 = ratio2

                    row_indexes.append(row)
                row_indexes = uniquelist(row_indexes)

                for i in range(6, 6+len(row_indexes)):
                    input_row = i - 6
                    for j in range(1, max_column + 1):
                        ws.cell(row=i, column=j).value = ws.cell(row=row_indexes[input_row], column=j).value

                for i in range(6+len(row_indexes), max_row+1):
                    for j in range(1, max_column + 1):
                        ws.cell(row=i, column=j).value = ""

            if sheetnames =="Joint Reactions":
                max_row = ws.max_row
                for m in range(6, max_row + 1):
                    if ws.cell(row=m, column=1).value is None:
                        max_row = m-1

                        break
                max_column = ws.max_column
                fz_column = column_identifier("FZ")
                sum_Area = 0
                sum_Percentage = 0

                titles = ["Unfactored Load", "Isolated Footing Area", "Percentage Coverage"]

                for i in range(max_column, max_column+3):
                    ws.cell(row=5, column=i).value = titles[i-max_column]
                    ws.cell(row=5, column=i ).fill = PatternFill(start_color='32CD32',end_color='32CD32', fill_type="solid")


                for i in range(6, max_row+1):
                    ws.cell(row=i, column=max_column).value = ws.cell(row=i, column=fz_column).value / 1.5
                    ws.cell(row=i, column=max_column + 1).value = ws.cell(row=i, column=max_column).value / Bearing_Capacity
                    ws.cell(row=i, column=max_column + 2).value =  (ws.cell(row=i, column=max_column).value * 100) / Plan_Area
                    sum_Area += ws.cell(row=i, column=max_column+1).value
                    sum_Percentage += ws.cell(row=i, column=max_column+2).value

                ws.cell(row=max_row+1, column=max_column+1).value = sum_Area
                ws.cell(row=max_row+1, column=max_column + 2).value = sum_Percentage
                merge_range = f'B{max_row+2}:I{max_row+3}'
                ws.merge_cells(merge_range)
                if sum_Percentage > 50:
                    ws.cell(row=max_row + 2, column=2).value = "The footing area requirement is greater than 50% of total " \
                                                               "plan area, it is recommended to use Raft as Foundation"

                else:
                    ws.cell(row=max_row + 2, column=2).value = "The footing area requirement is less than 50% of total " \
                                                               "plan area, it is recommended to use Isolated Footing as Foundation"

                ws.cell(row=max_row + 2, column=2).alignment = openpyxl.styles.Alignment(wrap_text=True)

                dims = {}
                # for i, row in enumerate(ws.rows):
                #     print(i)
                #     if i > 3:
                #         print("Entered", i)
                #         for cell in row:
                #             if cell.value:
                #                 dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
                # for col, value in dims.items():
                #     ws.column_dimensions[get_column_letter(col)].width = value

        wb.save(main.result_file)


    #
    # except:
    #     print(f"{Excel_Modifier} Excel file not found")






