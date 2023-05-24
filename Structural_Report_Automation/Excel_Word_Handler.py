import openpyxl as op
import docxtpl as dt
import docx
import main
from docxtpl import DocxTemplate, InlineImage
from docx.shared import Inches, Cm, Mm
import os
import cv2
Reporting_Dict = main.Reporting_Dict.copy()


items_dict = {}
context_word = {}
context_image = {}
wb = op.load_workbook(main.design_file, data_only=True)
ws = wb["Designer Sheet"]
design_code = ws['C42'].value


def items_grouper():
    # .Jpg File Gouping under main.extracted_data_path directory in items_dict with the name title given by items_group list
    # items_group_excel and items_group_etabs are simply the starting two words(excluding _) of placeholders in word
    items_group_etabs = ["LR Elevation", "LR Plan", "RP Elevation", "RP Plan", "SR Elevation", "SR Plan", "PMM 3D",
                         "BC 3D",
                         "CB 3D", "GR 3D", "IP 3D", "IS 3D", "IA 3D"]
    items_group_excel = main.excel_sheets
    items_group = items_group_etabs + items_group_excel
    for file in os.listdir(main.extracted_data_path):
        if file.endswith(".jpg"):
            for item in items_group:
                if file.startswith(item):
                    if item in items_dict:
                        items_dict[item].append(file)
                    else:
                        data = [file]
                        items_dict[item] = data
    return (items_group)


def placeholder_writer(para, word_1, word_2, word_3 = "1"):
    preceeding_para = "{{" + str(word_1) + "_" + str(word_2) + "_" + str(word_3) + "}}"
    looping_para = str(word_1) + " " + str(word_2)
    writing_para = "\n" + "{{" + str(word_1) + "_" + str(word_2) + "_"
    for i in range(len(para)):
        if para[i].text == preceeding_para:  # preceeding_para        As in Word
            if looping_para in items_dict:
                for no in range(1, len(items_dict[looping_para])):  # Looping based on number of elements in directory
                    para[i].add_run(writing_para + str(no + 1) + "}}")  # Para to be written (if preceeding is find)

#Collector From Excel (Placeholders defined by Bordered Cells of Table)
def word_context():

    indexing = {}
    for i in range(1, ws.max_row):
        first_cell = ws.cell(row=i, column=1).value
        if first_cell != None:
            indexing[first_cell] = i
    for j in range(9, 15):
        for i in range(1, ws.max_row):
            cell_value = ws.cell(row=i, column=j).value
            if cell_value != None:
                context_word[cell_value] = ws.cell(row=i, column=j - 6).value
    wb.close()
#Collector From Directory(Placeholders defined by Images)
def image_context(doc):
    def image_adjuster(doc, image_path):
        img = cv2.imread(image_path)
        width, height = img.shape[1], img.shape[0]
        lb_ratio = height/width
        l = lb_ratio * main.frame_width
        if l > main.frame_height:
            context_image[placeholder_name] = InlineImage(doc, image_path, height=Inches(main.frame_height))
        else:
            context_image[placeholder_name] = InlineImage(doc, image_path, width=Inches(main.frame_width))
        if file.split(" ")[0] in ["Cover"]:
            context_image[placeholder_name] = InlineImage(doc, image_path, width=Inches(main.frame_width))

    for file in os.listdir(main.extracted_data_path):
        if file.endswith(".jpg"):
            placeholder_name = (file.replace(".jpg", "")).replace(" ", "_")
            image_path = os.path.join(main.extracted_data_path, file)
            image_adjuster(doc, image_path)
            # if file.split(" ")[0] in ["Cover"]:
            #     width = 5
            # else:
            #     width = 6
            # # context_image[placeholder_name] = InlineImage(doc, image_path, Inches(width))
            # context_image[placeholder_name] = InlineImage(doc, image_path, Inches(width))




