import User_Input
from User_Input import *
from Pre_Processings import *
from ETABS import *
from openpyxl.styles import Font, Fill, Color, PatternFill  # Connect styles for text
from openpyxl.utils import get_column_letter, column_index_from_string
import openpyxl
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, RichTextProperties, Font, RegularTextRun
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice


global wb
global ws
global wb1
global ws1
block_start_row = []
block_end_row = []
end_column = 0
block_columns = 0
total_rows = 0
writing_row = 0
data_number = 0
start_column = 0
end_column = 0
start_row = 0
end_row = 0
sheet_name = ''
data_rowset = {}
indicator = 0
Col_ColTitle = {}


def intersection(lst1, lst2):
    lst3 = [value for value in lst1 if value in lst2]
    return lst3
def Processing_Initializor():
    print(f'Post_Processing.py is running')
    global wb1
    global ws1

    try:
        wb1 = load_workbook(refined_file)
    except:
        wb1 = Workbook()
    ws1 = wb1.create_sheet("Sheet_1")
    global output_sheet
    output_sheet = f"Output {len(wb1.sheetnames) - 1}"
    ws1.title = output_sheet


def indexes_identifier(sheet_name, title_name):
    # ws = wb[sheet_name]
    global start_column_i
    global end_column_i
    global block_end_row_i
    global start_row_i
    global end_row_i
    global max_number_results
    global min_number_results
    global max_num_res_row
    global min_num_res_row
    global title_not_found
    global numberresults_row
    global number_results
    global angle_deviation_col
    global load_case_col
    block_end_row_i = []
    start_row_i = 6
    end_row_i = 0
    max_col = ws.max_column
    max_row = ws.max_row
    number_res_column = 0
    max_num_res_row = 0
    min_num_res_row = 0
    numberresults_row = []
    number_results = []
    angle_deviation_col = 0
    load_case_col = 0

    try:
        title_not_found = True
        # First Block Boundary determination
        a = 0
        for i in range(1, max_col + 1):
            if ws.cell(row=5, column=i).value == "NumberResults":
                number_res_column = i
            if ws.cell(row=5, column=i).value == "Angle Deviation":
                angle_deviation_col = i
            if ws.cell(row=5, column=i).value == "LoadCase":
                load_case_col = i

            if a == 1 and ws.cell(row=5, column=i).value == None:
                end_column_i = i
            elif ws.cell(row=5, column=i).value == None or ws.cell(row=5, column=i).value != title_name:
                a = 0
            else:
                start_column_i = i
                a = 1
                title_not_found = False

        for j in range(6, max_row + 2):
            check_point = 0
            for k in range(0, 6):
                if ws.cell(row=j + k, column=number_res_column).value == None:
                    if k == 1:  # If j == x, then number is recorded if there encountered x rows none gap
                        block_end_row_i.append(j)  # For eg if x = 2 then it notes for 3 none values rows
                    if k == 5:
                        check_point = 2  # It breaks the system of loop
                    if j == max_row:
                        block_end_row_i.append(j)
                else:
                    check_point = 0
                    break
            if check_point == 2:
                break
            try:
                end_row = block_end_row_i[0]
            except:
                pass

        # Finding largest columns for a title and its row number
        max_number_results = 0
        min_number_results = 1000
        for i in range(start_row_i, max_row):
            try:
                if int(ws.cell(row=i, column=number_res_column).value) > max_number_results:
                    max_number_results = ws.cell(row=i, column=number_res_column).value
                    max_num_res_row = i
                if (int(ws.cell(row=i, column=number_res_column).value) < min_number_results) and (
                        (int(ws.cell(row=i, column=number_res_column).value) != 0) and
                        (ws.cell(row=i, column=number_res_column).value != None)):
                    min_number_results = int(ws.cell(row=i, column=number_res_column).value)
                    min_num_res_row = i

                if title_not_found == False:
                    if ((int(ws.cell(row=i, column=number_res_column).value) != 0) and
                            (ws.cell(row=i, column=number_res_column).value != None)):
                        if int(ws.cell(row=i, column=number_res_column).value) not in number_results:
                            number_results.append(int(ws.cell(row=i, column=number_res_column).value))
                            numberresults_row.append(i)


            except:
                pass
        print(numberresults_row)
        print(number_results)
    except:
        pass
def column_copier(column_number, output_column, reference=0):
    column_letter = get_column_letter(column_number)
    for cell in ws[f'{column_letter}:{column_letter}']:
        ws1.cell(row=cell.row, column=output_column, value=cell.value)

    global block_rows
    global y_column
    global x_column
    global block_end_row
    if reference == 1:
        block_end_row = []

        max_row = ws1.max_row
        check = False
        number = 0
        x_column = output_column
        for i in range(6, max_row):
            if ws1.cell(row=i, column=output_column).value == None:
                if check:
                    block_rows = number
                    break
                check = False
            else:
                check = True
                number += 1
        for i in range(6, max_row + 2):
            check1 = 0
            for j in range(0, 6):
                if ws1.cell(row=i + j, column=output_column).value == None:
                    check1 = 1
                    if check1 == 1 and j == 1:  # If j == x, then number is recorded if there encountered x rows none gap
                        block_end_row.append(i)  # For eg if x = 2 then it notes for 3 none values rows
                    if check1 == 1 and j == 5:
                        check1 = 2
                else:
                    check1 = 0
                    break
            if check1 == 2:
                break
    elif reference == 2:
        y_column = output_column
    else:
        pass
def data_arranger():
    global total_rows
    global block_start_row
    global block_columns
    global writing_row
    global start_column
    global end_column
    global previous_column

    block_columns = output_column - previous_column
    start_column = previous_column
    end_column = output_column
    total_rows = block_end_row[-1]
    block_start_row = [(x - block_rows) for x in block_end_row]
    writing_row = total_rows + len(assembled_pretitles) * 2 + 25

    # Pretitles Writer
    up = 1
    for items in range(0, len(assembled_pretitles)):
        indexes_identifier(sheet_name, assembled_pretitles[items])
        if title_not_found is False:
            for i in range(0, len(numberresults_row)):
                for col in range(start_column + 3, end_column):
                    row1 = writing_row - up
                    # row2 = writing_row - up - 1
                    ws1.cell(row=row1, column=col).value = ws.cell(row=numberresults_row[i],
                                                                   column=start_column_i + (
                                                                           col - start_column - 3)).value
                    ws1.cell(row=row1, column=col).fill = PatternFill(start_color='ff6289', end_color='ff6289',
                                                                      fill_type="solid")


                    if col == start_column + 3:
                        ws1.cell(row=row1, column=col - 1).value = ws.cell(row=numberresults_row[i],
                                                                           column=load_case_col + (
                                                                                   col - start_column - 3)).value
                        ws1.cell(row=row1, column=col - 2).value = ws.cell(row=numberresults_row[i],
                                                                           column=angle_deviation_col + (
                                                                                   col - start_column - 3)).value
                        ws1.cell(row=row1, column=col - 3).value = assembled_pretitles[items]

                up += 1

    # Titles Writer
    for col in range(start_column + 1, end_column):
        ws1.cell(row=writing_row, column=col).value = ws1.cell(row=5,
                                                               column=col).value
        ws1.cell(row=writing_row, column=col).fill = PatternFill(start_color='5cb800', end_color='5cb800',
                                                                 fill_type="solid")
    writing_row += 1
    # Content Writer
    for i in range(0, block_rows):
        for blocks in range(0, len(block_start_row)):
            for col in range(start_column + 1, end_column):
                ws1.cell(row=writing_row, column=col).value = ws1.cell(row=block_start_row[blocks] + i,
                                                                       column=col).value
            writing_row += 1
        writing_row += 1
def normal_charts_creater():
    global writing_row
    title_column = 2
    starting_row = total_rows + len(
        assembled_pretitles) * 2 + 26  # Assembled Block start row (Obtained after assembled data)
    finishing_row = starting_row + (len(block_start_row) - 1)  # Assembled Block end row
    title_row = starting_row - 1
    ending_row = writing_row
    if data_number % 2 == 0:
        ending_row += 15

    chart = openpyxl.chart.ScatterChart()
    chart.style = 5
    chart.x_axis.title = 'Angle Variation'
    chart.y_axis.title = ws1.cell(row=title_row, column=y_column).value
    xvalues = None
    values = None
    series = None
    xvalues = openpyxl.chart.Reference(ws1, min_col=x_column,
                                       min_row=starting_row,
                                       max_row=starting_row + len(block_start_row) - 1)
    chart.title = str(sheet_name) + " - " + str(ws1.cell(row=title_row, column=y_column).value)

    for k in range(0, block_rows):
        values = openpyxl.chart.Reference(ws1, min_col=y_column,
                                          min_row=starting_row + k * (len(block_start_row) + 1),
                                          max_row=finishing_row + k * (len(block_start_row) + 1))
        try:
            series = openpyxl.chart.Series(values, xvalues, title=cases_output[k])
        except:
            series = openpyxl.chart.Series(values, xvalues)

        chart.series.append(series)
    anchor = f'{get_column_letter(x_column)}{ending_row}'
    ws1.add_chart(chart, anchor)
def horizontal_values_charts_creater():
    global writing_row
    title_column = 2
    starting_row = total_rows + len(assembled_pretitles) * 2 + 26  # Block start row (Obtained after assembled data)
    finishing_row = starting_row + (len(block_start_row) - 1)  # Block end row
    title_row = starting_row - 1
    ending_row = writing_row

    for col in range(y_column, end_column):
        ws1.cell(row=title_row - 1, column=col).value = (col + 1) - y_column

    chart = openpyxl.chart.ScatterChart()
    chart.style = 5
    chart.x_axis.title = 'Modes Variation'
    chart.y_axis.title = ws1.cell(row=title_row, column=y_column).value
    xvalues = None
    values = None
    series = None
    xvalues = openpyxl.chart.Reference(ws1, min_col=y_column, max_col=end_column - 1,
                                       min_row=title_row - 1)
    chart.title = str(sheet_name) + " - " + str(ws1.cell(row=title_row, column=y_column).value)

    for k in range(starting_row, finishing_row + 1):
        values = openpyxl.chart.Reference(ws1, min_col=y_column, max_col=end_column - 1,
                                          min_row=k)
        series = openpyxl.chart.Series(values, xvalues, title=f"Deviation Angle {deviation_angles[k - starting_row]}")
        chart.series.append(series)
    anchor = f'{get_column_letter(x_column)}{ending_row}'
    ws1.add_chart(chart, anchor)

def control_center_Analysis_Results(sheet_name, title):
    global wb
    global ws
    wb = load_workbook(f_result, data_only=True)
    ws = wb[sheet_name]
    writing_row = 0
    global output_column
    global data_number
    global previous_column

    output_column += 1
    data_number += 1
    a = 0
    previous_column = output_column - 1
    max_col = ws.max_column
    for i in range(1, max_col + 1):
        for items in refined_firstcols:
            if ws.cell(row=5, column=i).value == items:
                reference = 0
                if refined_firstcols.index(items) == 0:
                    reference = 1
                column_copier(i, output_column, reference)
                output_column += 1

        if a == 1 and ws.cell(row=5, column=i).value == None:
            reference = 0
            column_copier(i, output_column, reference)
            output_column += 1
        elif ws.cell(row=5, column=i).value == None or ws.cell(row=5, column=i).value != title:
            a = 0
        else:
            a = 1
            reference = 2
            column_copier(i, output_column, reference)
            output_column += 1

    indexes_identifier(sheet_name, title_name=title)
    data_arranger()

    if sheet_name in horizontally_valued:
        horizontal_values_charts_creater()
    else:
        normal_charts_creater()
    output_column += 1
    wb1.save(filename=refined_file)
    wb.close()
    wb1.close()
















def tables_sorter():
    def data_copier(rows, columns, heading = ''):
        global output_column
        global output_row
        block_start= []
        block_end = []
        block_height = 0
        angle_dev_col = output_column
        print(f'Writing for {heading}')

        for j in range(1, len(columns) + 1):
            for i in range(5, len(rows)+1):
                cell_value = ws.cell(row = rows[i-1], column = columns [j-1]).value
                ws1.cell(row = i, column = output_column).value = cell_value
                if j == 1 and i == 5:
                    ws1.merge_cells(f'{get_column_letter(output_column)}{3}:{get_column_letter(output_column + 2)}{4}')
                    ws1.cell(row=3, column=output_column).value = heading
                    ws_h2 = ws1[f'{get_column_letter(output_column)}{3}']
                    ws_h2.font = Font(size=23, underline='single', color='FFBB00', bold=True, italic=False)
                output_row = i
            output_column += 1

        wb1.save(filename=refined_file)



        for i in range(6, output_row+2):
            check_point = 0
            for k in range(0, 6):
                if ws1.cell(row=i + k, column=angle_dev_col).value == None:
                    if k == 1:  # If j == x, then number is recorded if there encountered x rows none gap
                        block_end.append(i)  # For eg if x = 2 then it notes for 3 none values rows
                    if k == 5:
                        check_point = 2  # It breaks the system of loop
                    if j == output_row:
                        block_end.append(i)
                else:
                    check_point = 0
                    break
            if check_point == 2:
                break
        block_height = block_end[0] - 6
        for items in block_end:
            block_start.append(items - block_height)
        return(block_start, block_end,block_height)

    def data_arranger(block_start, block_end, block_height):
        global writing_row
        charts_y_columns = []
        writing_row = output_row + 5
        x_column = 0
        series_column = 0

        # Titles Writer
        for col in range(start_column, end_column):
            ws1.cell(row=writing_row, column=col).value = ws1.cell(row=5,
                                                                   column=col).value
            ws1.cell(row=writing_row, column=col).fill = PatternFill(start_color='5cb800', end_color='5cb800',
                                                                     fill_type="solid")
            for item in values[0][3]:
                if ws1.cell(row=writing_row, column=col).value == item:
                    charts_y_columns.append(col)
            if ws1.cell(row=writing_row, column=col).value == x_title:
                x_column = col
            if ws1.cell(row=writing_row, column=col).value == series_title or ws1.cell(row=writing_row, column=col).value == "Case":
                series_column = col



        writing_row += 1
        # Content Writer
        for i in range(0, block_height):
            for row in block_start:
                for col in range(start_column, end_column):
                    ws1.cell(row=writing_row, column=col).value = ws1.cell(row=row + i,
                                                                           column=col).value
                writing_row += 1
            writing_row += 1

        return(charts_y_columns, x_column, series_column)




    def tabled_charts_creater(block_start, block_end, block_height, y_columns, x_column, series_column):
        global writing_row
        starting_row = output_row + 5 + 1 # Assembled Block start row (Obtained after assembled data)
        finishing_row = starting_row + (len(block_start) - 1)  # Assembled Block end row
        title_row = starting_row - 1
        ending_row = writing_row

        for y_column in y_columns:
            chart = openpyxl.chart.ScatterChart()
            chart.style = 5
            chart.x_axis.title = 'Angle Variation'
            chart.y_axis.title = ws1.cell(row=title_row, column=y_column).value
            xvalues = None
            values = None
            series = None
            xvalues = openpyxl.chart.Reference(ws1, min_col=x_column,
                                               min_row=starting_row,
                                               max_row=starting_row + len(block_start) - 1)
            chart.title = str(sheet_name) + " - " + str(ws1.cell(row=title_row, column=y_column).value)

            for k in range(0, block_height):
                values = openpyxl.chart.Reference(ws1, min_col=y_column,
                                                  min_row=starting_row + k * (len(block_start) + 1),
                                                  max_row=finishing_row + k * (len(block_end) + 1))
                series = openpyxl.chart.Series(values, xvalues, title=ws1.cell(row=starting_row + k * (len(block_start) + 1), column=series_column).value)
                chart.series.append(series)
            anchor = f'{get_column_letter(x_column)}{ending_row + y_columns.index(y_column) * 15}'
            ws1.add_chart(chart, anchor)


    wb = load_workbook(f_result, data_only=True)
    global output_column
    global values
    global indicator
    sort_data = User_Input.sort_data
    for key, values in sort_data.items():
        output_column += 1
        sheet_name = key
        ws = wb[sheet_name]
        start_row = 6
        rows = []
        columns = []
        columns_refined = []
        max_row = ws.max_row
        heading = ws.cell(row=2, column=6).value
        UniqueNameSet = {}
        OutputCaseSet = {}
        UniqueName_column = 0
        OutputCase_column = 0
        ColTitle_Index = {}

        checker = True
        for key_1 , value_1 in data_rowset.items():
            if key == key_1:
                checker = False

        global start_column
        global end_column
        #
        for value in values:
            title = value[0]

            if ((title == "UniqueName" or title == "OutputCase")) and OCase_UniqueNameSet_Use and (
                    indicator != 0) and checker is False:
                print("I am inside If of Value")
                print(data_rowset[key][2])
                print(data_rowset[key][2][0])
                print(title)
                print(data_rowset[key][2][0][title])
                columns.append(data_rowset[key][2][0][title])
                Col_ColTitle[data_rowset[key][2][0][title]] = title
                ColTitle_Index[title] = values.index(value)


                if OCase_UniqueNameSet_Use :
                    if title == "UniqueName":
                        UniqueName_column = data_rowset[key][2][0][title]

                    if title == "OutputCase":
                        OutputCase_column = data_rowset[key][2][0][title]

                if value == values[0]:
                    if len(value[2]) > 0:
                        value[2].insert(0, "Angle Deviation")
                        for column in value[2]:
                            columns_refined.append(data_rowset[key][2][0][column])



            else:
                global prelist
                prelist = []
                TitleName_Col = {}

                indexer(wb, sheet_name, title)
                columns.append(Excel_Extractor_Preprocessing.start_column_i)
                Col_ColTitle[Excel_Extractor_Preprocessing.start_column_i] = title
                ColTitle_Index[title] = values.index(value)


                if OCase_UniqueNameSet_Use :
                    if title == "UniqueName":
                        UniqueName_column = Excel_Extractor_Preprocessing.start_column_i

                    if title == "OutputCase":
                        OutputCase_column = Excel_Extractor_Preprocessing.start_column_i

                if value == values[0]:
                    if len(value[2]) > 0:
                        value[2].insert(0, "Angle Deviation")
                        indexer(wb, sheet_name, "Angle Deviation")
                        TitleName_Col["Angle Deviation"] = Excel_Extractor_Preprocessing.start_column_i
                        for column in value[2]:
                            indexer(wb, sheet_name, column)
                            columns_refined.append(Excel_Extractor_Preprocessing.start_column_i)

                titles = values[0][2]
                for title_refined in titles:
                    indexer(wb, sheet_name, title_refined)
                    TitleName_Col[title_refined] = Excel_Extractor_Preprocessing.start_column_i
                prelist.append(TitleName_Col)
                print(TitleName_Col)


        if len(columns) > 0:
            for column in columns:
                rows_selective = []

                if ((column == UniqueName_column or column == OutputCase_column)) and OCase_UniqueNameSet_Use and (indicator != 0):
                    print("Under Set Use")
                    #Uses Data Set for Disp and drift, Other inside else is for general purpose
                    rows_selective = []
                    data = values[ColTitle_Index[Col_ColTitle[column]]][1]

                    OutputCaseSet = data_rowset[key][0]
                    UniqueNameSet = data_rowset[key][1]
                    for item in data:
                        try:
                            item = float(item)
                        except:
                            try:
                                item = int(item)
                            except:
                                pass
                        if column == UniqueName_column:

                            try:
                                rows_selective = rows_selective + UniqueNameSet[item]
                            except:
                                try:
                                    rows_selective = UniqueNameSet[item]
                                except:
                                    pass
                        if column == OutputCase_column:

                            try:
                                rows_selective = rows_selective + OutputCaseSet[item]
                            except:
                                try:
                                    rows_selective = OutputCaseSet[item]
                                except:
                                    pass

                    if column == columns[0]:
                        rows = rows_selective
                    else:
                        rows = intersection(rows, rows_selective)
                    rows.sort()

                else:
                    print("Under General")
                    if max_row > 0:
                        for row in range(start_row, max_row + 1):
                                data = values[ColTitle_Index[Col_ColTitle[column]]][1]

                                for item in data:
                                    if item is not None:
                                        try:
                                            item = float(item)
                                        except:
                                            try:
                                                item = int(item)
                                            except:
                                                pass

                                    if item == ws.cell(row=row, column=column).value:
                                        rows_selective.append(row)

                                if OCase_UniqueNameSet_Use and (column == UniqueName_column or column == OutputCase_column):
                                        try:
                                            item = float(ws.cell(row=row, column=column).value)
                                        except:
                                            try:
                                                item = int(ws.cell(row=row, column=column).value)
                                            except:
                                                item = ws.cell(row=row, column=column).value


                                        if column == OutputCase_column:
                                            try:
                                                OutputCaseSet[item].append(row)
                                            except:
                                                OutputCaseSet[item] = [row]

                                        if column == UniqueName_column:
                                            try:
                                                UniqueNameSet[item].append(row)
                                            except:
                                                UniqueNameSet[item] = [row]

                        if column == columns[0]:
                            rows = rows_selective
                        else:
                            rows = intersection(rows, rows_selective)


        if checker and OCase_UniqueNameSet_Use:
            if len(OutputCaseSet) !=0 and len(UniqueNameSet) != 0 :
                values1 = [OutputCaseSet, UniqueNameSet, prelist]
                # values1 = [OutputCaseSet, UniqueNameSet]


                data_rowset[key] = values1



        for i in range(1, 6):
            rows.insert(i-1, i)

        start_column = output_column
        block_start, block_end, block_height = data_copier(rows, columns_refined, heading)
        end_column = output_column


        y_columns, x_column, series_column = data_arranger(block_start, block_end, block_height)
        # tabled_charts_creater(block_start, block_end, block_height, y_columns, x_column, series_column)


        wb1.save(filename=refined_file)
    indicator += 1

def control_center_Table_Results():
    tables_sorter()







def Customized_Arranger(sheet_names):
    CA_InputFileName = "Disp Dirfts  + RS Cric Ref Y.xlsx"
    CA_InputSheetName = sheet_names                       #__#
    CA_OutputFileName = CA_InputFileName
    CA_OutputSheetName = CA_InputSheetName + " Max"                       #__#

    #All Are Absoulute Values of Excel Index
    CA_TitleRow = 5
    CA_BlockStartCol = 2
    CA_BlockEndCol = 9
    CA_BlockStartRow = 6
    CA_BlockEndRow = 9

    CA_BlockSpacingRows = 2
    CA_BlockSpacingCols = 1

    CA_BlocksNoVertically = 6
    CA_BlocksNoHorizontally = 8

    #For Charts
    CA_XValuesCol = 9            #__#
    CA_YValuesCol = 3
    CA_ChartTitleCol = 4
    CA_SeriesNameCol = 2
    CA_ChartTitle = ""


    CA_SortBy = "Max"
    CA_ReferenceCol = CA_XValuesCol
    CA_ConstantSeriesCol = 2
    CA_BlockSearchHorizontally = True

    #If values are manually provided above then the succedding if loops can be erased
    if CA_InputSheetName[0:4] == "Disp":
        CA_BlockStartCol = 2
        CA_BlockEndCol = 11
        CA_ChartTitle = "Displacement due to "

        if CA_InputSheetName[-1] == "X":
            CA_XValuesCol = 11  # __#
            CA_ReferenceCol = 11

        else:
            CA_XValuesCol = 11  # __#
            CA_ReferenceCol = 11
        print(sheet_names, CA_BlockEndCol, CA_XValuesCol)



    if CA_InputSheetName[0:5] == "Drift":
        CA_BlockStartCol = 2
        CA_BlockEndCol = 9
        CA_ChartTitle = "Drift due to "


        if CA_InputSheetName[-1] == "X":
            CA_XValuesCol = 9  # __#
            CA_ReferenceCol = 9

        else:
            CA_XValuesCol = 8  # __#
            CA_ReferenceCol = 8
        print(sheet_names, CA_BlockEndCol, CA_XValuesCol)




















    #Computational Requirements
    global CA_WritingRow
    CA_OriginalWritingRow = 5
    CA_WritingRow = CA_OriginalWritingRow

    global CA_ChartRowsGroup
    CA_RowColumn = {}
    CA_ChartRowsGroup = []


    CA_WBi = load_workbook(CA_InputFileName)
    CA_WSi = CA_WBi[CA_InputSheetName]

    try:
        CA_WBo = load_workbook(CA_OutputFileName)
    except:
        CA_WBo = Workbook()
    CA_WSo = CA_WBo.create_sheet("Sheet_1")
    CA_WSo.title = CA_OutputSheetName



    # Titles Writer
    for col in range(CA_BlockStartCol, CA_BlockEndCol + 1):
        CA_WSo.cell(row=CA_WritingRow, column=col).value = CA_WSi.cell(row=5, column=col).value
        CA_WSo.cell(row=CA_WritingRow, column=col).fill = PatternFill(start_color='5cb800', end_color='5cb800',
                                                                 fill_type="solid")
    CA_WritingRow += 1

    def Content_Writer(CA_RowColumn):
        def CA_StoryTextToHeight(CA_CellValue):
            if CA_CellValue == "Base":
                return 0
            if CA_CellValue == "Story1":
                return 3
            if CA_CellValue == "Story2":
                return 6
            if CA_CellValue == "Story3":
                return 9
            if CA_CellValue == "Story4":
                return 12
            if CA_CellValue == "Story5":
                return 15
            if CA_CellValue == "Story6":
                return 18
            if CA_CellValue == "Story7":
                return 21
            if CA_CellValue == "Story8":
                return 24

        global CA_WritingRow
        for CA_Key, CA_Value in CA_RowColumn.items():
            CA_ChartRows.append(CA_WritingRow)

            for column in CA_Value:
                CA_WritingColumn = CA_BlockStartCol + CA_Value.index(column)
                CA_CellValue = CA_WSi.cell(row=CA_Key, column=column).value
                try:
                    if CA_CellValue[0:5] == "Story" or CA_CellValue[0:4] == "Base":
                        CA_CellValue = CA_StoryTextToHeight(CA_CellValue)
                except: pass

                CA_WSo.cell(row=CA_WritingRow, column=CA_WritingColumn).value = CA_CellValue
        CA_WritingRow += 1

    def CA_Charts():
        CA_ChartTitleRows = CA_ChartRowsGroup[0]
        chart = openpyxl.chart.ScatterChart()
        chart.style = 4
        chart.x_axis.title =CA_WSo.cell(row=CA_TitleRow, column=CA_XValuesCol).value
        chart.y_axis.title = CA_WSo.cell(row=CA_TitleRow, column=CA_YValuesCol).value


        xvalues = None
        values = None
        series = None
        chart.title = CA_ChartTitle + str(CA_WSo.cell(row=CA_ChartTitleRows[0] + 1, column=CA_ChartTitleCol).value)



        for CA_XYChartRows in CA_ChartRowsGroup:
            xvalues = openpyxl.chart.Reference(CA_WSo, min_col=CA_XValuesCol, min_row=CA_XYChartRows[0],
                                               max_row=CA_XYChartRows[-1])

            values = openpyxl.chart.Reference(CA_WSo, min_col=CA_YValuesCol, min_row=CA_XYChartRows[0],  max_row=CA_XYChartRows[-1])
            series = openpyxl.chart.Series(values, xvalues, title=CA_WSo.cell(row=CA_XYChartRows[0] + 1, column=CA_SeriesNameCol).value)

            chart.series.append(series)

            series = chart.series[-1]
            Marker = CA_Markers[CA_ChartRowsGroup.index(CA_XYChartRows)]
            series.marker.symbol = Marker
            series.marker.size = 7

            if Marker == "star" or Marker == "plus" or Marker == 'x':
                series.marker.graphicalProperties.line.solidFill = CA_SeriesColour[
                    CA_ChartRowsGroup.index(CA_XYChartRows)]
                series.marker.graphicalProperties.line.width = pixels_to_EMU(2.5)
            else:
                series.marker.graphicalProperties.solidFill = CA_SeriesColour[CA_ChartRowsGroup.index(CA_XYChartRows)]
                series.marker.graphicalProperties.line.noFill = True

            series.graphicalProperties.line.solidFill = CA_SeriesColour[CA_ChartRowsGroup.index(CA_XYChartRows)]
            series.graphicalProperties.line.dashStyle = CA_Dashes[0]
            series.graphicalProperties.line.width = pixels_to_EMU(2.5)



            chart.legend.position = 'b'


        pp = ParagraphProperties(defRPr=CharacterProperties(latin=Font(typeface='Times New Roman'), sz=900, b=False))
        rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=CharacterProperties(latin=Font(typeface='Times New Roman'), sz=900, b=False))])

        chart.x_axis.txPr = rtp
        chart.y_axis.txPr = rtp

        pp = ParagraphProperties(defRPr=CharacterProperties(latin=Font(typeface='Times New Roman'), sz=1000, b=True))
        rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=CharacterProperties(latin=Font(typeface='Times New Roman'), sz=1000, b=True))])

        chart.legend.txPr = rtp
        chart.title.tx.rich.p[0].pPr = pp
        chart.x_axis.title.tx.rich.p[0].pPr = pp
        chart.y_axis.title.tx.rich.p[0].pPr = pp



        # Adjust graph size
        chart.width = 9     #Elselvier page halfwidth = 9cm        Writing area Only
        chart.height = 7   #Elselvier page full height = 24 cm        Writing area Only

        anchor = f'{get_column_letter(CA_BlockEndCol + CA_BlockSpacingCols * 2)}{CA_OriginalWritingRow}'
        CA_WSo.add_chart(chart, anchor)

    def CA_Max():
        CA_Rows = []
        CA_Cols = []
        CA_Row = 0
        CA_Col = 0

        if CA_BlockSearchHorizontally:
            for i in range(0, CA_BlocksNoVertically):
                CA_ConstantSeriesValue = 0
                Rows_Data = 0
                global CA_ChartRows
                CA_ChartRows = []
                for g in range(0, CA_BlocksNoHorizontally ):
                    CA_Column =  CA_ReferenceCol + g * (CA_BlockEndCol + 1 - CA_BlockStartCol + CA_BlockSpacingCols)
                    CA_MaxValue = 0
                    CA_RowColumn = {}
                    for j in range(CA_BlockStartRow, CA_BlockEndRow + 1):

                        CA_Row = i * (CA_BlockEndRow + 1 - CA_BlockStartRow + CA_BlockSpacingRows) + j
                        print(CA_WSi.cell(row=CA_Row, column=CA_Column).value)

                        CA_CellValue = CA_WSi.cell(row=CA_Row,  column=CA_Column).value
                        if CA_CellValue is None:
                            CA_CellValue = 0

                        if CA_CellValue is not None and CA_MaxValue <= CA_CellValue:
                            CA_MaxValue = CA_CellValue
                            Rows_Data = CA_Row



                    Columns_Data = []
                    if Rows_Data != 0:
                        for k in range(CA_BlockStartCol, CA_BlockEndCol + 1):
                            Columns_Data.append(k + g * (CA_BlockEndCol + 1 - CA_BlockStartCol + CA_BlockSpacingCols))
                        CA_RowColumn[Rows_Data] = Columns_Data
                        Content_Writer(CA_RowColumn)
                CA_ChartRowsGroup.append(CA_ChartRows)
            CA_Charts()


        else:
            for g in range(0, CA_BlocksNoHorizontally):
                CA_ChartRows = []
                CA_Column = CA_ReferenceCol + g * (CA_BlockEndCol + 1 - CA_BlockStartCol + CA_BlockSpacingCols)
                for i in range(0, CA_BlocksNoVertically):

                    CA_MaxValue = 0
                    Rows_Data = 0
                    CA_RowColumn = {}

                    for j in range(CA_BlockStartRow, CA_BlockEndRow + 1):
                        CA_Row = i * (CA_BlockEndRow + 1 - CA_BlockStartRow + CA_BlockSpacingRows) + j
                        print("Row", j, CA_Row, CA_Column)
                        print(CA_WSi.cell(row=CA_Row, column=CA_Column).value)

                        CA_CellValue = CA_WSi.cell(row=CA_Row,  column=CA_Column).value
                        if CA_CellValue is None:
                            CA_CellValue = 0

                        if CA_CellValue is not None and CA_MaxValue <= CA_CellValue:
                            CA_MaxValue = CA_CellValue
                            Rows_Data = CA_Row

                    Columns_Data = []
                    if Rows_Data != 0:
                        for k in range(CA_BlockStartCol, CA_BlockEndCol + 1):
                            Columns_Data.append(k + g * (CA_BlockEndCol + 1 - CA_BlockStartCol + CA_BlockSpacingCols))
                        CA_RowColumn[Rows_Data] = Columns_Data
                        Content_Writer(CA_RowColumn)
                        print(CA_RowColumn)
                CA_ChartRowsGroup.append(CA_ChartRows)

            CA_Charts()

    def CA_Min():
        CA_Rows = []
        CA_Cols = []
        CA_Row = 0
        CA_Col = 0

        if CA_BlockSearchHorizontally:
            for i in range(0, CA_BlocksNoVertically):
                CA_ConstantSeriesValue = 0
                Rows_Data = 0
                global CA_ChartRows
                CA_ChartRows = []
                for g in range(0, CA_BlocksNoHorizontally ):
                    CA_Column =  CA_ReferenceCol + g * (CA_BlockEndCol + 1 - CA_BlockStartCol + CA_BlockSpacingCols)
                    CA_MaxValue = 0
                    CA_MinValue = 0
                    CA_RowColumn = {}
                    for j in range(CA_BlockStartRow, CA_BlockEndRow + 1):

                        CA_Row = i * (CA_BlockEndRow + 1 - CA_BlockStartRow + CA_BlockSpacingRows) + j
                        CA_CellValue = CA_WSi.cell(row=CA_Row,  column=CA_Column).value

                        if CA_CellValue is None:
                            CA_CellValue = 0

                        if CA_CellValue is not None and CA_MaxValue <= CA_CellValue:
                            CA_MaxValue = CA_CellValue

                    CA_MinValue = CA_MaxValue

                    for j in range(CA_BlockStartRow, CA_BlockEndRow + 1):
                        CA_Row = i * (CA_BlockEndRow + 1 - CA_BlockStartRow + CA_BlockSpacingRows) + j

                        CA_CellValue = CA_WSi.cell(row=CA_Row,  column=CA_Column).value
                        if CA_CellValue is None:
                            CA_CellValue = 0

                        if CA_CellValue is not None and CA_MinValue >= CA_CellValue:
                            CA_MinValue = CA_CellValue
                            Rows_Data = CA_Row

                    Columns_Data = []
                    if Rows_Data != 0:
                        for k in range(CA_BlockStartCol, CA_BlockEndCol + 1):
                            Columns_Data.append(k + g * (CA_BlockEndCol + 1 - CA_BlockStartCol + CA_BlockSpacingCols))
                        CA_RowColumn[Rows_Data] = Columns_Data
                        Content_Writer(CA_RowColumn)
                CA_ChartRowsGroup.append(CA_ChartRows)
                CA_Charts()


        else:
            for g in range(0, CA_BlocksNoHorizontally):
                CA_ChartRows = []
                CA_Column = CA_ReferenceCol + g * (CA_BlockEndCol + 1 - CA_BlockStartCol + CA_BlockSpacingCols)
                for i in range(0, CA_BlocksNoVertically):
                    CA_MaxValue = 0
                    CA_MinValue = 0
                    Rows_Data = 0
                    CA_RowColumn = {}

                    for j in range(CA_BlockStartRow, CA_BlockEndRow + 1):
                        CA_Row = i * (CA_BlockEndRow + 1 - CA_BlockStartRow + CA_BlockSpacingRows) + j

                        CA_CellValue = CA_WSi.cell(row=CA_Row, column=CA_Column).value
                        if CA_CellValue is None:
                            CA_CellValue = 0

                        if CA_CellValue is not None and CA_MaxValue <= CA_CellValue:
                            CA_MaxValue = CA_CellValue

                    CA_MinValue = CA_MaxValue

                    for j in range(CA_BlockStartRow, CA_BlockEndRow + 1):
                        CA_Row = i * (CA_BlockEndRow + 1 - CA_BlockStartRow + CA_BlockSpacingRows) + j

                        CA_CellValue = CA_WSi.cell(row=CA_Row,  column=CA_Column).value
                        if CA_CellValue is None:
                            CA_CellValue = 0

                        if CA_CellValue is not None and CA_MinValue >= CA_CellValue:
                            CA_MinValue = CA_CellValue
                            Rows_Data = CA_Row

                    Columns_Data = []
                    if Rows_Data != 0:
                        for k in range(CA_BlockStartCol, CA_BlockEndCol + 1):
                            Columns_Data.append(k + g * (CA_BlockEndCol + 1 - CA_BlockStartCol + CA_BlockSpacingCols))
                        CA_RowColumn[Rows_Data] = Columns_Data
                        Content_Writer(CA_RowColumn)
                CA_Charts()


    if CA_SortBy == "Max":
        CA_Max()
    else:
        CA_Min()

    CA_WBo.save(CA_OutputFileName)
    CA_WBo.close()




