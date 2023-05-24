from User_Input import *
from openpyxl import *
import os
import shutil


#Index Identifier Variables
global block_end_row_i
# global load_case_col
max_number_results = 0
min_number_results = 0
start_column_i = 0
end_column_i = 0
block_end_row_i = []
block_start_row_i = []
start_row_i = 0
end_row_i = 0
max_num_res_row = 0
min_num_res_row = 0
numberresults_row = []
number_results = []
title_not_found = True
angle_deviation_col = 0
load_case_col = 0
block_height = 0
main_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop', "Python Programming")


def bool_converter(input):
    if (input == "True" or input == "TRUE") or input == "true":
        return (True)
    else:
        return (False)
#****************************************************"""Data Extraction For the Initialization of Code"""
os.chdir(main_path)
wb = load_workbook("Design Sheet.xlsx",data_only=True) #If data only is false it shows the formula of the cell
ws = wb["Modelling"]
giRow = int(ws.cell(row=3, column=12).value)
gdRow = int(ws.cell(row=4, column=12).value)
mdRow = int(ws.cell(row=5, column=12).value)
fsdRow =int( ws.cell(row=6, column=12).value)
lpRow = int(ws.cell(row=7, column=12).value)
lcRow = int(ws.cell(row=8, column=12).value)
ilcRow =int( ws.cell(row=9, column=12).value)
nlcRow =int( ws.cell(row=10, column=12).value)
fjRow = int(ws.cell(row=11, column=12).value)
laRow = int(ws.cell(row=12, column=12).value)
cdRow = int(ws.cell(row=13, column=12).value)
bdRow = int(ws.cell(row=14, column=12).value)
sdRow = int(ws.cell(row=15, column=12).value)
csaRow =int( ws.cell(row=16, column=12).value)

instance = bool_converter(ws.cell(row=3, column=10).value)
lockmodel = bool_converter(ws.cell(row=5, column=10).value)
defining_unit = ws.cell(row=gdRow + 6, column=5).value
modelling_unit = ws.cell(row=gdRow + 7, column=5).value
starting_angle = ws.cell(row=gdRow + 5, column=6).value
increment_angle = ws.cell(row=gdRow + 5, column=8).value
ending_angle = ws.cell(row=gdRow + 5, column=10).value
user_name = ws.cell(row=giRow, column=4).value
chart_position = bool_converter(ws.cell(row=7, column=10).value)
number_of_stories = ws.cell(row=giRow + 10, column=4).value



# #---------------------Creating Directory and exporting Design Excel File---------------------------------
main_folder = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Documents', 'ETABS Design Report')
ResearchFolder = os.path.join(main_folder,user_name, "ETABS", "Research Folder")
excelscr = os.path.join(main_path, 'Design Sheet.xlsx')
exceldst = main_folder + "\\" + user_name
etabspath = os.path.join(exceldst, "ETABS")
if os.path.exists(ResearchFolder) is False:
    os.makedirs(ResearchFolder)

shutil.copy(excelscr, exceldst)             #Never Open excel sheet while running else this code doesnot run
ModelPath = etabspath + os.sep + 'Design File.edb'
ExcelPath = exceldst + os.sep + "Design Sheet.xlsx"
FilePath = os.path.join(main_path, "ETABS File", "Model - Seismic.EDB")
os.chdir(exceldst)



#Identifies title column, column range, first block starting and ending row, from provided title
def indexer(wb, sheet_name, title_name):
    ws = wb[sheet_name]
    global start_column_i
    global end_column_i
    global block_end_row_i
    global block_start_row_i
    global start_row_i
    global end_row_i
    global max_number_results
    global min_number_results
    global max_num_res_row
    global min_num_res_row
    global title_not_found
    global numberresults_row
    global number_results
    global angle_deviation_col
    global load_case_col
    global block_height
    # block_end_row_i = [] #Never ever define with = inside function to make it as global variable
    # block_start_row_i = []

    start_row_i = 6
    end_row_i = 0
    max_col = ws.max_column
    max_row = ws.max_row
    number_res_column = 0
    max_num_res_row = 0
    min_num_res_row = 0
    numberresults_row = []
    number_results = []
    angle_deviation_col = 0
    # load_case_col = 0
    # block_height = 0

    del block_start_row_i[0: len(block_start_row_i) + 1]
    del block_end_row_i[0: len(block_end_row_i) + 1]

    title_not_found = True
    # First Block Boundary determination
    a = 0
    for i in range(1, max_col + 1):
        if ws.cell(row=5, column=i).value == "NumberResults":
            number_res_column = i
        if ws.cell(row=5, column=i).value == "Angle Deviation":
            angle_deviation_col = i
        if ws.cell(row=5, column=i).value == "LoadCase" or ws.cell(row=5, column=i).value=="OutputCase":
            load_case_col = i

        if a == 1 and ws.cell(row=5, column=i).value == None:
            end_column_i = i
        elif ws.cell(row=5, column=i).value == None or ws.cell(row=5, column=i).value != title_name:
            a = 0
        else:
            start_column_i = i
            a = 1
            title_not_found = False

    if number_res_column == 0:
        number_res_column = angle_deviation_col

    for j in range(6, max_row + 2):
            check_point = 0
            for k in range(0, 6):
                if ws.cell(row=j + k, column=number_res_column).value == None:
                    if k == 1:  # If j == x, then number is recorded if there encountered x rows none gap
                        block_end_row_i.append(j)  # For eg if x = 2 then it notes for 3 none values rows
                    if k == 5:
                        check_point = 2  # It breaks the system of loop
                    if j == max_row:
                        block_end_row_i.append(j)
                else:
                    check_point = 0
                    break
            if check_point == 2:
                break
            try:
                end_row = block_end_row_i[0]
            except:
                pass

    # Finding largest columns for a title and its row number
    max_number_results = 0
    min_number_results = 1000
    for i in range(start_row_i, max_row):
        try:
            if int(ws.cell(row=i, column=number_res_column).value) > max_number_results:
                max_number_results = ws.cell(row=i, column=number_res_column).value
                max_num_res_row = i
            if (int(ws.cell(row=i, column=number_res_column).value) < min_number_results) and (
                    (int(ws.cell(row=i, column=number_res_column).value) != 0) and
                    (ws.cell(row=i, column=number_res_column).value != None)):
                min_number_results = int(ws.cell(row=i, column=number_res_column).value)
                min_num_res_row = i

            if title_not_found == False:
                if ((int(ws.cell(row=i, column=number_res_column).value) != 0) and
                        (ws.cell(row=i, column=number_res_column).value != None)):
                    if int(ws.cell(row=i, column=number_res_column).value) not in number_results:
                        number_results.append(int(ws.cell(row=i, column=number_res_column).value))
                        numberresults_row.append(i)


        except:
            pass
    if len(block_end_row_i) > 0 :
        end_row_i = block_end_row_i [0]
        block_height = end_row_i - start_row_i
    for items in block_end_row_i:
        block_start_row_i.append(items - block_height)
    # print(numberresults_row)
    # print(number_results)
    # print(block_height, block_start_row_i, block_end_row_i)
    # print(load_case_col)

print(f'Excel_Extractor_Preprocessing is running')
