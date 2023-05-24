import Pre_Processings
from User_Input import *
from Pre_Processings import *
import sys
import comtypes.client
import xlwings as xw
from openpyxl.styles import Font, Fill, Color, PatternFill  # Connect styles for text
from openpyxl.utils import get_column_letter, column_index_from_string
import math
from math import *
import numpy as np
from itertools import combinations
from scipy.spatial import Delaunay


##_____________________ User_Inputs and Containers for Labels Storage________________________________________
angles = []
X_CoOrdiatesFloor = []
Y_CoOrdiatesFloor = []
f_name = "Design Sheet.xlsx"
f_result = "Result.xlsx"
refined_file = "Refined Output.xlsx"
baysX = ws.cell(row=gdRow, column=4).value
baysY = ws.cell(row=gdRow + 1, column=4).value
storey = ws.cell(row=giRow + 10, column=4).value
storey_height = ws.cell(row=giRow + 12, column=7).value
col_section = ws.cell(row=fsdRow, column=3).value
beam_section = ws.cell(row=fsdRow + 5, column=3).value
slab_section = ws.cell(row=fsdRow + 10, column=3).value
# material = ws.cell(row=fsdRow+5, column=4).value
id_unique = {}  # All kept in managed order of storey as key -> (Columns , BeamsX, BeamsY) as value for storey and key for unique id
id_fname = {}  # All seperated with element Like beam columns as keys and their frame name (..+"to"+..) as list
id_uname = {}  # All seperated with element Like beam columns as keys and their unique name (".."+"..") as list
ids = {}  # All kept in random order
slab_uname = []  # All kept unique name in list form ... First value indicates storey number
slab_floorwise = [] #All first floor slabs are kept in first list, second in second and so on as [[First floor slabs], ...]
columns_floorwise = [] #All first floor columns are kept in first list, second in second and so on as [[First floor columns], ...]
joints_floorwise = [] #All base floor columns are kept in first list, first in second and so on as [[First floor columns], ...]
roofSlab_uname = []
floorSlab_uname = []
beams_outer = []
beams_inner = []
columns_inner = []
columns_outer = []
colsection_dict = {} #contains storey as key and sections assigned to that storey's column as value

output_column = 1 #Used for output refinery


alphabet = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]
number = ["1", "2", "3", "4", "5", "6", "7", "8", "9", "10"]
original_RS_SFx = Resp_Spectrum_SFx[0]
original_RS_SFy = Resp_Spectrum_SFy[0]

#Special_ Case Outputs
story_drifts = []
# # # --------------------------------------------Initializing Variables
SapModel = None
wb = Pre_Processings.wb
ws = Pre_Processings.ws
def system_adjuster():
    global angles
    if CoOrdinatesFrom_Excel:
        for k in range(starting_angle, ending_angle, increment_angle):
            angles.append(k)
    else:
        if len(deviation_angles) > 0 :
            angles = deviation_angles
    number_of_iterations = int(len(angles))

def model_initializer():
    global SapModel
    EtabsObject = None
    AttachToInstance = 1
    if AttachToInstance:
        try:
            myETABSObject = comtypes.client.GetActiveObject("CSI.ETABS.API.ETABSObject")
        except (OSError, comtypes.COMError):
            print("No running instance of the program found or failed to attach.")
            sys.exit(-1)
    SapModel = myETABSObject.SapModel
    ret = SapModel.File.OpenFile(FilePath)  # ******** Just Open Etabs And No need to create NEW Model for opening file
def material_definition():
    # Set present Units
    ret = SapModel.SetPresentUnits(defining_unit)
    # Concrete Material Definition
    row = mdRow - 3
    col = 4
    if ws.cell(row=row, column=col).value == "Auto Defining":
        a = ws.cell(row=mdRow, column=3).value
        b = int(ws.cell(row=mdRow, column=4).value)
        c = ws.cell(row=mdRow, column=5).value
        d = ws.cell(row=mdRow, column=6).value
        e = ws.cell(row=mdRow, column=7).value
        ret = SapModel.PropMaterial.AddMaterial(a, b, c, d, e)
    else:
        a = ws.cell(row=mdRow + 5, column=3).value
        b = int(ws.cell(row=mdRow + 5, column=4).value)
        c = int(ws.cell(row=mdRow + 5, column=5).value)
        d = bool_converter(ws.cell(row=mdRow + 5, column=6).value)
        e = int(ws.cell(row=mdRow + 5, column=7).value)
        f = int(ws.cell(row=mdRow + 5, column=8).value)
        g = int(ws.cell(row=mdRow + 5, column=9).value)
        h = int(ws.cell(row=mdRow + 5, column=10).value)
        i = int(ws.cell(row=mdRow + 5, column=11).value)
        j = int(ws.cell(row=mdRow + 5, column=12).value)
        k = int(ws.cell(row=mdRow + 5, column=13).value)
        l = int(ws.cell(row=mdRow + 5, column=14).value)
        ret = SapModel.PropMaterial.SetMaterial(a, b)
        ret = SapModel.PropMaterial.SetOConcrete_1(a, c, d, e, f, g, h, i, j, k, l)
    # Rebar Material Definition
    row = mdRow + 7
    col = 4
    if ws.cell(row=row, column=col).value == "Auto Defining":
        a = ws.cell(row=mdRow + 10, column=3).value
        b = int(ws.cell(row=mdRow + 10, column=4).value)
        c = ws.cell(row=mdRow + 10, column=5).value
        d = ws.cell(row=mdRow + 10, column=6).value
        e = ws.cell(row=mdRow + 10, column=7).value
        ret = SapModel.PropMaterial.AddMaterial(a, b, c, d, e)
    else:
        a = ws.cell(row=mdRow + 15, column=3).value
        b = int(ws.cell(row=mdRow + 15, column=4).value)
        c = int(ws.cell(row=mdRow + 15, column=5).value)
        d = int(ws.cell(row=mdRow + 15, column=6).value)
        e = int(ws.cell(row=mdRow + 15, column=7).value)
        f = int(ws.cell(row=mdRow + 15, column=8).value)
        g = int(ws.cell(row=mdRow + 15, column=9).value)
        h = int(ws.cell(row=mdRow + 15, column=10).value)
        i = int(ws.cell(row=mdRow + 15, column=11).value)
        j = int(ws.cell(row=mdRow + 15, column=12).value)
        k = bool_converter(ws.cell(row=mdRow + 15, column=13).value)
        l = int(ws.cell(row=mdRow + 15, column=14).value)
        ret = SapModel.PropMaterial.SetMaterial(a, b)
        ret = SapModel.PropMaterial.SetORebar(a, c, d, e, f, g, h, i, j, k, l)
def section_definition():
    # Column Section Definition
    for loop in range(0, int(ws.cell(row=cdRow - 3, column=7).value)):
        if ws.cell(row=cdRow - 3, column=5).value == "Rectangular":
            a = ws.cell(row=cdRow + (loop*3), column=3).value
            b = ws.cell(row=cdRow + (loop*3), column=4).value
            c = int(ws.cell(row=cdRow + (loop*3), column=5).value)
            d = int(ws.cell(row=cdRow + (loop*3), column=6).value)
            e = ws.cell(row=cdRow + (loop*3), column=7).value
            f = ws.cell(row=cdRow + (loop*3), column=8).value
            g = int(ws.cell(row=cdRow + (loop*3), column=9).value)
            h = int(ws.cell(row=cdRow + (loop*3), column=10).value)
            i = int(ws.cell(row=cdRow + (loop*3), column=11).value)
            j = int(ws.cell(row=cdRow + (loop*3), column=12).value)
            k = int(ws.cell(row=cdRow + (loop*3), column=13).value)
            l = int(ws.cell(row=cdRow + (loop*3), column=14).value)
            m = str(ws.cell(row=cdRow + (loop*3), column=15).value)
            n = str(ws.cell(row=cdRow + (loop*3), column=16).value)
            o = int(ws.cell(row=cdRow + (loop*3), column=17).value)
            p = int(ws.cell(row=cdRow + (loop*3), column=18).value)
            q = int(ws.cell(row=cdRow + (loop*3), column=19).value)
            r = bool_converter((ws.cell(row=cdRow + (loop*3), column=20).value))
            ret = SapModel.PropFrame.SetRectangle(a, b, c, d)
            ret = SapModel.PropFrame.SetRebarColumn(a, e, f, g, h, i, j, k, l, m, n, o, p, q, r)


    # Beam Section Definition
    for loop in range(0, ws.cell(row=bdRow - 3, column=7).value):
        if ws.cell(row=bdRow - 3, column=5).value == "Rectangular":
            a = ws.cell(row=bdRow + (loop*3), column=3).value
            b = ws.cell(row=bdRow + (loop*3), column=4).value
            c = int(ws.cell(row=bdRow + (loop*3), column=5).value)
            d = int(ws.cell(row=bdRow + (loop*3), column=6).value)
            e = ws.cell(row=bdRow + (loop*3), column=7).value
            f = ws.cell(row=bdRow + (loop*3), column=8).value
            g = int(ws.cell(row=bdRow + (loop*3), column=9).value)
            h = int(ws.cell(row=bdRow + (loop*3), column=10).value)
            i = int(ws.cell(row=bdRow + (loop*3), column=11).value)
            j = int(ws.cell(row=bdRow + (loop*3), column=12).value)
            k = int(ws.cell(row=bdRow + (loop*3), column=13).value)
            l = int(ws.cell(row=bdRow + (loop*3), column=14).value)
            ret = SapModel.PropFrame.SetRectangle(a, b, c, d)
            ret = SapModel.PropFrame.SetRebarBeam(a, e, f, g, h, i, j, k, l)

    # Slab Section Definition
    a = ws.cell(row=sdRow, column=3).value
    b = int(ws.cell(row=sdRow , column=4).value)
    c = int(ws.cell(row=sdRow , column=5).value)
    d = ws.cell(row=sdRow , column=6).value
    e = int(ws.cell(row=sdRow , column=7).value)
    ret = SapModel.PropArea.SetSlab(a, b, c, d, e)
def loadpattern_definition():
    for i in range(0, 14):
        a = ws.cell(row=lpRow + i, column=3).value
        b = int(ws.cell(row=lpRow + i, column=4).value)
        c = int(ws.cell(row=lpRow + i, column=5).value)
        ret = SapModel.LoadPatterns.Add(a, b, c, True)
def loadcase_definition(i = 0):
    # Also Captures Required RS cases for output(Recommended to enable this phase)
    #Response Spectrum Cases Definition for determining Critical angle of seismic incidence
    if Critical_Angle_Determination:
        global cad_increment
        if i == 0:
            try:
                cases_output.remove("Response Spectrum")
            except:
                pass
            lower_limit = max_Angle
            upper_limit = critical_angle_UpperLimit + 1
            cad_increment = 1
        else:
            if (max_Angle - 1) < 0:
                lower_limit = 0
            else:
                lower_limit = 100 * (max_Angle - 1)
            upper_limit = 100*(max_Angle + 1)

            for RS_cases in critical_cases:
                try:
                    cases_output.remove(RS_cases)
                except:
                    pass


        for j in range(lower_limit, upper_limit, cad_increment):
            if i != 0 :
                j = j/100
            name = f'{loads_applied[0]} {j}'
            ret = SapModel.LoadCases.ResponseSpectrum.SetCase(name)

            Name = name
            NumberLoads = loads_applied[1]
            LoadName = loads_applied[2]
            Func = loads_applied[3]
            SF = loads_applied[4]
            CSys = loads_applied[5]
            Ang = [j]
            ret = SapModel.LoadCases.ResponseSpectrum.SetLoads(Name, NumberLoads,
                                                               LoadName,
                                                               Func,
                                                               SF,
                                                               CSys,
                                                               Ang)
            ret = SapModel.LoadCases.ResponseSpectrum.SetEccentricity(Name, Diaphragm_Ecc)
            critical_cases.append(name)

    try:
        cases_output.remove("Response Spectrum")
        NumberNames = 0
        MyName = []
        CaseType = 4
        [NumberNames, MyName, ret] = SapModel.LoadCases.GetNameList(NumberNames, MyName, CaseType)
        for RS_cases in MyName:
            cases_output.append(RS_cases)
    except:
        pass


def ResponseSpectrum_Modifier(angle = 0):
    Name = Resp_Spectrumx[0]
    NumberLoads = Resp_Spectrumx[1]
    LoadName = Resp_Spectrumx[2]
    Func = Resp_Spectrumx[3]
    SF = Resp_Spectrum_SFx
    CSys = Resp_Spectrumx[5]
    Ang = Resp_Spectrumx[6]

    ret = SapModel.LoadCases.ResponseSpectrum.SetLoads(Name, NumberLoads,
                                                      LoadName,
                                                      Func,
                                                      SF,
                                                      CSys,
                                                      Ang)
    Name = Resp_Spectrumy[0]
    NumberLoads = Resp_Spectrumy[1]
    LoadName = Resp_Spectrumy[2]
    Func = Resp_Spectrumy[3]
    SF = Resp_Spectrum_SFy
    CSys = Resp_Spectrumy[5]
    Ang = Resp_Spectrumy[6]

    ret = SapModel.LoadCases.ResponseSpectrum.SetLoads(Name, NumberLoads,
                                                      LoadName,
                                                      Func,
                                                      SF,
                                                      CSys,
                                                      Ang)

    Name = Resp_SpectrumCric[0]
    NumberLoads = Resp_SpectrumCric[1]
    LoadName = Resp_SpectrumCric[2]
    Func = Resp_SpectrumCric[3]
    SF = Resp_Spectrum_SF_Cr
    CSys = Resp_SpectrumCric[5]
    Ang = [CriticalAngles_DA_RS[angle]]
    print(angle, Ang)

    ret = SapModel.LoadCases.ResponseSpectrum.SetLoads(Name, NumberLoads,
                                                      LoadName,
                                                      Func,
                                                      SF,
                                                      CSys,
                                                      Ang)
    for value in Diaphragm_eccentricity:
        Name = value[0]
        Eccentricity = value[1]
        ret = SapModel.LoadCases.ResponseSpectrum.SetEccentricity(Name, Eccentricity)
def loadcombo_definition():
    combos_input = []
    # General Combo
    for i in range(lcRow, lcRow + 20):
        if ws.cell(row=i, column=2).value != None:
            a = ws.cell(row=i, column=3).value
            b = ws.cell(row=i, column=4).value
            SapModel.RespCombo.Add(a, b)
            combos_input.append(a)
            for n in range(0, 10):
                c = ws.cell(row=i + n, column=5).value
                if (c == 0) or (c == 1):
                    d = ws.cell(row=i + n, column=6).value
                    e = ws.cell(row=i + n, column=7).value
                    SapModel.RespCombo.SetCaseList(a, c, d, e)
                else:
                    break

    # Combos_input for the Codes
    if ws.cell(row=gdRow + 8, column=4).value == "IS":
        for i in range(ilcRow, ilcRow + 57):
            if ws.cell(row=i, column=2).value != None:
                a = ws.cell(row=i, column=3).value
                b = ws.cell(row=i, column=4).value
                SapModel.RespCombo.Add(a, b)
                combos_input.append(a)
                for n in range(0, 10):
                    c = ws.cell(row=i + n, column=5).value
                    if (c == 0) or (c == 1):
                        pass
                    else:
                        break
                for j in range(0, n):
                    c = ws.cell(row=i + j, column=5).value
                    d = ws.cell(row=i + j, column=6).value
                    e = ws.cell(row=i + j, column=7).value
                    SapModel.RespCombo.SetCaseList(a, c, d, e)
    else:
        for i in range(nlcRow, nlcRow + 22):
            if ws.cell(row=i, column=2).value != None:
                a = ws.cell(row=i, column=3).value
                b = ws.cell(row=i, column=4).value
                SapModel.RespCombo.Add(a, b)
                combos_input.append(a)
                for n in range(0, 10):
                    c = ws.cell(row=i + n, column=5).value
                    if (c == 0) or (c == 1):
                        pass
                    else:
                        break
                for j in range(0, n):
                    c = ws.cell(row=i + j, column=5).value
                    d = ws.cell(row=i + j, column=6).value
                    e = ws.cell(row=i + j, column=7).value
                    SapModel.RespCombo.SetCaseList(a, c, d, e)
def modelling_preparation():
    global angles
    global SapModel
    # # -----------------------------Saving Model -----------------------
    ret = SapModel.File.Save(ModelPath)
    wb.close()
    # Preparation for the iterations
    SapModel = None
    EtabsObject = None
def excelmodifier_modellinginitializor(k):
    global SapModel
    experiment_file = ResearchFolder + os.sep + f'Design File {k}.edb'

    # Modify Angle cell and save it by activating related formula
    wb = xw.Book(f_name)
    ws = wb.sheets["Modelling"]
    ws.range('F30').value = k
    wb.save()
    wb.close()

    # clean up variables and Restart the Sapmodel for the file opening
    SapModel = None
    EtabsObject = None
    AttachToInstance = 1
    if AttachToInstance:
        try:
            myETABSObject = comtypes.client.GetActiveObject("CSI.ETABS.API.ETABSObject")
        except (OSError, comtypes.COMError):
            print("No running instance of the program found or failed to attach.")
            sys.exit(-1)
    SapModel = myETABSObject.SapModel
    ret = SapModel.File.OpenFile(ModelPath)
    ret = SapModel.File.Save(experiment_file)
    ret = SapModel.File.OpenFile(experiment_file)


##----------------------------------------
def Co_Ordinate_Computer(Angle):
    def Chord_Fit(lnner_length, Curve_Angle, Curve_Radius):
        l = lnner_length
        R = Curve_Radius
        Angle = Curve_Angle
        D = 2 * R * math.sin(math.radians(Angle / 2))
        Ns = math.ceil(D / l)
        Angle_interval = 5
        Magnification = 100
        print("Hi i am here")
        if Ns > 1:
            for i in range(5, 180, Angle_interval):
                Angles = []
                Difference = []
                ratio1 = D / l
                ratio2 = sin(radians(i / 2)) / sin(radians(i / (2 * Ns)))

                if ratio2 < ratio1:
                    for j in range((i - Angle_interval) * Magnification, i * Magnification, Angle_interval):
                        if j == 0:
                            j = 1
                        Angle_J = j / Magnification


                        ratio22 = sin(radians(Angle_J / 2)) / sin(radians(Angle_J / (2 * Ns)))
                        if ratio22 < ratio1:
                            for k in range(j * Magnification - (Angle_interval * Magnification), j * Magnification,
                                           Angle_interval):
                                if k == 0:
                                    k = 1
                                Angle_K = k / (Magnification * Magnification)
                                ratio22 = sin(radians(Angle_K / 2)) / sin(radians(Angle_K / (2 * Ns)))
                                Angles.append(Angle_K)
                                Difference.append(abs(ratio1 - ratio22))

                    Angle_Output = Angles[Difference.index((min(Difference)))]
                    Radius_Output = D / (2 * sin(radians(Angle_Output / 2)))
                    Sides_No = Ns
                    print(Angle_Output, Radius_Output, Sides_No)
                    return ([Angle_Output, Radius_Output, Sides_No])

        else:
            Angle_Output =  Angle
            Radius_Output = l
            Sides_No =      1
            print(Angle_Output, Radius_Output, Sides_No)
            return ([Angle_Output, Radius_Output, Sides_No])

    def Triangles_Finder(points):
        #Finds maximum number of triangles
        def get_triangles(points):
            tri = Delaunay(points)
            triangles = []
            for simplex in tri.simplices:
                triangle = [tuple(points[i]) for i in simplex]
                triangles.append(tuple(sorted(triangle)))
            return list(set(triangles))

        return(get_triangles(points))








    if CoOrdinates_Type == 1:
        def Computer(Angle):
            # Co Ordinates Modelling
            # Level1 = Floor wise Co ordinates, Level 2 = Along Horizontal bays co, ordinates  [[], [], []]
            # Similarly same for y co_ordinates
            wb = load_workbook("Design Sheet.xlsx", data_only=True)  # If data only is false it shows the formula of the cell
            ws = wb["Modelling"]
            Angle = Angle


            Angle_AddGrids = {}
            X_BaysSpacing = []
            Y_BaysSpacing = []
            Origin_X = 0
            Origin_Y = 0

            Inclined_Bay = ws.cell(column=4, row=gdRow + 5).value
            for i in range(4, 13):
                X_BaysSpacing.append(ws.cell(column=i, row=gdRow + 3).value)
                Y_BaysSpacing.append(ws.cell(column=i, row=gdRow + 4).value)
            for row in range(gdRow + 1, gdRow + 7):
                Angle_AddGrids[int(ws.cell(column=14, row=row).value)] = int(ws.cell(column=14 + 1, row=row).value)
            X_Previous = Origin_X

            for i in range(0, baysX + 1):
                X_Value = X_Previous
                Y_Previous = Origin_Y
                Y_CoOrdBay = []
                X_CoOrdBay = []

                if i < Inclined_Bay:
                    for j in range(0, baysY + 1):
                        X_CoOrdBay.append(X_Value)
                    X_CoOrdiatesFloor.append(X_CoOrdBay)

                    for j in range(0, baysY + 1):
                        Y_Value = Y_Previous
                        Y_CoOrdBay.append(Y_Value)
                        Y_Value = Y_Previous + Y_BaysSpacing[j]
                        Y_Previous = Y_Value
                    Y_CoOrdiatesFloor.append(Y_CoOrdBay)

                if i == Inclined_Bay:
                    for k in range(1, Angle_AddGrids[Angle] + 2):
                        Grid_Angle = k * Angle / (Angle_AddGrids[Angle] + 1)
                        Inner_GridX = []
                        Inner_GridY = []
                        for j1 in range(0, baysY + 1):
                            length = Y_CoOrdiatesFloor[Inclined_Bay - 1][j1] - Y_CoOrdiatesFloor[Inclined_Bay - 1][0]
                            xi = X_CoOrdiatesFloor[Inclined_Bay - 1][0] + length * math.sin(math.radians(Grid_Angle))
                            yi = Y_CoOrdiatesFloor[Inclined_Bay - 1][0] + length * math.cos(math.radians(Grid_Angle))
                            Inner_GridX.append(xi)
                            Inner_GridY.append(yi)
                        X_CoOrdiatesFloor.append(Inner_GridX)
                        Y_CoOrdiatesFloor.append(Inner_GridY)

                if i > Inclined_Bay + 1:
                    for j in range(0, baysY + 1):
                        xi = X_CoOrdiatesFloor[-1][j]
                        yi = Y_CoOrdiatesFloor[-1][j]

                        length = X_BaysSpacing[i - 1]

                        xi1 = xi + length * math.cos(math.radians(Angle))
                        yi1 = yi - length * math.sin(math.radians(Angle))

                        X_CoOrdBay.append(xi1)
                        Y_CoOrdBay.append(yi1)

                    X_CoOrdiatesFloor.append(X_CoOrdBay)
                    Y_CoOrdiatesFloor.append(Y_CoOrdBay)


                X_Value = X_BaysSpacing[i] + X_Previous
                X_Previous = X_Value

            print(math.sin(math.radians(30)))
            print(X_CoOrdiatesFloor)
            print(Y_CoOrdiatesFloor)
        def BeamColSlab_CodeModelling():
            for h in range(1, storey + 1):
                # Beams along X Modelling
                loop_AlongY = len(X_CoOrdiatesFloor[0])
                loop_AlongX = len(X_CoOrdiatesFloor)
                for i in range(0, loop_AlongY):
                    for j in range(0, loop_AlongX):
                        try:
                            x1 = X_CoOrdiatesFloor[j][i]
                            y1 = Y_CoOrdiatesFloor[j][i]
                            z1 = h * storey_height
                            x2 = X_CoOrdiatesFloor[j + 1][i]
                            y2 = Y_CoOrdiatesFloor[j + 1][i]
                            z2 = h * storey_height

                            frame_key = f'{h}{i + 1}{j + 1}to{h}{i + 2}{j + 2}'
                            unique_name = f'{h}{i + 1}{j + 1}{h}{i + 2}{j + 2}'
                            [frame_key, ret] = SapModel.FrameObj.AddByCoord(x1, y1, z1, x2, y2, z2, frame_key,
                                                                            beam_section,
                                                                            unique_name)
                        except:
                            pass

                # Beams along Y Modelling
                for i in range(0, loop_AlongX):
                    for j in range(0, loop_AlongY):
                        try:
                            x1 = X_CoOrdiatesFloor[i][j]
                            y1 = Y_CoOrdiatesFloor[i][j]
                            z1 = h * storey_height
                            x2 = X_CoOrdiatesFloor[i][j + 1]
                            y2 = Y_CoOrdiatesFloor[i][j + 1]
                            z2 = h * storey_height

                            frame_key = f'{h}{j + 1}{i + 1}to{h}{j + 2}{i + 2}'
                            unique_name = f'{h + 1}{j + 1}{i + 1}{h + 1}{j + 2}{i + 2}'
                            [frame_key, ret] = SapModel.FrameObj.AddByCoord(x1, y1, z1, x2, y2, z2, frame_key,
                                                                            beam_section,
                                                                            unique_name)
                        except:
                            pass

                # Columns along Z Modelling
                for i in range(0, loop_AlongY):
                    for j in range(0, loop_AlongX):
                        try:
                            x1 = X_CoOrdiatesFloor[j][i]
                            y1 = Y_CoOrdiatesFloor[j][i]
                            z1 = (h - 1) * storey_height
                            x2 = X_CoOrdiatesFloor[j][i]
                            y2 = Y_CoOrdiatesFloor[j][i]
                            z2 = h * storey_height

                            frame_key = f'{h}{i + 1}{j + 1}to{h + 1}{i + 1}{j + 1}'
                            unique_name = f'{h}{i + 1}{j + 1}{h + 1}{i + 1}{j + 1}'
                            [frame_key, ret] = SapModel.FrameObj.AddByCoord(x1, y1, z1, x2, y2, z2, frame_key,
                                                                            col_section,
                                                                            unique_name)
                        except:
                            pass

                # Slabs Modelling
                for i in range(0, loop_AlongY):
                    for j in range(0, loop_AlongX):
                        try:
                            x1 = X_CoOrdiatesFloor[j][i]
                            y1 = Y_CoOrdiatesFloor[j][i]
                            z1 = h * storey_height
                            x2 = X_CoOrdiatesFloor[j + 1][i]
                            y2 = Y_CoOrdiatesFloor[j + 1][i]
                            z2 = h * storey_height
                            x3 = X_CoOrdiatesFloor[j + 1][i + 1]
                            y3 = Y_CoOrdiatesFloor[j + 1][i + 1]
                            z3 = h * storey_height
                            x4 = X_CoOrdiatesFloor[j][i + 1]
                            y4 = Y_CoOrdiatesFloor[j][i + 1]
                            z4 = h * storey_height

                            x = [x1, x2, x3, x4]
                            y = [y1, y2, y3, y4]
                            z = [z1, z2, z3, z4]

                            unique_name = f'{h}{alphabet[j]}{number[i]})'
                            slab_fname = f'{h}{i + 1}{j + 1}{h}{i + 1}{j + 2}{h}{i + 2}{j + 2}{h}{i + 2}{j + 1}'
                            ret = SapModel.AreaObj.AddByCoord(4, x, y, z, slab_fname, slab_section, unique_name)

                            frame_key = f'{h + 1}{i + 1}{j + 1}to{h}{i + 2}{j + 2}'
                            unique_name = f'{h + 1}{i + 1}{j + 1}{h}{i + 2}{j + 2}'
                            [frame_key, ret] = SapModel.FrameObj.AddByCoord(x1, y1, z1, x2, y2, z2, frame_key,
                                                                            beam_section,
                                                                            unique_name)
                        except:
                            pass

        Computer(Angle)
        BeamColSlab_CodeModelling()

    if CoOrdinates_Type == 2:
        triangles = []

        blocks = []
        bent_CoOrd = []

        def Computer(Angle):

            Angle_AddGrids = {}
            X_BaysSpacing = []
            Y_BaysSpacing = []
            Origin_X = 0
            Origin_Y = 0

            Inclined_Bay = ws.cell(column=4, row=gdRow + 5).value
            for i in range(4, 13):
                X_BaysSpacing.append(ws.cell(column=i, row=gdRow + 3).value)
                Y_BaysSpacing.append(ws.cell(column=i, row=gdRow + 4).value)
            for row in range(gdRow + 1, gdRow + 7):
                Angle_AddGrids[int(ws.cell(column=14, row=row).value)] = int(ws.cell(column=14 + 1, row=row).value)
            X_Previous = Origin_X
            Y_Previous = Origin_Y

            block1 = []
            block2 = []
            block3 = []
            print("A angle", Angle)

            for h in range(0, baysX + 1):

                # Phase 1 - Before bent
                if h < Inclined_Bay:
                    Xs = []
                    Ys = []
                    for i in range(0, baysY + 1):
                        if h == 0:
                            Xs.append(X_Previous)
                        else:
                            Xs.append(X_Previous + X_BaysSpacing[h - 1])

                        if i == 0:
                            Ys.append(Y_Previous)
                        else:
                            Ys.append(Ys[i - 1] + Y_BaysSpacing[i - 1])
                    line = []
                    for i in range(0, len(Xs)):
                        xy = [Xs[i], Ys[i]]
                        line.append(xy)

                    block1.append(line)
                    X_Previous = Xs[0]
                    print("Before bent angle", Angle)


                if h == Inclined_Bay:
                    blocks.append(block1)
                    line = []

                # Phase 2 - Bent
                if h == Inclined_Bay:
                    # Along X beam in bent portion
                    last_line = block1[-1]
                    centre_Coordinates = last_line[0]
                    for k in last_line:
                        if last_line.index(k) == 0:
                            x1 = k[0]
                            y1 = k[1]
                            xy = [x1, y1]
                            line.append(xy)
                            block2.append(line)  # Line in horizontal dir
                            line = []

                        if last_line.index(k) != 0:
                            x1 = k[0]
                            y1 = k[1]
                            xy = [x1, y1]
                            line.append(xy)

                            X_spacing = X_BaysSpacing[h - 1]
                            R = y1
                            if R > 0:
                                Result = Chord_Fit(X_spacing, Angle, R)
                                print((X_spacing, Angle, R))

                                Angle_Output = Result[0]
                                Radius_Output = Result[1]
                                Sides_No = Result[2]

                                beta = 180 - (Angle / 2) - ((360 - Angle_Output) / 2)
                                x0 = centre_Coordinates[0] + Radius_Output * sin(radians(beta))
                                y0 = R - Radius_Output * cos(radians(beta))

                                alpha = Angle_Output / Sides_No
                                for k in range(1, Sides_No + 1):
                                    gamma = k * alpha - beta
                                    xi = x0 + Radius_Output * sin(radians(gamma))
                                    yi = y0 + Radius_Output * cos(radians(gamma))

                                    xy = [xi, yi]
                                    line.append(xy)
                                block2.append(line)  # Line in horizontal dir
                                line = []

                    blocks.append(block2)
                    line = []

                    # Along Y beam in bent portion
                    Y_line = []

                    last_line = blocks[-1][-1]
                    S_last = blocks[-1][-2]

                    a = floor((len(last_line) - 2) / 2)
                    for index in range(1, a+1):
                        start_y_line = []
                        end_y_line = []
                        for lines in block2:
                            if len(lines) > (index*2):
                                start_y_line.append(lines[index])
                                end_y_line.append(lines[-index-1])
                        Y_line.append(start_y_line)
                        Y_line.append(end_y_line)


                    if len(last_line)%2 == 1:
                        len_last = len(last_line)
                        len_Slast = len(S_last)

                        #Case 1
                        if len_last == (len_Slast + 1):
                            co_index11 = floor((len_Slast)/2)-1
                            co_index12 = floor((len_Slast)/2)
                            co_index2 = floor(len_last/2)

                            line.append(S_last[co_index11])
                            line.append(last_line[co_index2])
                            Y_line.append(line)
                            line = []

                            line.append(S_last[co_index12])
                            line.append(last_line[co_index2])
                            Y_line.append(line)
                            line = []




                        #Case 2
                        if Quadrilateral_Boundary is False:
                            if len_last == (len_Slast + 2):
                                co_index11 = floor((len_Slast) / 2)
                                co_index2 = floor(len_last / 2)

                                line.append(S_last[co_index11])
                                line.append(last_line[co_index2])
                                Y_line.append(line)
                                line = []


                    blocks.append(Y_line)




                # Phase 3 - After bent

                if h > Inclined_Bay:
                    if h == (Inclined_Bay + 1):
                        last_line = block1[-1]
                        for k in last_line:
                            x1 = k[0]
                            y1 = k[1]

                            xi = x1 + y1 * math.sin(math.radians(Angle))
                            yi = y1 * math.cos(math.radians(Angle))

                            # yi = y1 - y1 * math.cos(math.radians(Angle))

                            xy = [xi, yi]
                            line.append(xy)
                    block3.append(line)
                    line = []

                    last_line = block3[-1]
                    for k in last_line:
                        X_spacing = X_BaysSpacing[h - 1]
                        x1 = k[0]
                        y1 = k[1]

                        xi = x1 + X_spacing * math.cos(math.radians(Angle))
                        yi = y1 - X_spacing * math.sin(math.radians(Angle))

                        xy = [xi, yi]
                        line.append(xy)
                    block3.append(line)

                    print("After bent angle", Angle)

                if h == baysX:
                    blocks.append(block3)
                    line = []

        def Slab_for_bent():
            #Determination of all co_ordinates in bent
            block1 = blocks[1]
            for lines in block1:
                for Co_Ord in lines:
                    bent_CoOrd.append(Co_Ord)

            shapes = Triangles_Finder([tuple(x) for x in bent_CoOrd])
            for x in shapes:
                triangles.append(x)



        def BeamColSlab_CodeModelling():

            for h in range(1, storey + 1):
                unique_name = ""
                frame_key = ""
                # Columns Modelling for bent
                for point in bent_CoOrd:
                    x1 = point[0]
                    y1 = point[1]
                    z1 = (h - 1) * storey_height
                    x2 = point[0]
                    y2 = point[1]
                    z2 = h * storey_height
                    [frame_key, ret] = SapModel.FrameObj.AddByCoord(x1, y1, z1, x2, y2, z2, frame_key,
                                                                    col_section,
                                                                    unique_name)

                # Slabs Modelling for bent portion
                for triangle in triangles:
                    x1 = triangle[0][0]
                    y1 = triangle[0][1]
                    z1 = h * storey_height
                    x2 = triangle[1][0]
                    y2 = triangle[1][1]
                    z2 = h * storey_height
                    x3 = triangle[2][0]
                    y3 = triangle[2][1]
                    z3 = h * storey_height

                    x = [x1, x2, x3, x1]
                    y = [y1, y2, y3, y1]
                    z = [z1, z2, z3, z1]
                    slab_fname = " "

                    ret = SapModel.AreaObj.AddByCoord(4, x, y, z, slab_fname, slab_section, unique_name)

            for block in blocks:
                for h in range(1, storey + 1):
                    if (blocks.index(block) == 1 or blocks.index(block) == 2):                # Bent type of block - horizontal line
                        # ALong X (block1) and Y Beams (block2) in bent
                        for line in block:
                            for i in range(0, len(line) - 1):
                                co_ordinates1 = line[i]
                                co_ordinates2 = line[i + 1]

                                x1 = co_ordinates1[0]
                                y1 = co_ordinates1[1]
                                z1 = h * storey_height
                                x2 = co_ordinates2[0]
                                y2 = co_ordinates2[1]
                                z2 = h * storey_height

                                y1_key = alphabet[block.index(line) + 1]
                                y2_key = alphabet[block.index(line) + 1]

                                x1_key = alphabet[line.index(co_ordinates1) + 1]
                                x2_key = alphabet[line.index(co_ordinates2) + 1]

                                z1_key = alphabet[h]
                                z2_key = alphabet[h]

                                frame_key = f'{z1_key}{y1_key}{x1_key}to{z2_key}{y2_key}{x2_key}'
                                unique_name = f'{z1_key}{y1_key}{x1_key}{z2_key}{y2_key}{x2_key}'
                                [frame_key, ret] = SapModel.FrameObj.AddByCoord(x1, y1, z1, x2, y2, z2, frame_key,
                                                                                beam_section,
                                                                                unique_name)





                    if (blocks.index(block) == 0 or  blocks.index(block) == 3):
                        # ALong Y Beams
                        for line in block:
                            for i in range(0, len(line) - 1):
                                co_ordinates1 = line[i]
                                co_ordinates2 = line[i + 1]

                                x1 = co_ordinates1[0]
                                y1 = co_ordinates1[1]
                                z1 = h * storey_height
                                x2 = co_ordinates2[0]
                                y2 = co_ordinates2[1]
                                z2 = h * storey_height

                                if blocks.index(block) == 0:
                                    x1_key = block.index(line) +1
                                    x2_key = block.index(line) + 1

                                if blocks.index(block) == 2:
                                    x1_key = len(blocks[0]) + block.index(line) +1
                                    x2_key = len(blocks[0]) + block.index(line) + 1

                                y1_key = line.index(co_ordinates1) + 1
                                y2_key = line.index(co_ordinates2) + 1


                                z1_key = h + 1
                                z2_key = h + 1

                                frame_key = f'{z1_key}{y1_key}{x1_key}to{z2_key}{y2_key}{x2_key}'
                                unique_name = f'{z1_key}{y1_key}{x1_key}{z2_key}{y2_key}{x2_key}'
                                [frame_key, ret] = SapModel.FrameObj.AddByCoord(x1, y1, z1, x2, y2, z2, frame_key,
                                                                                beam_section,
                                                                                unique_name)


                        # ALong X Beams
                        for i in range(0, len(line)):
                            for j in range(0, len(block) - 1):
                                line1 = block[j]
                                line2 = block[j + 1]
                                co_ordinates1 = line1[i]
                                co_ordinates2 = line2[i]

                                x1 = co_ordinates1[0]
                                y1 = co_ordinates1[1]
                                z1 = h * storey_height
                                x2 = co_ordinates2[0]
                                y2 = co_ordinates2[1]
                                z2 = h * storey_height

                                if blocks.index(block) == 0:
                                    x1_key = block.index(line) + 1
                                    x2_key = block.index(line) + 1

                                if blocks.index(block) == 2:
                                    x1_key = len(blocks[0]) + block.index(line) + 1
                                    x2_key = len(blocks[0]) + block.index(line) + 1

                                y1_key = line1.index(co_ordinates1) + 1
                                y2_key = line2.index(co_ordinates2) + 1

                                z1_key = h + 1
                                z2_key = h + 1

                                frame_key = f'{z1_key}{y1_key}{x1_key}to{z2_key}{y2_key}{x2_key}'
                                unique_name = f'{z1_key}{y1_key}{x1_key}{z2_key}{y2_key}{x2_key}'
                                [frame_key, ret] = SapModel.FrameObj.AddByCoord(x1, y1, z1, x2, y2, z2, frame_key,
                                                                                beam_section,
                                                                                unique_name)


                        # ALong Z Columns
                        for line in block:
                            for i in range(0, len(line)):
                                co_ordinates1 = line[i]

                                x1 = co_ordinates1[0]
                                y1 = co_ordinates1[1]
                                z1 = (h-1) * storey_height
                                x2 = co_ordinates1[0]
                                y2 = co_ordinates1[1]
                                z2 = h * storey_height

                                if blocks.index(block) == 0:
                                    x1_key = block.index(line) + 1
                                    x2_key = block.index(line) + 1

                                if blocks.index(block) == 2:
                                    x1_key = len(blocks[0]) + block.index(line) + 1
                                    x2_key = len(blocks[0]) + block.index(line) + 1


                                y1_key = line.index(co_ordinates1) + 1
                                y2_key = line.index(co_ordinates1) + 1

                                z1_key = h
                                z2_key = h + 1

                                frame_key = f'{z1_key}{y1_key}{x1_key}to{z2_key}{y2_key}{x2_key}'
                                unique_name = f'{z1_key}{y1_key}{x1_key}{z2_key}{y2_key}{x2_key}'
                                [frame_key, ret] = SapModel.FrameObj.AddByCoord(x1, y1, z1, x2, y2, z2, frame_key,
                                                                                col_section,
                                                                                unique_name)


                        # Slabs Modelling
                        for i in range(0, len(block) - 1):
                            for j in range(0, len(line) - 1):
                                line1 = block[i]
                                line2 = block[i + 1]

                                co_ordinates1 = line1[j]
                                co_ordinates2 = line1[j+1]
                                co_ordinates3 = line2[j+1]
                                co_ordinates4 = line2[j]

                                x1 = co_ordinates1[0]
                                y1 = co_ordinates1[1]
                                z1 = h * storey_height
                                x2 = co_ordinates2[0]
                                y2 = co_ordinates2[1]
                                z2 = h * storey_height
                                x3 = co_ordinates3[0]
                                y3 = co_ordinates3[1]
                                z3 = h * storey_height
                                x4 = co_ordinates4[0]
                                y4 = co_ordinates4[1]
                                z4 = h * storey_height

                                x = [x1, x2, x3, x4]
                                y = [y1, y2, y3, y4]
                                z = [z1, z2, z3, z4]

                                unique_name = f'{h}{alphabet[i]}{number[j]}'
                                slab_fname = f'{h}{i + 1}{j + 1}{h}{i + 1}{j + 2}{h}{i + 2}{j + 2}{h}{i + 2}{j + 1}'
                                ret = SapModel.AreaObj.AddByCoord(4, x, y, z, slab_fname, slab_section, unique_name)



        Computer(Angle)
        Slab_for_bent()
        BeamColSlab_CodeModelling()

    if CoOrdinates_Type == 3:
        triangles = []

        blocks = []
        bent_CoOrd = []

        def Computer(Angle):

            Angle_AddGrids = {}
            X_BaysSpacing = []
            Y_BaysSpacing = []
            Origin_X = 0
            Origin_Y = 0

            Inclined_Bay = ws.cell(column=4, row=gdRow + 5).value
            for i in range(4, 13):
                X_BaysSpacing.append(ws.cell(column=i, row=gdRow + 3).value)
                Y_BaysSpacing.append(ws.cell(column=i, row=gdRow + 4).value)
            for row in range(gdRow + 1, gdRow + 7):
                Angle_AddGrids[int(ws.cell(column=14, row=row).value)] = int(ws.cell(column=14 + 1, row=row).value)
            X_Previous = Origin_X
            Y_Previous = Origin_Y

            block1 = []
            block2 = []
            block3 = []
            print("A angle", Angle)

            for h in range(0, baysX + 1):

                # Phase 1 - Before bent
                if h < Inclined_Bay:
                    Xs = []
                    Ys = []
                    for i in range(0, baysY + 1):
                        if h == 0:
                            Xs.append(X_Previous)
                        else:
                            Xs.append(X_Previous + X_BaysSpacing[h - 1])

                        if i == 0:
                            Ys.append(Y_Previous)
                        else:
                            Ys.append(Ys[i - 1] + Y_BaysSpacing[i - 1])
                    line = []
                    for i in range(0, len(Xs)):
                        xy = [Xs[i], Ys[i]]
                        line.append(xy)

                    block1.append(line)
                    X_Previous = Xs[0]
                    print("Before bent angle", Angle)

                if h == Inclined_Bay:
                    blocks.append(block1)
                    line = []

                # Phase 2 - Bent
                if h == Inclined_Bay:
                    # Along X beam in bent portion
                    last_line = block1[-1]
                    centre_Coordinates = last_line[0]
                    for k in last_line:
                        if last_line.index(k) == 0:
                            x1 = k[0]
                            y1 = k[1]
                            xy = [x1, y1]
                            line.append(xy)
                            block2.append(line)  # Line in horizontal dir
                            line = []

                        if last_line.index(k) != 0:
                            x1 = k[0]
                            y1 = k[1]
                            xy = [x1, y1]
                            line.append(xy)

                            X_spacing = X_BaysSpacing[h - 1]
                            R = y1

                            circumference = 2 * math.pi * R * Angle / 360

                            Sides_No = ceil(circumference / X_spacing)
                            del_Angle = Angle / Sides_No
                            if R > 0:

                                for k in range(1, Sides_No + 1):
                                    gamma = k * del_Angle
                                    xi = x1 + R * sin(radians(gamma))
                                    yi = R * cos(radians(gamma))

                                    xy = [xi, yi]
                                    line.append(xy)
                                block2.append(line)  # Line in horizontal dir
                                line = []

                    blocks.append(block2)
                    line = []

                    # Along Y beam in bent portion
                    Y_line = []

                    last_line = blocks[-1][-1]
                    S_last = blocks[-1][-2]

                    a = floor((len(last_line) - 2) / 2)
                    for index in range(1, a + 1):
                        start_y_line = []
                        end_y_line = []
                        for lines in block2:
                            if len(lines) > (index * 2):
                                start_y_line.append(lines[index])
                                end_y_line.append(lines[-index - 1])
                        Y_line.append(start_y_line)
                        Y_line.append(end_y_line)

                    if len(last_line) % 2 == 1:
                        len_last = len(last_line)
                        len_Slast = len(S_last)

                        # Case 1
                        if len_last == (len_Slast + 1):
                            co_index11 = floor((len_Slast) / 2) - 1
                            co_index12 = floor((len_Slast) / 2)
                            co_index2 = floor(len_last / 2)

                            line.append(S_last[co_index11])
                            line.append(last_line[co_index2])
                            Y_line.append(line)
                            line = []

                            line.append(S_last[co_index12])
                            line.append(last_line[co_index2])
                            Y_line.append(line)
                            line = []

                        # Case 2
                        if Quadrilateral_Boundary is False:
                            if len_last == (len_Slast + 2):
                                co_index11 = floor((len_Slast) / 2)
                                co_index2 = floor(len_last / 2)

                                line.append(S_last[co_index11])
                                line.append(last_line[co_index2])
                                Y_line.append(line)
                                line = []

                    blocks.append(Y_line)

                # Phase 3 - After bent

                if h > Inclined_Bay:
                    if h == (Inclined_Bay + 1):
                        last_line = block1[-1]
                        for k in last_line:
                            x1 = k[0]
                            y1 = k[1]

                            xi = x1 + y1 * math.sin(math.radians(Angle))
                            yi = y1 * math.cos(math.radians(Angle))

                            # yi = y1 - y1 * math.cos(math.radians(Angle))

                            xy = [xi, yi]
                            line.append(xy)
                    block3.append(line)
                    line = []

                    last_line = block3[-1]
                    for k in last_line:
                        X_spacing = X_BaysSpacing[h - 1]
                        x1 = k[0]
                        y1 = k[1]

                        xi = x1 + X_spacing * math.cos(math.radians(Angle))
                        yi = y1 - X_spacing * math.sin(math.radians(Angle))

                        xy = [xi, yi]
                        line.append(xy)
                    block3.append(line)

                    print("After bent angle", Angle)

                if h == baysX:
                    blocks.append(block3)
                    line = []

        def Slab_for_bent():
            # Determination of all co_ordinates in bent
            block1 = blocks[1]
            for lines in block1:
                for Co_Ord in lines:
                    bent_CoOrd.append(Co_Ord)

            shapes = Triangles_Finder([tuple(x) for x in bent_CoOrd])
            for x in shapes:
                triangles.append(x)

        def BeamColSlab_CodeModelling():
            # Slabs  and Columns Modelling for bent portion
            for h in range(1, storey + 1):
                unique_name = ""
                frame_key = ""
                # Columns Modelling for bent
                for point in bent_CoOrd:
                    x1 = point[0]
                    y1 = point[1]
                    z1 = (h - 1) * storey_height
                    x2 = point[0]
                    y2 = point[1]
                    z2 = h * storey_height
                    [frame_key, ret] = SapModel.FrameObj.AddByCoord(x1, y1, z1, x2, y2, z2, frame_key,
                                                                    col_section,
                                                                    unique_name)

                # Slabs Modelling for bent portion
                for triangle in triangles:
                    x1 = triangle[0][0]
                    y1 = triangle[0][1]
                    z1 = h * storey_height
                    x2 = triangle[1][0]
                    y2 = triangle[1][1]
                    z2 = h * storey_height
                    x3 = triangle[2][0]
                    y3 = triangle[2][1]
                    z3 = h * storey_height

                    x = [x1, x2, x3, x1]
                    y = [y1, y2, y3, y1]
                    z = [z1, z2, z3, z1]
                    slab_fname = " "

                    ret = SapModel.AreaObj.AddByCoord(4, x, y, z, slab_fname, slab_section, unique_name)

            for block in blocks:
                for h in range(1, storey + 1):
                    # Bent type of block - horizontal line
                    if (blocks.index(block) == 1 or blocks.index(block) == 2):
                        # ALong X (block1) and Y Beams (block2) in bent
                        for line in block:
                            for i in range(0, len(line) - 1):
                                co_ordinates1 = line[i]
                                co_ordinates2 = line[i + 1]

                                x1 = co_ordinates1[0]
                                y1 = co_ordinates1[1]
                                z1 = h * storey_height
                                x2 = co_ordinates2[0]
                                y2 = co_ordinates2[1]
                                z2 = h * storey_height

                                y1_key = alphabet[block.index(line) + 1]
                                y2_key = alphabet[block.index(line) + 1]

                                x1_key = alphabet[line.index(co_ordinates1) + 1]
                                x2_key = alphabet[line.index(co_ordinates2) + 1]

                                z1_key = alphabet[h]
                                z2_key = alphabet[h]

                                frame_key = f'{z1_key}{y1_key}{x1_key}to{z2_key}{y2_key}{x2_key}'
                                unique_name = f'{z1_key}{y1_key}{x1_key}{z2_key}{y2_key}{x2_key}'
                                [frame_key, ret] = SapModel.FrameObj.AddByCoord(x1, y1, z1, x2, y2, z2, frame_key,
                                                                                beam_section,
                                                                                unique_name)

                    # Regular type of block
                    if (blocks.index(block) == 0 or blocks.index(block) == 3):
                        # ALong Y Beams
                        for line in block:
                            for i in range(0, len(line) - 1):
                                co_ordinates1 = line[i]
                                co_ordinates2 = line[i + 1]

                                x1 = co_ordinates1[0]
                                y1 = co_ordinates1[1]
                                z1 = h * storey_height
                                x2 = co_ordinates2[0]
                                y2 = co_ordinates2[1]
                                z2 = h * storey_height

                                if blocks.index(block) == 0:
                                    x1_key = block.index(line) + 1
                                    x2_key = block.index(line) + 1

                                if blocks.index(block) == 3:
                                    x1_key = len(blocks[0]) + block.index(line) + 1
                                    x2_key = len(blocks[0]) + block.index(line) + 1

                                y1_key = line.index(co_ordinates1) + 1
                                y2_key = line.index(co_ordinates2) + 1

                                z1_key = h + 1
                                z2_key = h + 1

                                frame_key = f'{z1_key}{y1_key}{x1_key}to{z2_key}{y2_key}{x2_key}'
                                unique_name = f'{z1_key}{y1_key}{x1_key}{z2_key}{y2_key}{x2_key}'
                                [frame_key, ret] = SapModel.FrameObj.AddByCoord(x1, y1, z1, x2, y2, z2, frame_key,
                                                                                beam_section,
                                                                                unique_name)

                        # ALong X Beams
                        for i in range(0, len(line)):
                            for j in range(0, len(block) - 1):
                                line1 = block[j]
                                line2 = block[j + 1]
                                co_ordinates1 = line1[i]
                                co_ordinates2 = line2[i]

                                x1 = co_ordinates1[0]
                                y1 = co_ordinates1[1]
                                z1 = h * storey_height
                                x2 = co_ordinates2[0]
                                y2 = co_ordinates2[1]
                                z2 = h * storey_height

                                if blocks.index(block) == 0:
                                    x1_key = block.index(line) + 1
                                    x2_key = block.index(line) + 1

                                if blocks.index(block) == 3:
                                    x1_key = len(blocks[0]) + block.index(line) + 1
                                    x2_key = len(blocks[0]) + block.index(line) + 1

                                y1_key = line1.index(co_ordinates1) + 1
                                y2_key = line2.index(co_ordinates2) + 1

                                z1_key = h + 1
                                z2_key = h + 1

                                frame_key = f'{z1_key}{y1_key}{x1_key}to{z2_key}{y2_key}{x2_key}'
                                unique_name = f'{z1_key}{y1_key}{x1_key}{z2_key}{y2_key}{x2_key}'
                                [frame_key, ret] = SapModel.FrameObj.AddByCoord(x1, y1, z1, x2, y2, z2, frame_key,
                                                                                beam_section,
                                                                                unique_name)

                        # ALong Z Columns
                        for line in block:
                            for i in range(0, len(line)):
                                co_ordinates1 = line[i]

                                x1 = co_ordinates1[0]
                                y1 = co_ordinates1[1]
                                z1 = (h - 1) * storey_height
                                x2 = co_ordinates1[0]
                                y2 = co_ordinates1[1]
                                z2 = h * storey_height

                                if blocks.index(block) == 0:
                                    x1_key = block.index(line) + 1
                                    x2_key = block.index(line) + 1

                                if blocks.index(block) == 3:
                                    x1_key = len(blocks[0]) + block.index(line) + 1
                                    x2_key = len(blocks[0]) + block.index(line) + 1

                                y1_key = line.index(co_ordinates1) + 1
                                y2_key = line.index(co_ordinates1) + 1

                                z1_key = h
                                z2_key = h + 1

                                frame_key = f'{z1_key}{y1_key}{x1_key}to{z2_key}{y2_key}{x2_key}'
                                unique_name = f'{z1_key}{y1_key}{x1_key}{z2_key}{y2_key}{x2_key}'
                                [frame_key, ret] = SapModel.FrameObj.AddByCoord(x1, y1, z1, x2, y2, z2, frame_key,
                                                                                col_section,
                                                                                unique_name)

                        # Slabs Modelling
                        for i in range(0, len(block) - 1):
                            for j in range(0, len(line) - 1):
                                line1 = block[i]
                                line2 = block[i + 1]

                                co_ordinates1 = line1[j]
                                co_ordinates2 = line1[j + 1]
                                co_ordinates3 = line2[j + 1]
                                co_ordinates4 = line2[j]

                                x1 = co_ordinates1[0]
                                y1 = co_ordinates1[1]
                                z1 = h * storey_height
                                x2 = co_ordinates2[0]
                                y2 = co_ordinates2[1]
                                z2 = h * storey_height
                                x3 = co_ordinates3[0]
                                y3 = co_ordinates3[1]
                                z3 = h * storey_height
                                x4 = co_ordinates4[0]
                                y4 = co_ordinates4[1]
                                z4 = h * storey_height

                                x = [x1, x2, x3, x4]
                                y = [y1, y2, y3, y4]
                                z = [z1, z2, z3, z4]

                                unique_name = f'{h}{alphabet[i]}{number[j]}'
                                slab_fname = f'{h}{i + 1}{j + 1}{h}{i + 1}{j + 2}{h}{i + 2}{j + 2}{h}{i + 2}{j + 1}'
                                ret = SapModel.AreaObj.AddByCoord(4, x, y, z, slab_fname, slab_section, unique_name)

        Computer(Angle)
        Slab_for_bent()
        BeamColSlab_CodeModelling()




##----------------------------------------
def beamcolumnslab_modelling():
    for h in range(13, (int(storey)) * 13 + 1, 13):
        fjcRow = fjRow + h
        columns_unique = []
        columns_fname = []
        columns_uname = []
        beamsX_unique = []
        beamsX_fname = []
        beamsX_uname = []
        beamsY_unique = []
        beamsY_fname = []
        beamsY_uname = []
        slabs_floor = []
        columns_floor = []
        joints_floor = []

        # Joints Unique name assignment
        for i in range(0, baysY + 1):
            for j in range(0, baysX + 1):
                if len(joints_floorwise) == 0:
                    joints_floor.append(j * 2 + 1 + (i) * 2 * (baysX + 1))
                elif len(joints_floorwise) == 1:
                    joints_floor.append((j + 1) * 2 + (i) *  2 *(baysX + 1))
                else:
                    joints_floor.append((baysY + 1) * (baysX + 1) * ((h-13) / 13) + (j + 1) + (i) * (baysX + 1))
        joints_floorwise.append(joints_floor)


        # Columns Along Z Modelling
        for i in range(0, baysY + 1):
            for j in range(3, 4 * baysX + 4, 4):
                p1 = ws.cell(row=fjcRow - 13 + i, column=j).value
                p2 = ws.cell(row=fjcRow, column=j).value
                x1 = (ws.cell(row=fjcRow - 13 + i, column=j + 1).value)
                y1 = (ws.cell(row=fjcRow - 13 + i, column=j + 2).value)
                z1 = (ws.cell(row=fjcRow - 13 + i, column=j + 3).value)
                x2 = (ws.cell(row=fjcRow + i, column=j + 1).value)
                y2 = (ws.cell(row=fjcRow + i, column=j + 2).value)
                z2 = (ws.cell(row=fjcRow + i, column=j + 3).value)
                frame_key = str(p1) + "to" + str(p2)
                unique_namez = str(p1) + str(p2)  # Creates the unique name to the each of the frame element
                columns_unique.append(unique_namez)
                columns_fname.append(frame_key)
                columns_uname.append(unique_namez)
                ids[frame_key] = unique_namez
                [frame_key, ret] = SapModel.FrameObj.AddByCoord(x1, y1, z1, x2, y2, z2, frame_key, col_section,
                                                                unique_namez)
                if ((i == 0 or j == 3) or (i == baysY or j == 4 * baysX + 4 - 1)):
                    columns_outer.append(unique_namez)
                else:
                    columns_inner.append(unique_namez)
                columns_floor.append(unique_namez)

        # Beams Along X Modelling
        for i in range(0, baysY + 1):
            for j in range(3, 4 * baysX, 4):
                p1 = ws.cell(row=fjcRow + i, column=j).value
                p2 = ws.cell(row=fjcRow + i, column=j + 4).value
                x1 = (ws.cell(row=fjcRow + i, column=j + 1).value)
                y1 = (ws.cell(row=fjcRow + i, column=j + 2).value)
                z1 = (ws.cell(row=fjcRow + i, column=j + 3).value)
                x2 = (ws.cell(row=fjcRow + i, column=j + 5).value)
                y2 = (ws.cell(row=fjcRow + i, column=j + 6).value)
                z2 = (ws.cell(row=fjcRow + i, column=j + 7).value)
                frame_key = str(p1) + "to" + str(p2)
                unique_namex = str(p1) + str(p2)
                beamsX_unique.append(unique_namex)
                beamsX_fname.append(frame_key)
                beamsX_uname.append(unique_namex)
                ids[frame_key] = unique_namex
                [frame_key, ret] = SapModel.FrameObj.AddByCoord(x1, y1, z1, x2, y2, z2, frame_key, beam_section,
                                                                unique_namex)
                if (i == 0 or i == baysY):
                    beams_outer.append(unique_namex)
                else:
                    beams_inner.append(unique_namex)

        # Beams Along Y Modelling
        for j in range(3, 4 * baysX + 4, 4):
            for i in range(0, baysY):
                p1 = ws.cell(row=fjcRow + i, column=j).value
                p2 = ws.cell(row=fjcRow + i + 1, column=j).value
                x1 = ws.cell(row=fjcRow + i, column=j + 1).value
                y1 = ws.cell(row=fjcRow + i, column=j + 2).value
                z1 = ws.cell(row=fjcRow + i, column=j + 3).value
                x2 = ws.cell(row=fjcRow + i + 1, column=j + 1).value
                y2 = ws.cell(row=fjcRow + i + 1, column=j + 2).value
                z2 = ws.cell(row=fjcRow + i + 1, column=j + 3).value
                frame_key = str(p1) + "to" + str(p2)
                unique_namey = str(p1) + str(p2)
                beamsY_unique.append(unique_namey)
                beamsY_fname.append(frame_key)
                beamsY_uname.append(unique_namey)

                ids[frame_key] = unique_namey
                [frame_key, ret] = SapModel.FrameObj.AddByCoord(x1, y1, z1, x2, y2, z2, frame_key, beam_section,
                                                                unique_namey)
                if (j == 3 or j == 4 * baysX + 4 - 1):
                    beams_outer.append(unique_namey)
                else:
                    beams_inner.append(unique_namey)
        id_unique[f"{int(h / 13)} Columns"] = columns_unique
        id_unique[f"{int(h / 13)} BeamsY"] = beamsX_unique
        id_unique[f"{int(h / 13)} BeamsX"] = beamsY_unique
        id_fname[f"{int(h / 13)} Columns"] = columns_fname
        id_fname[f"{int(h / 13)} BeamsX"] = beamsX_fname
        id_fname[f"{int(h / 13)} BeamsY"] = beamsY_fname
        id_uname[f"{int(h / 13)} Columns"] = columns_uname
        id_uname[f"{int(h / 13)} BeamsX"] = beamsX_uname
        id_uname[f"{int(h / 13)} BeamsY"] = beamsY_uname

        #   #Slab Modelling
        for i in range(0, baysY):  # Along Y Modelling
            for j in range(3, 4 * baysX, 4):  # Along X Modelling
                p1 = (ws.cell(row=fjcRow + i, column=j).value)
                x1 = (ws.cell(row=fjcRow + i, column=j + 1).value)
                y1 = (ws.cell(row=fjcRow + i, column=j + 2).value)
                z1 = (ws.cell(row=fjcRow + i, column=j + 3).value)
                p4 = (ws.cell(row=fjcRow + i, column=j + 4).value)
                x4 = (ws.cell(row=fjcRow + i, column=j + 5).value)
                y4 = (ws.cell(row=fjcRow + i, column=j + 6).value)
                z4 = (ws.cell(row=fjcRow + i, column=j + 7).value)
                p2 = (ws.cell(row=fjcRow + i + 1, column=j).value)
                x2 = (ws.cell(row=fjcRow + i + 1, column=j + 1).value)
                y2 = (ws.cell(row=fjcRow + i + 1, column=j + 2).value)
                z2 = (ws.cell(row=fjcRow + i + 1, column=j + 3).value)
                p3 = (ws.cell(row=fjcRow + i + 1, column=j + 4).value)
                x3 = (ws.cell(row=fjcRow + i + 1, column=j + 5).value)
                y3 = (ws.cell(row=fjcRow + i + 1, column=j + 6).value)
                z3 = (ws.cell(row=fjcRow + i + 1, column=j + 7).value)
                x = [x1, x2, x3, x4]
                y = [y1, y2, y3, y4]
                z = [z1, z2, z3, z4]
                unique_name = str(str(int(h / 13)) + alphabet[int((j - 3) / 4)] + number[i])
                slab_fname = str(p1) + str(p2) + str(p3) + str(p4)
                slab_uname.append(unique_name)
                ret = SapModel.AreaObj.AddByCoord(4, x, y, z, slab_fname, slab_section, unique_name)
                if int(h / 13) != storey:
                    floorSlab_uname.append(unique_name)
                else:
                    roofSlab_uname.append(unique_name)
                slabs_floor.append(unique_name)
        slab_floorwise.append(slabs_floor)
        columns_floorwise.append(columns_floor)
def section_assigner():
    for loop in range(0, int(ws.cell(row=cdRow - 3, column=7).value)):
        for numbers in range(4,12):
            if ws.cell(row=csaRow + loop, column=numbers).value != None:
                colsection_dict[ws.cell(row=csaRow + loop, column=numbers).value] = ws.cell(row=csaRow + loop, column=3).value

    for floor in range(1, len(columns_floorwise) + 1):
        for columns in range(0, len(columns_floorwise[floor-1])):
            ret = SapModel.FrameObj.SetSection(columns_floorwise[floor-1][columns], colsection_dict.get(floor))
def restraint_assignment():
    fjcRow = fjRow + 13
    for i in range(0, baysY + 1):
        for j in range(3, 4 * baysX + 4, 4):
            point1 = " "
            point2 = " "
            restraint = [True, True, True, True, True, True]
            p1 = ws.cell(row=fjcRow - 13 + i, column=j).value
            p2 = ws.cell(row=fjcRow, column=j).value
            unique_namez = str(p1) + str(p2)
            [point1, point2, ret] = SapModel.FrameObj.GetPoints(unique_namez, point1, point2)
            ret = SapModel.PointObj.SetRestraint(point1, restraint)
def load_application():
    loads = {}
    for i in range(0, 11):
        loads[ws.cell(row=laRow + i, column=3).value] = ws.cell(row=laRow + i, column=4).value

    # Outer Wall Load
    for i in beams_outer:
        ret = SapModel.FrameObj.SetLoadDistributed(i, list(loads.keys())[0], 1, 2, 0, 1,
                                                   -(loads[list(loads.keys())[0]]),
                                                   -(loads[list(loads.keys())[0]]), "Local")
    #
    # #Inner Wall Load
    for i in beams_inner:
        #     ret = SapModel.FrameObj.SetLoadDistributed(list(ids.keys())[list(ids.values()).index(i)],list(loads.keys())[1], 1, 10, 0, 1,
        #                                                (loads[list(loads.keys())[1]]), (loads[list(loads.keys())[1]]))
        ret = SapModel.FrameObj.SetLoadDistributed(i, list(loads.keys())[1], 1, 2,
                                                   0, 1, -(loads[list(loads.keys())[1]]),
                                                   -(loads[list(loads.keys())[1]]), "Local")
    for i in slab_uname:
        Load_name = ws.cell(row=laRow + 3, column=3).value
        Load_value = ws.cell(row=laRow + 3, column=4).value
        ret = SapModel.AreaObj.SetLoadUniform(i, Load_name, (float(Load_value)), 3, True, "Local")
    for i in floorSlab_uname:
        Load_name = ws.cell(row=laRow + 5, column=3).value
        Load_value = ws.cell(row=laRow + 5, column=4).value
        ret = SapModel.AreaObj.SetLoadUniform(i, Load_name, (float(Load_value)), 3, True, "Local")
    for i in floorSlab_uname:
        Load_name = ws.cell(row=laRow + 6, column=3).value
        Load_value = ws.cell(row=laRow + 6, column=4).value
        ret = SapModel.AreaObj.SetLoadUniform(i, Load_name, (float(Load_value)), 3, True, "Local")
    for i in roofSlab_uname:
        Load_name = ws.cell(row=laRow + 7, column=3).value
        Load_value = ws.cell(row=laRow + 7, column=4).value
        ret = SapModel.AreaObj.SetLoadUniform(i, Load_name, (float(Load_value)), 3, True, "Local")

    # Closing the excel of design file as all data are succesfully entered for modelling
    number_of_stories = ws.cell(row=giRow + 10, column=4).value
    wb.close()
def diaphragm_assignment():
    for story in range(1, len(slab_floorwise) + 1):
        diaphragm_name = "Diaphragm " + str(story)
        ret = SapModel.Diaphragm.SetDiaphragm(diaphragm_name, SemiRigid := False)

        for slabs in range(0, len(slab_floorwise[story-1])):
             ret = SapModel.AreaObj.SetDiaphragm(slab_floorwise[story-1][slabs], diaphragm_name)
def run_analysis(k):
    ret = SapModel.Analyze.RunAnalysis()
    print(f"Required Analysis Completed for Angle {k}")


#Results Section__________________________________________________________________________________________________
def output_writer(sheet_name, titles, outputs, nth_iteration = 1, k = 0, special_case = 0 ):
    global f_result
    os.chdir(exceldst)
    try:
        wb = load_workbook(f_result)
    except:
        wb = Workbook()

    if special_case ==0:
        try:
            ws = wb[sheet_name]
        except:
            ws = wb.create_sheet("Sheet_1")
            ws.title = sheet_name
            ws.merge_cells('F2:M3')
            ws.cell(row=2, column=6).value = sheet_name
            ws_h2 = ws['F2']
            ws_h2.font = Font(size=23, underline='single', color='FFBB00', bold=True, italic=False)
        # Initialization
        start_column = 1
        row = ws.max_row
        column = start_column
        titles.insert(0, "Angle Deviation");
        titles.insert(0, "Model Number")
        for h in range(0, len(outputs)):
            outputs[h].insert(0, k);
            outputs[h].insert(0, nth_iteration)

        max_numbers = 1
        try:
            index = titles.index("NumberResults")
            for n in range(0, len(outputs)):
                if outputs[n][index] >= max_numbers:
                    max_numbers = outputs[n][index]
        except:
            max_numbers = len(outputs[0])


        # Titles Writer
        for i in range(0, len(outputs[0])):
            try:
                list_checker = len(outputs[0][i])
                for j in range(0, max_numbers):
                    if j == 0:
                        ws.cell(row=5, column=column).value = titles[i]
                        ws.cell(row=5, column=column).fill = PatternFill(start_color='32CD32',
                                                                                 end_color='32CD32', fill_type="solid")

                        column += 1
                    else:
                        ws.cell(row=5, column=column).fill = PatternFill(start_color='32CD32',
                                                                                 end_color='32CD32', fill_type="solid")
                        column += 1
            except:
                ws.cell(row=5, column=column).value = titles[i]
                ws.cell(row=5, column=column).fill = PatternFill(start_color='32CD32',
                                                                 end_color='32CD32', fill_type="solid")
                column += 1

        # Results Writer
        for h in range(0, len(outputs)):
            column = start_column
            for i in range(0, len(outputs[h])):
                try:
                    list_checker = len(outputs[h][i])
                    for j in range(0, max_numbers):
                        try:
                            ws.cell(row=row + h + 3, column=column).value = outputs[h][i][j]
                            column += 1
                        except:
                            column += 1
                except:
                    ws.cell(row=row + h + 3, column=column).value = outputs[h][i]
                    column += 1
    if special_case ==1:
        try:
            ws = wb[sheet_name]
            for row in ws[f'A5:{get_column_letter(ws.max_column)}{ws.max_row}']:
                for cell in row:
                    cell.value = None
        except:
            ws = wb.create_sheet("Sheet_1")
            ws.title = sheet_name
            ws.merge_cells('F2:M3')
            ws.cell(row=2, column=6).value = sheet_name
            ws_h2 = ws['F2']
            ws_h2.font = Font(size=23, underline='single', color='FFBB00', bold=True, italic=False)
        # Initialization
        start_column = 1
        row = 3
        column = start_column
        titles.insert(0, "Angle Deviation");
        titles.insert(0, "Model Number")
        for h in range(0, len(outputs)):
            outputs[h].insert(0, k);
            outputs[h].insert(0, nth_iteration)

        global story_drifts
        story_drifts.append(outputs)

        index = titles.index("NumberResults")
        max_numbers = 1
        for m in range(0, len(story_drifts)):
            for n in range(0, len(outputs)):
                if story_drifts[m][n][index] >= max_numbers:
                    max_numbers = outputs[n][index]

        # Titles Writer
        for i in range(0, len(outputs[0])):
            try:
                list_checker = len(outputs[0][i])
                for j in range(0, max_numbers):
                    if j == 0:
                        ws.cell(row=5, column=column).value = titles[i]
                        ws.cell(row=5, column=column).fill = PatternFill(start_color='32CD32',
                                                                                 end_color='32CD32', fill_type="solid")
                        column += 1
                    else:
                        ws.cell(row=5, column=column).fill = PatternFill(start_color='32CD32',
                                                                                 end_color='32CD32', fill_type="solid")
                        column += 1
            except:
                ws.cell(row=5, column=column).value = titles[i]
                ws.cell(row=5, column=column).fill = PatternFill(start_color='32CD32',
                                                                 end_color='32CD32', fill_type="solid")
                column += 1

        # Results Writer
        for g in range(0, len(story_drifts)):
            for h in range(0, len(story_drifts[g])):
                column = start_column
                block_rows = len(story_drifts[g])
                assigned_block_rows = block_rows + 2
                for i in range(0, len(story_drifts[g][h])):
                    try:
                        list_checker = len(story_drifts[g][h][i])
                        for j in range(0, max_numbers):
                            try:
                                ws.cell(row=row + g*(assigned_block_rows) + h + 3, column=column).value = story_drifts[g][h][i][j]
                                column += 1
                            except:
                                column += 1
                    except:
                        ws.cell(row=row + g*(assigned_block_rows) + h + 3, column=column).value = story_drifts[g][h][i]
                        column += 1

    wb.save(filename=f_result)
    wb.close()
def BaseReact(nth_iteration, k):
    sheet_name = "BaseReact"
    outputs = []

    for j in range(0, len(cases_output)):
        case = cases_output[j]
        ret = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
        ret = SapModel.Results.Setup.SetCaseSelectedForOutput(case)

        NumberResults = 0
        LoadCase = []
        StepType = []
        StepNum = []
        Fx = []
        Fy = []
        Fz = []
        Mx = []
        ParamMy = []
        Mz = []
        gx = 0.00
        gy = 0.00
        gz = 0.00

        [NumberResults, LoadCase, StepType, StepNum, Fx, Fy, Fz, Mx, ParamMy, Mz,
         gx, gy, gz, ret] = SapModel.Results.BaseReact(NumberResults, LoadCase, StepType, StepNum, Fx, Fy, Fz, Mx,
                                                       ParamMy, Mz,
                                                       gx, gy, gz)
        titles = ["NumberResults", "LoadCase", "StepType", "StepNum", "Fx", "Fy", "Fz", "Mx", "ParamMy", "Mz",
                  "gx", "gy", "gz"]
        results = [NumberResults, LoadCase, StepType, StepNum, Fx, Fy, Fz, Mx, ParamMy, Mz,
                   gx, gy, gz]
        outputs.append(results)
    for j in range(0, len(combos_output)):
        combo = combos_output[j]
        ret = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
        ret = SapModel.Results.Setup.SetComboSelectedForOutput(combo)

        NumberResults = 0
        LoadCase = []
        StepType = []
        StepNum = []
        Fx = []
        Fy = []
        Fz = []
        Mx = []
        ParamMy = []
        Mz = []
        gx = 0.00
        gy = 0.00
        gz = 0.00

        [NumberResults, LoadCase, StepType, StepNum, Fx, Fy, Fz, Mx, ParamMy, Mz,
         gx, gy, gz, ret] = SapModel.Results.BaseReact(NumberResults, LoadCase, StepType, StepNum, Fx, Fy, Fz, Mx,
                                                       ParamMy, Mz,
                                                       gx, gy, gz)
        titles = ["NumberResults", "LoadCase", "StepType", "StepNum", "Fx", "Fy", "Fz", "Mx", "ParamMy", "Mz",
                  "gx", "gy", "gz"]
        results = [NumberResults, LoadCase, StepType, StepNum, Fx, Fy, Fz, Mx, ParamMy, Mz,
                   gx, gy, gz]
        outputs.append(results)

    output_writer(sheet_name, titles, outputs, nth_iteration, k)
def StoryDrifts(nth_iteration, k):
    sheet_name = "StoryDrifts"
    outputs = []

    for j in range(0, len(cases_output)):
        combo = cases_output[j]
        ret = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
        ret = SapModel.Results.Setup.SetCaseSelectedForOutput(combo)
        NumberResults = 0
        Story = []
        LoadCase = []
        StepType = []
        StepNum = []
        Direction = []
        Drift = []
        Label = []
        X = []
        Y = []
        Z = []
        [NumberResults, Story, LoadCase, StepType, StepNum, Direction, Drift, Label, X, Y, Z, ret] = \
            SapModel.Results.StoryDrifts(NumberResults, Story, LoadCase, StepType, StepNum, Direction, Drift,
                                         Label, X, Y, Z)
        titles = ["NumberResults", "Story", "LoadCase", "StepType", "StepNum", "Direction", "Drift", "Label", "X", "Y",
                  "Z"]
        results = [NumberResults, Story, LoadCase, StepType, StepNum, Direction, Drift, Label, X, Y, Z]
        outputs.append(results)
    for j in range(0, len(combos_output)):
        combo = combos_output[j]
        ret = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
        ret = SapModel.Results.Setup.SetComboSelectedForOutput(combo)
        NumberResults = 0
        Story = []
        LoadCase = []
        StepType = []
        StepNum = []
        Direction = []
        Drift = []
        Label = []
        X = []
        Y = []
        Z = []
        [NumberResults, Story, LoadCase, StepType, StepNum, Direction, Drift, Label, X, Y, Z, ret] = \
            SapModel.Results.StoryDrifts(NumberResults, Story, LoadCase, StepType, StepNum, Direction, Drift,
                                         Label, X, Y, Z)
        titles = ["NumberResults", "Story", "LoadCase", "StepType", "StepNum", "Direction", "Drift", "Label", "X", "Y",
                  "Z"]
        results = [NumberResults, Story, LoadCase, StepType, StepNum, Direction, Drift, Label, X, Y, Z]
        outputs.append(results)

    output_writer(sheet_name, titles, outputs, nth_iteration, k, special_case=1)
def BaseReactWithCentroid(nth_iteration, k, i = 0):
    global ref_Reaction
    ref_Reaction = 0
    C_Method = ''

    sheet_name = 'BaseReactWithCentroid'
    outputs = []
    for j in range(0, len(cases_output)):
        cases = cases_output[j]
        ret = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
        ret = SapModel.Results.Setup.SetCaseSelectedForOutput(cases)
        NumberResults = 0
        LoadCase = []
        StepType = []
        StepNum = []
        FX = []
        FY = []
        FZ = []
        MX = []
        ParamMy = []
        MZ = []
        GX = 0.00
        GY = 0.00
        GZ = 0.00
        XCentroidForFX = []
        YCentroidForFX = []
        ZCentroidForFX = []
        XCentroidForFY = []
        YCentroidForFY = []
        ZCentroidForFY = []
        XCentroidForFZ = []
        YCentroidForFZ = []
        ZCentroidForFZ = []
        [NumberResults,
         LoadCase,
         StepType,
         StepNum,
         FX,
         FY,
         FZ,
         MX,
         ParamMy,
         MZ,
         GX,
         GY,
         GZ,
         XCentroidForFX,
         YCentroidForFX,
         ZCentroidForFX,
         XCentroidForFY,
         YCentroidForFY,
         ZCentroidForFY,
         XCentroidForFZ,
         YCentroidForFZ,
         ZCentroidForFZ, ret] = SapModel.Results.BaseReactWithCentroid(NumberResults,
                                                                       LoadCase,
                                                                       StepType,
                                                                       StepNum,
                                                                       FX,
                                                                       FY,
                                                                       FZ,
                                                                       MX,
                                                                       ParamMy,
                                                                       MZ,
                                                                       GX,
                                                                       GY,
                                                                       GZ,
                                                                       XCentroidForFX,
                                                                       YCentroidForFX,
                                                                       ZCentroidForFX,
                                                                       XCentroidForFY,
                                                                       YCentroidForFY,
                                                                       ZCentroidForFY,
                                                                       XCentroidForFZ,
                                                                       YCentroidForFZ,
                                                                       ZCentroidForFZ)
        titles = ["NumberResults", "LoadCase", "StepType", "StepNum", "FX", "FY", "FZ", "MX", "ParamMy", "MZ", "GX",
                  "GY", "GZ", "XCentroidForFX", "YCentroidForFX", "ZCentroidForFX", "XCentroidForFY", "YCentroidForFY",
                  "ZCentroidForFY", "XCentroidForFZ", "YCentroidForFZ", "ZCentroidForFZ"]
        results = [NumberResults,
                   LoadCase,
                   StepType,
                   StepNum,
                   FX,
                   FY,
                   FZ,
                   MX,
                   ParamMy,
                   MZ,
                   GX,
                   GY,
                   GZ,
                   XCentroidForFX,
                   YCentroidForFX,
                   ZCentroidForFX,
                   XCentroidForFY,
                   YCentroidForFY,
                   ZCentroidForFY,
                   XCentroidForFZ,
                   YCentroidForFZ,
                   ZCentroidForFZ]
        outputs.append(results)


        #Critical angle determination Step
        try:
            global max_Angle
            if Calculation_Option == 1:
                C_Method = "Max Base Reaction(Along Reference Axis)"
                if LoadCase[0] in critical_cases:
                    if ref_Reaction == 0 or FX[0] > ref_Reaction:
                        ref_Reaction = FX[0]
                        if i == 0:
                            max_Angle = int(LoadCase[0].split(" ")[-1])
                        else:
                            max_Angle = float(LoadCase[0].split(" ")[-1])

            if Calculation_Option == 2:
                C_Method = "Min Base Reaction(Transverse of Reference Axis)"
                if LoadCase[0] in critical_cases:
                    if ref_Reaction == 0 or FY[0] < ref_Reaction:
                        ref_Reaction = FY[0]
                        if i == 0:
                            max_Angle = int(LoadCase[0].split(" ")[-1])
                        else:
                            max_Angle = float(LoadCase[0].split(" ")[-1])

            if Calculation_Option == 3:
                C_Method = "Max Base Reaction Difference(Along Reference Axis - Transverse Axis)"
                if LoadCase[0] in critical_cases:
                    if ref_Reaction == 0 or  (FX[0] - FY[0]) > ref_Reaction:
                        ref_Reaction = FX[0] - FY[0]
                        if i == 0:
                            max_Angle = int(LoadCase[0].split(" ")[-1])
                        else:
                            max_Angle = float(LoadCase[0].split(" ")[-1])
        except:
            pass
    output_writer("Critical Angle", ["Method", "Refining Step", "Critical Angle"], [[[C_Method], i+1, max_Angle]], nth_iteration, k)
    print(f"Critical Angle for {k} degree model in {i+1} level of refinement with option number {Calculation_Option} is {max_Angle}")
    for j in range(0, len(combos_output)):
        combo = combos_output[j]
        ret = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
        ret = SapModel.Results.Setup.SetComboSelectedForOutput(combo)
        NumberResults = 0
        LoadCase = []
        StepType = []
        StepNum = []
        FX = []
        FY = []
        FZ = []
        MX = []
        ParamMy = []
        MZ = []
        GX = 0.00
        GY = 0.00
        GZ = 0.00
        XCentroidForFX = []
        YCentroidForFX = []
        ZCentroidForFX = []
        XCentroidForFY = []
        YCentroidForFY = []
        ZCentroidForFY = []
        XCentroidForFZ = []
        YCentroidForFZ = []
        ZCentroidForFZ = []
        [NumberResults,
         LoadCase,
         StepType,
         StepNum,
         FX,
         FY,
         FZ,
         MX,
         ParamMy,
         MZ,
         GX,
         GY,
         GZ,
         XCentroidForFX,
         YCentroidForFX,
         ZCentroidForFX,
         XCentroidForFY,
         YCentroidForFY,
         ZCentroidForFY,
         XCentroidForFZ,
         YCentroidForFZ,
         ZCentroidForFZ, ret] = SapModel.Results.BaseReactWithCentroid(NumberResults,
                                                                       LoadCase,
                                                                       StepType,
                                                                       StepNum,
                                                                       FX,
                                                                       FY,
                                                                       FZ,
                                                                       MX,
                                                                       ParamMy,
                                                                       MZ,
                                                                       GX,
                                                                       GY,
                                                                       GZ,
                                                                       XCentroidForFX,
                                                                       YCentroidForFX,
                                                                       ZCentroidForFX,
                                                                       XCentroidForFY,
                                                                       YCentroidForFY,
                                                                       ZCentroidForFY,
                                                                       XCentroidForFZ,
                                                                       YCentroidForFZ,
                                                                       ZCentroidForFZ)
        titles = ["NumberResults", "LoadCase", "StepType", "StepNum", "FX", "FY", "FZ", "MX", "ParamMy", "MZ", "GX",
                  "GY", "GZ", "XCentroidForFX", "YCentroidForFX", "ZCentroidForFX", "XCentroidForFY", "YCentroidForFY",
                  "ZCentroidForFY", "XCentroidForFZ", "YCentroidForFZ", "ZCentroidForFZ"]
        results = [NumberResults,
                   LoadCase,
                   StepType,
                   StepNum,
                   FX,
                   FY,
                   FZ,
                   MX,
                   ParamMy,
                   MZ,
                   GX,
                   GY,
                   GZ,
                   XCentroidForFX,
                   YCentroidForFX,
                   ZCentroidForFX,
                   XCentroidForFY,
                   YCentroidForFY,
                   ZCentroidForFY,
                   XCentroidForFZ,
                   YCentroidForFZ,
                   ZCentroidForFZ]
        outputs.append(results)

    output_writer(sheet_name, titles, outputs, nth_iteration, k)
def BucklingFactor(nth_iteration, k):
    sheet_name = 'BucklingFactor'
    outputs = []

    for j in range(0, len(cases_output)):
        combo = cases_output[j]
        ret = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
        ret = SapModel.Results.Setup.SetCaseSelectedForOutput(combo)
        NumberResults = 0
        LoadCase = []
        StepType = []
        StepNum = []
        Factor = []
        [NumberResults,
         LoadCase,
         StepType,
         StepNum,
         Factor, ret] = \
            SapModel.Results.BucklingFactor(NumberResults,
                                            LoadCase,
                                            StepType,
                                            StepNum,
                                            Factor)
        titles = ["NumberResults", "LoadCase", "StepType", "StepNum", "Factor"]
        results = [NumberResults,
                   LoadCase,
                   StepType,
                   StepNum,
                   Factor]
        outputs.append(results)
    for j in range(0, len(combos_output)):
        combo = cases_output[j]
        ret = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
        ret = SapModel.Results.Setup.SetComboSelectedForOutput(combo)
        NumberResults = 0
        LoadCase = []
        StepType = []
        StepNum = []
        Factor = []
        [NumberResults,
         LoadCase,
         StepType,
         StepNum,
         Factor, ret] = \
            SapModel.Results.BucklingFactor(NumberResults,
                                            LoadCase,
                                            StepType,
                                            StepNum,
                                            Factor)
        titles = ["NumberResults", "LoadCase", "StepType", "StepNum", "Factor"]
        results = [NumberResults,
                   LoadCase,
                   StepType,
                   StepNum,
                   Factor]
        outputs.append(results)

    output_writer(sheet_name, titles, outputs, nth_iteration, k)
def FrameForce(nth_iteration, k):
    sheet_name = 'FrameForce'
    outputs = []

    for i in range(0, len(elements)):
        for j in range(0, len(cases_output)):
            combo = cases_output[j]
            ret = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
            ret = SapModel.Results.Setup.SetCaseSelectedForOutput(combo)
            ItemTypeElm = 0
            NumberResults = 0
            Obj = []
            ObjSta = []
            Elm = []
            ElmSta = []
            LoadCase = []
            StepType = []
            StepNum = []
            P = []
            V2 = []
            V3 = []
            T = []
            M2 = []
            M3 = []
            [NumberResults,
             Obj,
             ObjSta,
             Elm,
             ElmSta,
             LoadCase,
             StepType,
             StepNum,
             P,
             V2,
             V3,
             T,
             M2,
             M3, ret] = SapModel.Results.FrameForce(elements[i],
                                                    ItemTypeElm,
                                                    NumberResults,
                                                    Obj,
                                                    ObjSta,
                                                    Elm,
                                                    ElmSta,
                                                    LoadCase,
                                                    StepType,
                                                    StepNum,
                                                    P,
                                                    V2,
                                                    V3,
                                                    T,
                                                    M2,
                                                    M3)
        titles = ["NumberResults", "Obj", "ObjSta", "Elm", "ElmSta", "LoadCase", "StepType", "StepNum", "P", "V2", "V3",
                  "T", "M2", "M3"]
        results = [NumberResults,
                   Obj,
                   ObjSta,
                   Elm,
                   ElmSta,
                   LoadCase,
                   StepType,
                   StepNum,
                   P,
                   V2,
                   V3,
                   T,
                   M2,
                   M3]
        outputs.append(results)
    for i in range(0, len(elements)):
        for j in range(0, len(combos_output)):
            combo = combos_output[j]
            ret = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
            ret = SapModel.Results.Setup.SetComboSelectedForOutput(combo)
            ItemTypeElm = 0
            NumberResults = 0
            Obj = []
            ObjSta = []
            Elm = []
            ElmSta = []
            LoadCase = []
            StepType = []
            StepNum = []
            P = []
            V2 = []
            V3 = []
            T = []
            M2 = []
            M3 = []
            [NumberResults,
             Obj,
             ObjSta,
             Elm,
             ElmSta,
             LoadCase,
             StepType,
             StepNum,
             P,
             V2,
             V3,
             T,
             M2,
             M3, ret] = SapModel.Results.FrameForce(elements[i],
                                                    ItemTypeElm,
                                                    NumberResults,
                                                    Obj,
                                                    ObjSta,
                                                    Elm,
                                                    ElmSta,
                                                    LoadCase,
                                                    StepType,
                                                    StepNum,
                                                    P,
                                                    V2,
                                                    V3,
                                                    T,
                                                    M2,
                                                    M3)
        titles = ["NumberResults", "Obj", "ObjSta", "Elm", "ElmSta", "LoadCase", "StepType", "StepNum", "P", "V2", "V3",
                  "T", "M2", "M3"]
        results = [NumberResults,
                   Obj,
                   ObjSta,
                   Elm,
                   ElmSta,
                   LoadCase,
                   StepType,
                   StepNum,
                   P,
                   V2,
                   V3,
                   T,
                   M2,
                   M3]
        outputs.append(results)

    output_writer(sheet_name, titles, outputs, nth_iteration, k)
def FrameJointForce(nth_iteration, k):
    sheet_name = 'FrameJointForce'
    outputs = []

    for i in range(0, len(elements)):
        for j in range(0, len(cases_output)):
            combo = cases_output[j]
            ret = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
            ret = SapModel.Results.Setup.SetCaseSelectedForOutput(combo)
            ItemTypeElm = 0
            NumberResults = 0
            Obj = []
            Elm = []
            PointElm = []
            LoadCase = []
            StepType = []
            StepNum = []
            F1 = []
            F2 = []
            F3 = []
            M1 = []
            M2 = []
            M3 = []

            [NumberResults,
             Obj,
             Elm,
             PointElm,
             LoadCase,
             StepType,
             StepNum,
             F1,
             F2,
             F3,
             M1,
             M2,
             M3,
             ret] = SapModel.Results.FrameJointForce(elements[i],
                                                     ItemTypeElm,
                                                     NumberResults,
                                                     Obj,
                                                     Elm,
                                                     PointElm,
                                                     LoadCase,
                                                     StepType,
                                                     StepNum,
                                                     F1,
                                                     F2,
                                                     F3,
                                                     M1,
                                                     M2,
                                                     M3)

            titles = ["NumberResults", "Obj", "Elm", "PointElm", "LoadCase", "StepType", "StepNum", "F1", "F2", "F3",
                      "M1", "M2", "M3"]
            results = [NumberResults, Obj, Elm, PointElm, LoadCase, StepType, StepNum, F1, F2, F3, M1, M2, M3]
            outputs.append(results)
    for i in range(0, len(elements)):
        for j in range(0, len(combos_output)):
            combo = combos_output[j]
            ret = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
            ret = SapModel.Results.Setup.SetComboSelectedForOutput(combo)
            ItemTypeElm = 0
            NumberResults = 0
            Obj = []
            Elm = []
            PointElm = []
            LoadCase = []
            StepType = []
            StepNum = []
            F1 = []
            F2 = []
            F3 = []
            M1 = []
            M2 = []
            M3 = []

            [NumberResults,
             Obj,
             Elm,
             PointElm,
             LoadCase,
             StepType,
             StepNum,
             F1,
             F2,
             F3,
             M1,
             M2,
             M3,
             ret] = SapModel.Results.FrameJointForce(elements[i],
                                                     ItemTypeElm,
                                                     NumberResults,
                                                     Obj,
                                                     Elm,
                                                     PointElm,
                                                     LoadCase,
                                                     StepType,
                                                     StepNum,
                                                     F1,
                                                     F2,
                                                     F3,
                                                     M1,
                                                     M2,
                                                     M3)

            titles = ["NumberResults", "Obj", "Elm", "PointElm", "LoadCase", "StepType", "StepNum", "F1", "F2", "F3",
                      "M1", "M2", "M3"]
            results = [NumberResults, Obj, Elm, PointElm, LoadCase, StepType, StepNum, F1, F2, F3, M1, M2, M3]
            outputs.append(results)

    output_writer(sheet_name, titles, outputs, nth_iteration, k)
def JointAcc(nth_iteration, k):
    sheet_name = 'JointAcc'
    outputs = []
    for i in range(0, len(joints)):
        for j in range(0, len(cases_output)):
            combo = cases_output[j]
            ret = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
            ret = SapModel.Results.Setup.SetCaseSelectedForOutput(combo)
            ItemTypeElm = 0
            NumberResults = 0
            Obj = []
            Elm = []
            LoadCase = []
            StepType = []
            StepNum = []
            U1 = []
            U2 = []
            U3 = []
            R1 = []
            R2 = []
            R3 = []

            [NumberResults,
             Obj,
             Elm,
             LoadCase,
             StepType,
             StepNum,
             U1,
             U2,
             U3,
             R1,
             R2,
             R3, ret] = SapModel.Results.JointAcc(joints[i],
                                                  ItemTypeElm,
                                                  NumberResults,
                                                  Obj,
                                                  Elm,
                                                  LoadCase,
                                                  StepType,
                                                  StepNum,
                                                  U1,
                                                  U2,
                                                  U3,
                                                  R1,
                                                  R2,
                                                  R3)
            titles = ["NumberResults", "Obj", "Elm", "LoadCase", "StepType", "StepNum", "U1", "U2", "U3", "R1", "R2",
                      "R3"]
            results = [NumberResults,
                       Obj,
                       Elm,
                       LoadCase,
                       StepType,
                       StepNum,
                       U1,
                       U2,
                       U3,
                       R1,
                       R2,
                       R3]
            outputs.append(results)
    for i in range(0, len(joints)):
        for j in range(0, len(combos_output)):
            combo = combos_output[j]
            ret = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
            ret = SapModel.Results.Setup.SetComboSelectedForOutput(combo)
            ItemTypeElm = 0
            NumberResults = 0
            Obj = []
            Elm = []
            LoadCase = []
            StepType = []
            StepNum = []
            U1 = []
            U2 = []
            U3 = []
            R1 = []
            R2 = []
            R3 = []

            [NumberResults,
             Obj,
             Elm,
             LoadCase,
             StepType,
             StepNum,
             U1,
             U2,
             U3,
             R1,
             R2,
             R3, ret] = SapModel.Results.JointAcc(joints[i],
                                                  ItemTypeElm,
                                                  NumberResults,
                                                  Obj,
                                                  Elm,
                                                  LoadCase,
                                                  StepType,
                                                  StepNum,
                                                  U1,
                                                  U2,
                                                  U3,
                                                  R1,
                                                  R2,
                                                  R3)
            titles = ["NumberResults", "Obj", "Elm", "LoadCase", "StepType", "StepNum", "U1", "U2", "U3", "R1", "R2",
                      "R3"]
            results = [NumberResults,
                       Obj,
                       Elm,
                       LoadCase,
                       StepType,
                       StepNum,
                       U1,
                       U2,
                       U3,
                       R1,
                       R2,
                       R3]
            outputs.append(results)

    output_writer(sheet_name, titles, outputs, nth_iteration, k)
# Only Applicable in case of Time History Load Cases to Find out the absolute movement of the Structure
def JointAccAbs(nth_iteration, k):
    sheet_name = 'JointAccAbs'
    outputs = []

    for i in range(0, len(joints)):
        for j in range(0, len(cases_output)):
            combo = cases_output[j]
            ret = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
            ret = SapModel.Results.Setup.SetCaseSelectedForOutput(combo)
            ItemTypeElm = 0
            NumberResults = 0
            Obj = []
            Elm = []
            LoadCase = []
            StepType = []
            StepNum = []
            U1 = []
            U2 = []
            U3 = []
            R1 = []
            R2 = []
            R3 = []

            [NumberResults,
             Obj,
             Elm,
             LoadCase,
             StepType,
             StepNum,
             U1,
             U2,
             U3,
             R1,
             R2,
             R3, ret] = SapModel.Results.JointAccAbs(joints[i],
                                                     ItemTypeElm,
                                                     NumberResults,
                                                     Obj,
                                                     Elm,
                                                     LoadCase,
                                                     StepType,
                                                     StepNum,
                                                     U1,
                                                     U2,
                                                     U3,
                                                     R1,
                                                     R2,
                                                     R3)
            titles = ["NumberResults", "Obj", "Elm", "LoadCase", "StepType", "StepNum", "U1", "U2", "U3", "R1", "R2",
                      "R3"]
            results = [NumberResults,
                       Obj,
                       Elm,
                       LoadCase,
                       StepType,
                       StepNum,
                       U1,
                       U2,
                       U3,
                       R1,
                       R2,
                       R3]
            outputs.append(results)
    for i in range(0, len(joints)):
        for j in range(0, len(combos_output)):
            combo = combos_output[j]
            ret = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
            ret = SapModel.Results.Setup.SetComboSelectedForOutput(combo)
            ItemTypeElm = 0
            NumberResults = 0
            Obj = []
            Elm = []
            LoadCase = []
            StepType = []
            StepNum = []
            U1 = []
            U2 = []
            U3 = []
            R1 = []
            R2 = []
            R3 = []

            [NumberResults,
             Obj,
             Elm,
             LoadCase,
             StepType,
             StepNum,
             U1,
             U2,
             U3,
             R1,
             R2,
             R3, ret] = SapModel.Results.JointAccAbs(joints[i],
                                                     ItemTypeElm,
                                                     NumberResults,
                                                     Obj,
                                                     Elm,
                                                     LoadCase,
                                                     StepType,
                                                     StepNum,
                                                     U1,
                                                     U2,
                                                     U3,
                                                     R1,
                                                     R2,
                                                     R3)
            titles = ["NumberResults", "Obj", "Elm", "LoadCase", "StepType", "StepNum", "U1", "U2", "U3", "R1", "R2",
                      "R3"]
            results = [NumberResults,
                       Obj,
                       Elm,
                       LoadCase,
                       StepType,
                       StepNum,
                       U1,
                       U2,
                       U3,
                       R1,
                       R2,
                       R3]
            outputs.append(results)

    output_writer(sheet_name, titles, outputs, nth_iteration, k)
def JointDrifts(nth_iteration, k):
    sheet_name = 'JointDrifts'
    outputs = []

    for j in range(0, len(cases_output)):
        combo = cases_output[j]
        ret = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
        ret = SapModel.Results.Setup.SetCaseSelectedForOutput(combo)
        NumberResults = 0
        Story = []
        Label = []
        Name = []
        LoadCase = []
        StepType = []
        StepNum = []
        DisplacementX = []
        DisplacementY = []
        DriftX = []
        DriftY = []

        [NumberResults,
         Story,
         Label,
         Name,
         LoadCase,
         StepType,
         StepNum,
         DisplacementX,
         DisplacementY,
         DriftX,
         DriftY, ret] = SapModel.Results.JointDrifts(NumberResults,
                                                     Story,
                                                     Label,
                                                     Name,
                                                     LoadCase,
                                                     StepType,
                                                     StepNum,
                                                     DisplacementX,
                                                     DisplacementY,
                                                     DriftX,
                                                     DriftY)
        titles = ["NumberResults", "Story", "Label", "Name", "LoadCase", "StepType", "StepNum", "DisplacementX",
                  "DisplacementY", "DriftX", "DriftY"]
        results = [NumberResults,
                   Story,
                   Label,
                   Name,
                   LoadCase,
                   StepType,
                   StepNum,
                   DisplacementX,
                   DisplacementY,
                   DriftX,
                   DriftY]
        outputs.append(results)
    for j in range(0, len(combos_output)):
        combo = combos_output[j]
        ret = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
        ret = SapModel.Results.Setup.SetComboSelectedForOutput(combo)
        NumberResults = 0
        Story = []
        Label = []
        Name = []
        LoadCase = []
        StepType = []
        StepNum = []
        DisplacementX = []
        DisplacementY = []
        DriftX = []
        DriftY = []

        [NumberResults,
         Story,
         Label,
         Name,
         LoadCase,
         StepType,
         StepNum,
         DisplacementX,
         DisplacementY,
         DriftX,
         DriftY, ret] = SapModel.Results.JointDrifts(NumberResults,
                                                     Story,
                                                     Label,
                                                     Name,
                                                     LoadCase,
                                                     StepType,
                                                     StepNum,
                                                     DisplacementX,
                                                     DisplacementY,
                                                     DriftX,
                                                     DriftY)
        titles = ["NumberResults", "Story", "Label", "Name", "LoadCase", "StepType", "StepNum", "DisplacementX",
                  "DisplacementY", "DriftX", "DriftY"]
        results = [NumberResults,
                   Story,
                   Label,
                   Name,
                   LoadCase,
                   StepType,
                   StepNum,
                   DisplacementX,
                   DisplacementY,
                   DriftX,
                   DriftY]
        outputs.append(results)

    output_writer(sheet_name, titles, outputs, nth_iteration, k)
def JointReact(nth_iteration, k):
    sheet_name = 'JointReact'
    outputs = []

    for i in range(0, len(joints)):
        for j in range(0, len(cases_output)):
            combo = cases_output[j]
            ret = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
            ret = SapModel.Results.Setup.SetCaseSelectedForOutput(combo)
            ItemTypeElm = 0
            NumberResults = 0
            Obj = []
            Elm = []
            LoadCase = []
            StepType = []
            StepNum = []
            F1 = []
            F2 = []
            F3 = []
            M1 = []
            M2 = []
            M3 = []

            [NumberResults,
             Obj,
             Elm,
             LoadCase,
             StepType,
             StepNum,
             F1,
             F2,
             F3,
             M1,
             M2,
             M3, ret] = SapModel.Results.JointReact(joints[i],
                                                    ItemTypeElm,
                                                    NumberResults,
                                                    Obj,
                                                    Elm,
                                                    LoadCase,
                                                    StepType,
                                                    StepNum,
                                                    F1,
                                                    F2,
                                                    F3,
                                                    M1,
                                                    M2,
                                                    M3)
            titles = ["NumberResults", "Obj", "Elm", "LoadCase", "StepType", "StepNum", "F1", "F2", "F3", "M1", "M2",
                      "M3"]
            results = [NumberResults,
                       Obj,
                       Elm,
                       LoadCase,
                       StepType,
                       StepNum,
                       F1,
                       F2,
                       F3,
                       M1,
                       M2,
                       M3]
            outputs.append(results)
    for i in range(0, len(joints)):
        for j in range(0, len(combos_output)):
            combo = combos_output[j]
            ret = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
            ret = SapModel.Results.Setup.SetComboSelectedForOutput(combo)
            ItemTypeElm = 0
            NumberResults = 0
            Obj = []
            Elm = []
            LoadCase = []
            StepType = []
            StepNum = []
            F1 = []
            F2 = []
            F3 = []
            M1 = []
            M2 = []
            M3 = []

            [NumberResults,
             Obj,
             Elm,
             LoadCase,
             StepType,
             StepNum,
             F1,
             F2,
             F3,
             M1,
             M2,
             M3, ret] = SapModel.Results.JointReact(joints[i],
                                                    ItemTypeElm,
                                                    NumberResults,
                                                    Obj,
                                                    Elm,
                                                    LoadCase,
                                                    StepType,
                                                    StepNum,
                                                    F1,
                                                    F2,
                                                    F3,
                                                    M1,
                                                    M2,
                                                    M3)
            titles = ["NumberResults", "Obj", "Elm", "LoadCase", "StepType", "StepNum", "F1", "F2", "F3", "M1", "M2",
                      "M3"]
            results = [NumberResults,
                       Obj,
                       Elm,
                       LoadCase,
                       StepType,
                       StepNum,
                       F1,
                       F2,
                       F3,
                       M1,
                       M2,
                       M3]
            outputs.append(results)

    output_writer(sheet_name, titles, outputs, nth_iteration, k)
def JointVel(nth_iteration, k):
    sheet_name = 'JointVel'
    outputs = []

    for i in range(0, len(joints)):
        for j in range(0, len(cases_output)):
            combo = cases_output[j]
            ret = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
            ret = SapModel.Results.Setup.SetCaseSelectedForOutput(combo)
            ItemTypeElm = 0
            NumberResults = 0
            Obj = []
            Elm = []
            LoadCase = []
            StepType = []
            StepNum = []
            U1 = []
            U2 = []
            U3 = []
            R1 = []
            R2 = []
            R3 = []

            [NumberResults,
             Obj,
             Elm,
             LoadCase,
             StepType,
             StepNum,
             U1,
             U2,
             U3,
             R1,
             R2,
             R3, ret] = SapModel.Results.JointVel(joints[i],
                                                  ItemTypeElm,
                                                  NumberResults,
                                                  Obj,
                                                  Elm,
                                                  LoadCase,
                                                  StepType,
                                                  StepNum,
                                                  U1,
                                                  U2,
                                                  U3,
                                                  R1,
                                                  R2,
                                                  R3)
            titles = ["NumberResults", "Obj", "Elm", "LoadCase", "StepType", "StepNum", "U1", "U2", "U3", "R1", "R2",
                      "R3"]
            results = [NumberResults,
                       Obj,
                       Elm,
                       LoadCase,
                       StepType,
                       StepNum,
                       U1,
                       U2,
                       U3,
                       R1,
                       R2,
                       R3]
            outputs.append(results)
    for i in range(0, len(joints)):
        for j in range(0, len(combos_output)):
            combo = combos_output[j]
            ret = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
            ret = SapModel.Results.Setup.SetComboSelectedForOutput(combo)
            ItemTypeElm = 0
            NumberResults = 0
            Obj = []
            Elm = []
            LoadCase = []
            StepType = []
            StepNum = []
            U1 = []
            U2 = []
            U3 = []
            R1 = []
            R2 = []
            R3 = []

            [NumberResults,
             Obj,
             Elm,
             LoadCase,
             StepType,
             StepNum,
             U1,
             U2,
             U3,
             R1,
             R2,
             R3, ret] = SapModel.Results.JointVel(joints[i],
                                                  ItemTypeElm,
                                                  NumberResults,
                                                  Obj,
                                                  Elm,
                                                  LoadCase,
                                                  StepType,
                                                  StepNum,
                                                  U1,
                                                  U2,
                                                  U3,
                                                  R1,
                                                  R2,
                                                  R3)
            titles = ["NumberResults", "Obj", "Elm", "LoadCase", "StepType", "StepNum", "U1", "U2", "U3", "R1", "R2",
                      "R3"]
            results = [NumberResults,
                       Obj,
                       Elm,
                       LoadCase,
                       StepType,
                       StepNum,
                       U1,
                       U2,
                       U3,
                       R1,
                       R2,
                       R3]
            outputs.append(results)

    output_writer(sheet_name, titles, outputs, nth_iteration, k)
def JointVelAbs(nth_iteration, k):
    sheet_name = 'JointVelAbs'
    outputs = []

    for i in range(0, len(joints)):
        for j in range(0, len(cases_output)):
            combo = cases_output[j]
            ret = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
            ret = SapModel.Results.Setup.SetCaseSelectedForOutput(combo)
            ItemTypeElm = 0
            NumberResults = 0
            Obj = []
            Elm = []
            LoadCase = []
            StepType = []
            StepNum = []
            U1 = []
            U2 = []
            U3 = []
            R1 = []
            R2 = []
            R3 = []

            [NumberResults,
             Obj,
             Elm,
             LoadCase,
             StepType,
             StepNum,
             U1,
             U2,
             U3,
             R1,
             R2,
             R3, ret] = SapModel.Results.JointVelAbs(joints[i],
                                                     ItemTypeElm,
                                                     NumberResults,
                                                     Obj,
                                                     Elm,
                                                     LoadCase,
                                                     StepType,
                                                     StepNum,
                                                     U1,
                                                     U2,
                                                     U3,
                                                     R1,
                                                     R2,
                                                     R3)
            titles = ["NumberResults", "Obj", "Elm", "LoadCase", "StepType", "StepNum", "U1", "U2", "U3", "R1", "R2",
                      "R3"]
            results = [NumberResults,
                       Obj,
                       Elm,
                       LoadCase,
                       StepType,
                       StepNum,
                       U1,
                       U2,
                       U3,
                       R1,
                       R2,
                       R3]
            outputs.append(results)
    for i in range(0, len(joints)):
        for j in range(0, len(combos_output)):
            combo = combos_output[j]
            ret = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
            ret = SapModel.Results.Setup.SetComboSelectedForOutput(combo)
            ItemTypeElm = 0
            NumberResults = 0
            Obj = []
            Elm = []
            LoadCase = []
            StepType = []
            StepNum = []
            U1 = []
            U2 = []
            U3 = []
            R1 = []
            R2 = []
            R3 = []

            [NumberResults,
             Obj,
             Elm,
             LoadCase,
             StepType,
             StepNum,
             U1,
             U2,
             U3,
             R1,
             R2,
             R3, ret] = SapModel.Results.JointVelAbs(joints[i],
                                                     ItemTypeElm,
                                                     NumberResults,
                                                     Obj,
                                                     Elm,
                                                     LoadCase,
                                                     StepType,
                                                     StepNum,
                                                     U1,
                                                     U2,
                                                     U3,
                                                     R1,
                                                     R2,
                                                     R3)
            titles = ["NumberResults", "Obj", "Elm", "LoadCase", "StepType", "StepNum", "U1", "U2", "U3", "R1", "R2",
                      "R3"]
            results = [NumberResults,
                       Obj,
                       Elm,
                       LoadCase,
                       StepType,
                       StepNum,
                       U1,
                       U2,
                       U3,
                       R1,
                       R2,
                       R3]
            outputs.append(results)

    output_writer(sheet_name, titles, outputs, nth_iteration, k)
def ModalLoadParticipationRatios(nth_iteration, k):
    sheet_name = 'ModalLoadParticipationRatios'
    outputs = []

    for j in range(0, len(cases_output)):
        combo = cases_output[j]
        ret = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
        ret = SapModel.Results.Setup.SetCaseSelectedForOutput(combo)
        NumberResults = 0
        LoadCase = []
        ItemType = []
        Item = []
        Stat = []
        Dyn = []

        [NumberResults,
         LoadCase,
         ItemType,
         Item,
         Stat,
         Dyn, ret] = SapModel.Results.ModalLoadParticipationRatios(NumberResults,
                                                                   LoadCase,
                                                                   ItemType,
                                                                   Item,
                                                                   Stat,
                                                                   Dyn)
        titles = ["NumberResults", "LoadCase", "ItemType", "Item", "Stat", "Dyn"]
        results = [NumberResults,
                   LoadCase,
                   ItemType,
                   Item,
                   Stat,
                   Dyn]
        outputs.append(results)
    for j in range(0, len(combos_output)):
        combo = combos_output[j]
        ret = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
        ret = SapModel.Results.Setup.SetComboSelectedForOutput(combo)
        NumberResults = 0
        LoadCase = []
        ItemType = []
        Item = []
        Stat = []
        Dyn = []

        [NumberResults,
         LoadCase,
         ItemType,
         Item,
         Stat,
         Dyn, ret] = SapModel.Results.ModalLoadParticipationRatios(NumberResults,
                                                                   LoadCase,
                                                                   ItemType,
                                                                   Item,
                                                                   Stat,
                                                                   Dyn)
        titles = ["NumberResults", "LoadCase", "ItemType", "Item", "Stat", "Dyn"]
        results = [NumberResults,
                   LoadCase,
                   ItemType,
                   Item,
                   Stat,
                   Dyn]
        outputs.append(results)

    output_writer(sheet_name, titles, outputs, nth_iteration, k)
def ModalParticipatingMassRatios(nth_iteration, k):
    sheet_name = 'ModalParticipatingMassRatios'
    outputs = []

    for j in range(0, len(cases_output)):
        combo = cases_output[j]
        ret = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
        ret = SapModel.Results.Setup.SetCaseSelectedForOutput(combo)
        NumberResults = 0
        LoadCase = []
        StepType = []
        StepNum = []
        Period = []
        UX = []
        UY = []
        UZ = []
        SumUX = []
        SumUY = []
        SumUZ = []
        RX = []
        RY = []
        RZ = []
        SumRX = []
        SumRY = []
        SumRZ = []

        [NumberResults,
         LoadCase,
         StepType,
         StepNum,
         Period,
         UX,
         UY,
         UZ,
         SumUX,
         SumUY,
         SumUZ,
         RX,
         RY,
         RZ,
         SumRX,
         SumRY,
         SumRZ
            , ret] = SapModel.Results.ModalParticipatingMassRatios(NumberResults,
                                                                   LoadCase,
                                                                   StepType,
                                                                   StepNum,
                                                                   Period,
                                                                   UX,
                                                                   UY,
                                                                   UZ,
                                                                   SumUX,
                                                                   SumUY,
                                                                   SumUZ,
                                                                   RX,
                                                                   RY,
                                                                   RZ,
                                                                   SumRX,
                                                                   SumRY,
                                                                   SumRZ
                                                                   )
        titles = ["NumberResults", "LoadCase", "StepType", "StepNum", "Period", "UX", "UY", "UZ", "SumUX", "SumUY",
                  "SumUZ", "RX", "RY", "RZ", "SumRX", "SumRY", "SumRZ"]
        results = [NumberResults,
                   LoadCase,
                   StepType,
                   StepNum,
                   Period,
                   UX,
                   UY,
                   UZ,
                   SumUX,
                   SumUY,
                   SumUZ,
                   RX,
                   RY,
                   RZ,
                   SumRX,
                   SumRY,
                   SumRZ]
        outputs.append(results)
    for j in range(0, len(combos_output)):
        combo = combos_output[j]
        ret = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
        ret = SapModel.Results.Setup.SetComboSelectedForOutput(combo)
        NumberResults = 0
        LoadCase = []
        StepType = []
        StepNum = []
        Period = []
        UX = []
        UY = []
        UZ = []
        SumUX = []
        SumUY = []
        SumUZ = []
        RX = []
        RY = []
        RZ = []
        SumRX = []
        SumRY = []
        SumRZ = []

        [NumberResults,
         LoadCase,
         StepType,
         StepNum,
         Period,
         UX,
         UY,
         UZ,
         SumUX,
         SumUY,
         SumUZ,
         RX,
         RY,
         RZ,
         SumRX,
         SumRY,
         SumRZ
            , ret] = SapModel.Results.ModalParticipatingMassRatios(NumberResults,
                                                                   LoadCase,
                                                                   StepType,
                                                                   StepNum,
                                                                   Period,
                                                                   UX,
                                                                   UY,
                                                                   UZ,
                                                                   SumUX,
                                                                   SumUY,
                                                                   SumUZ,
                                                                   RX,
                                                                   RY,
                                                                   RZ,
                                                                   SumRX,
                                                                   SumRY,
                                                                   SumRZ
                                                                   )
        titles = ["NumberResults", "LoadCase", "StepType", "StepNum", "Period", "UX", "UY", "UZ", "SumUX", "SumUY",
                  "SumUZ", "RX", "RY", "RZ", "SumRX", "SumRY", "SumRZ"]
        results = [NumberResults,
                   LoadCase,
                   StepType,
                   StepNum,
                   Period,
                   UX,
                   UY,
                   UZ,
                   SumUX,
                   SumUY,
                   SumUZ,
                   RX,
                   RY,
                   RZ,
                   SumRX,
                   SumRY,
                   SumRZ]
        outputs.append(results)

    output_writer(sheet_name, titles, outputs, nth_iteration, k)
def ModalPeriod(nth_iteration, k):
    sheet_name = 'ModalPeriod'
    outputs = []

    for j in range(0, len(cases_output)):
        combo = cases_output[j]
        ret = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
        ret = SapModel.Results.Setup.SetCaseSelectedForOutput(combo)

        NumberResults = 00
        LoadCase = []
        StepType = []
        StepNum = []
        Period = []
        Frequency = []
        CircFreq = []
        EigenValue = []

        [NumberResults,
         LoadCase,
         StepType,
         StepNum,
         Period,
         Frequency,
         CircFreq,
         EigenValue
            , ret] = SapModel.Results.ModalPeriod(NumberResults,
                                                  LoadCase,
                                                  StepType,
                                                  StepNum,
                                                  Period,
                                                  Frequency,
                                                  CircFreq,
                                                  EigenValue
                                                  )
        titles = ["NumberResults", "LoadCase", "StepType", "StepNum", "Period", "Frequency", "CircFreq", "EigenValue"]
        results = [NumberResults,
                   LoadCase,
                   StepType,
                   StepNum,
                   Period,
                   Frequency,
                   CircFreq,
                   EigenValue]
        outputs.append(results)
    for j in range(0, len(combos_output)):
        combo = combos_output[j]
        ret = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
        ret = SapModel.Results.Setup.SetComboSelectedForOutput(combo)

        NumberResults = 00
        LoadCase = []
        StepType = []
        StepNum = []
        Period = []
        Frequency = []
        CircFreq = []
        EigenValue = []

        [NumberResults,
         LoadCase,
         StepType,
         StepNum,
         Period,
         Frequency,
         CircFreq,
         EigenValue
            , ret] = SapModel.Results.ModalPeriod(NumberResults,
                                                  LoadCase,
                                                  StepType,
                                                  StepNum,
                                                  Period,
                                                  Frequency,
                                                  CircFreq,
                                                  EigenValue
                                                  )
        titles = ["NumberResults", "LoadCase", "StepType", "StepNum", "Period", "Frequency", "CircFreq", "EigenValue"]
        results = [NumberResults,
                   LoadCase,
                   StepType,
                   StepNum,
                   Period,
                   Frequency,
                   CircFreq,
                   EigenValue]
        outputs.append(results)

    output_writer(sheet_name, titles, outputs, nth_iteration, k)
def ModeShape(nth_iteration, k):
    sheet_name = 'ModeShape'
    outputs = []

    for i in range(0, len(joints)):
        for j in range(0, len(cases_output)):
            combo = cases_output[j]
            ret = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
            ret = SapModel.Results.Setup.SetCaseSelectedForOutput(combo)

            ItemTypeElm = 0
            NumberResults = 0
            Obj = []
            Elm = []
            LoadCase = []
            StepType = []
            StepNum = []
            U1 = []
            U2 = []
            U3 = []
            R1 = []
            R2 = []
            R3 = []

            [NumberResults,
             Obj,
             Elm,
             LoadCase,
             StepType,
             StepNum,
             U1,
             U2,
             U3,
             R1,
             R2,
             R3
                , ret] = SapModel.Results.ModeShape(joints[i],
                                                    ItemTypeElm,
                                                    NumberResults,
                                                    Obj,
                                                    Elm,
                                                    LoadCase,
                                                    StepType,
                                                    StepNum,
                                                    U1,
                                                    U2,
                                                    U3,
                                                    R1,
                                                    R2,
                                                    R3)
            titles = ["NumberResults", "Obj", "Elm", "LoadCase", "StepType", "StepNum", "U1", "U2", "U3", "R1", "R2",
                      "R3"]
            results = [NumberResults,
                       Obj,
                       Elm,
                       LoadCase,
                       StepType,
                       StepNum,
                       U1,
                       U2,
                       U3,
                       R1,
                       R2,
                       R3]
            outputs.append(results)
    for i in range(0, len(joints)):
        for j in range(0, len(combos_output)):
            combo = combos_output[j]
            ret = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
            ret = SapModel.Results.Setup.SetComboSelectedForOutput(combo)

            ItemTypeElm = 0
            NumberResults = 0
            Obj = []
            Elm = []
            LoadCase = []
            StepType = []
            StepNum = []
            U1 = []
            U2 = []
            U3 = []
            R1 = []
            R2 = []
            R3 = []

            [NumberResults,
             Obj,
             Elm,
             LoadCase,
             StepType,
             StepNum,
             U1,
             U2,
             U3,
             R1,
             R2,
             R3
                , ret] = SapModel.Results.ModeShape(joints[i],
                                                    ItemTypeElm,
                                                    NumberResults,
                                                    Obj,
                                                    Elm,
                                                    LoadCase,
                                                    StepType,
                                                    StepNum,
                                                    U1,
                                                    U2,
                                                    U3,
                                                    R1,
                                                    R2,
                                                    R3)
            titles = ["NumberResults", "Obj", "Elm", "LoadCase", "StepType", "StepNum", "U1", "U2", "U3", "R1", "R2",
                      "R3"]
            results = [NumberResults,
                       Obj,
                       Elm,
                       LoadCase,
                       StepType,
                       StepNum,
                       U1,
                       U2,
                       U3,
                       R1,
                       R2,
                       R3]
            outputs.append(results)

    output_writer(sheet_name, titles, outputs, nth_iteration, k)
def SectionCutAnalysis(nth_iteration, k):
    sheet_name = 'SectionCutAnalysis'
    outputs = []

    for i in range(0, len(joints)):
        for j in range(0, len(cases_output)):
            combo = cases_output[j]
            ret = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
            ret = SapModel.Results.Setup.SetCaseSelectedForOutput(combo)

            NumberResults = 0
            SCut = []
            LoadCase = []
            StepType = []
            StepNum = []
            F1 = []
            F2 = []
            F3 = []
            M1 = []
            M2 = []
            M3 = []

            [NumberResults,
             SCut,
             LoadCase,
             StepType,
             StepNum,
             F1,
             F2,
             F3,
             M1,
             M2,
             M3
                , ret] = SapModel.Results.SectionCutAnalysis(NumberResults,
                                                             SCut,
                                                             LoadCase,
                                                             StepType,
                                                             StepNum,
                                                             F1,
                                                             F2,
                                                             F3,
                                                             M1,
                                                             M2,
                                                             M3)
            titles = ["NumberResults", "SCut", "LoadCase", "StepType", "StepNum", "F1", "F2", "F3", "M1", "M2", "M3"]
            results = [NumberResults,
                       SCut,
                       LoadCase,
                       StepType,
                       StepNum,
                       F1,
                       F2,
                       F3,
                       M1,
                       M2,
                       M3]
            outputs.append(results)
    for i in range(0, len(joints)):
        for j in range(0, len(combos_output)):
            combo = combos_output[j]
            ret = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
            ret = SapModel.Results.Setup.SetComboSelectedForOutput(combo)

            NumberResults = 0
            SCut = []
            LoadCase = []
            StepType = []
            StepNum = []
            F1 = []
            F2 = []
            F3 = []
            M1 = []
            M2 = []
            M3 = []

            [NumberResults,
             SCut,
             LoadCase,
             StepType,
             StepNum,
             F1,
             F2,
             F3,
             M1,
             M2,
             M3
                , ret] = SapModel.Results.SectionCutAnalysis(NumberResults,
                                                             SCut,
                                                             LoadCase,
                                                             StepType,
                                                             StepNum,
                                                             F1,
                                                             F2,
                                                             F3,
                                                             M1,
                                                             M2,
                                                             M3)
            titles = ["NumberResults", "SCut", "LoadCase", "StepType", "StepNum", "F1", "F2", "F3", "M1", "M2", "M3"]
            results = [NumberResults,
                       SCut,
                       LoadCase,
                       StepType,
                       StepNum,
                       F1,
                       F2,
                       F3,
                       M1,
                       M2,
                       M3]
            outputs.append(results)

    output_writer(sheet_name, titles, outputs, nth_iteration, k)
def SectionCutDesign(nth_iteration, k):
    sheet_name = 'SectionCutDesign'
    outputs = []

    for i in range(0, len(joints)):
        for j in range(0, len(cases_output)):
            combo = cases_output[j]
            ret = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
            ret = SapModel.Results.Setup.SetCaseSelectedForOutput(combo)
            NumberResults = 0
            SCut = []
            LoadCase = []
            StepType = []
            StepNum = []
            P = []
            V2 = []
            V3 = []
            T = []
            M2 = []
            M3 = []
            [NumberResults,
             SCut,
             LoadCase,
             StepType,
             StepNum,
             P,
             V2,
             V3,
             T,
             M2,
             M3
                , ret] = SapModel.Results.SectionCutDesign(NumberResults,
                                                           SCut,
                                                           LoadCase,
                                                           StepType,
                                                           StepNum,
                                                           P,
                                                           V2,
                                                           V3,
                                                           T,
                                                           M2,
                                                           M3)
            titles = ["NumberResults", "SCut", "LoadCase", "StepType", "StepNum", "P", "V2", "V3", "T", "M2", "M3"]
            results = [NumberResults,
                       SCut,
                       LoadCase,
                       StepType,
                       StepNum,
                       P,
                       V2,
                       V3,
                       T,
                       M2,
                       M3]
            outputs.append(results)
    for i in range(0, len(joints)):
        for j in range(0, len(combos_output)):
            combo = combos_output[j]
            ret = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
            ret = SapModel.Results.Setup.SetComboSelectedForOutput(combo)
            NumberResults = 0
            SCut = []
            LoadCase = []
            StepType = []
            StepNum = []
            P = []
            V2 = []
            V3 = []
            T = []
            M2 = []
            M3 = []
            [NumberResults,
             SCut,
             LoadCase,
             StepType,
             StepNum,
             P,
             V2,
             V3,
             T,
             M2,
             M3
                , ret] = SapModel.Results.SectionCutDesign(NumberResults,
                                                           SCut,
                                                           LoadCase,
                                                           StepType,
                                                           StepNum,
                                                           P,
                                                           V2,
                                                           V3,
                                                           T,
                                                           M2,
                                                           M3)
            titles = ["NumberResults", "SCut", "LoadCase", "StepType", "StepNum", "P", "V2", "V3", "T", "M2", "M3"]
            results = [NumberResults,
                       SCut,
                       LoadCase,
                       StepType,
                       StepNum,
                       P,
                       V2,
                       V3,
                       T,
                       M2,
                       M3]
            outputs.append(results)

    output_writer(sheet_name, titles, outputs, nth_iteration, k)
def JointDispl(nth_iteration, k):
    sheet_name = 'JointDispl'
    outputs = []
    for j in range(0, len(cases_output)):
        combo = cases_output[j]
        ret = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
        ret = SapModel.Results.Setup.SetCaseSelectedForOutput(combo)
        ret = SapModel.Results.Setup.SetOptionModalHist(2)
        ItemTypeElm = 2
        NumberResults = 0
        Obj = []
        Elm = []
        LoadCase = []
        StepType = []
        StepNum = []
        U1 = []
        U2 = []
        U3 = []
        R1 = []
        R2 = []
        R3 = []

        [NumberResults,
         Obj,
         Elm,
         LoadCase,
         StepType,
         StepNum,
         U1,
         U2,
         U3,
         R1,
         R2,
         R3, ret] = SapModel.Results.JointDispl("ALL",
                                                ItemTypeElm,
                                                NumberResults,
                                                Obj,
                                                Elm,
                                                LoadCase,
                                                StepType,
                                                StepNum,
                                                U1,
                                                U2,
                                                U3,
                                                R1,
                                                R2,
                                                R3)
        titles = ["NumberResults",
                  "Obj",
                  "Elm",
                  "LoadCase",
                  "StepType",
                  "StepNum",
                  "U1",
                  "U2",
                  "U3",
                  "R1",
                  "R2",
                  "R3"]
        results = [NumberResults,
                   Obj,
                   Elm,
                   LoadCase,
                   StepType,
                   StepNum,
                   U1,
                   U2,
                   U3,
                   R1,
                   R2,
                   R3]
        outputs.append(results)
    for j in range(0, len(combos_output)):
        combo = combos_output[j]
        ret = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
        ret = SapModel.Results.Setup.SetComboSelectedForOutput(combo)
        ret = SapModel.Results.Setup.SetOptionModalHist(2)
        ItemTypeElm = 2
        NumberResults = 0
        Obj = []
        Elm = []
        LoadCase = []
        StepType = []
        StepNum = []
        U1 = []
        U2 = []
        U3 = []
        R1 = []
        R2 = []
        R3 = []

        [NumberResults,
         Obj,
         Elm,
         LoadCase,
         StepType,
         StepNum,
         U1,
         U2,
         U3,
         R1,
         R2,
         R3, ret] = SapModel.Results.JointDispl("ALL",
                                                ItemTypeElm,
                                                NumberResults,
                                                Obj,
                                                Elm,
                                                LoadCase,
                                                StepType,
                                                StepNum,
                                                U1,
                                                U2,
                                                U3,
                                                R1,
                                                R2,
                                                R3)
        titles = ["NumberResults",
                  "Obj",
                  "Elm",
                  "LoadCase",
                  "StepType",
                  "StepNum",
                  "U1",
                  "U2",
                  "U3",
                  "R1",
                  "R2",
                  "R3"]
        results = [NumberResults,
                   Obj,
                   Elm,
                   LoadCase,
                   StepType,
                   StepNum,
                   U1,
                   U2,
                   U3,
                   R1,
                   R2,
                   R3]
        outputs.append(results)

    output_writer(sheet_name, titles, outputs, nth_iteration, k)
def JointDisplAbs(nth_iteration, k):
    sheet_name = 'JointDisplAbs'
    outputs = []
    for j in range(0, len(cases_output)):
        combo = cases_output[j]
        ret = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
        ret = SapModel.Results.Setup.SetCaseSelectedForOutput(combo)
        Name = "ALL"
        ItemTypeElm = 2
        NumberResults = 0
        Obj = []
        Elm = []
        LoadCase = []
        StepType = []
        StepNum = []
        U1 = []
        U2 = []
        U3 = []
        R1 = []
        R2 = []
        R3 = []

        [NumberResults,
         Obj,
         Elm,
         LoadCase,
         StepType,
         StepNum,
         U1,
         U2,
         U3,
         R1,
         R2,
         R3, ret] = SapModel.Results.JointDisplAbs(Name,
                                                   ItemTypeElm,
                                                   NumberResults,
                                                   Obj,
                                                   Elm,
                                                   LoadCase,
                                                   StepType,
                                                   StepNum,
                                                   U1,
                                                   U2,
                                                   U3,
                                                   R1,
                                                   R2,
                                                   R3)
        titles = ["NumberResults",
                  "Obj",
                  "Elm",
                  "LoadCase",
                  "StepType",
                  "StepNum",
                  "U1",
                  "U2",
                  "U3",
                  "R1",
                  "R2",
                  "R3"]
        results = [NumberResults,
                   Obj,
                   Elm,
                   LoadCase,
                   StepType,
                   StepNum,
                   U1,
                   U2,
                   U3,
                   R1,
                   R2,
                   R3]
        outputs.append(results)
    for j in range(0, len(combos_output)):
        combo = combos_output[j]
        ret = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
        ret = SapModel.Results.Setup.SetComboSelectedForOutput(combo)
        Name = "ALL"
        ItemTypeElm = 2
        NumberResults = 0
        Obj = []
        Elm = []
        LoadCase = []
        StepType = []
        StepNum = []
        U1 = []
        U2 = []
        U3 = []
        R1 = []
        R2 = []
        R3 = []

        [NumberResults,
         Obj,
         Elm,
         LoadCase,
         StepType,
         StepNum,
         U1,
         U2,
         U3,
         R1,
         R2,
         R3, ret] = SapModel.Results.JointDisplAbs(Name,
                                                   ItemTypeElm,
                                                   NumberResults,
                                                   Obj,
                                                   Elm,
                                                   LoadCase,
                                                   StepType,
                                                   StepNum,
                                                   U1,
                                                   U2,
                                                   U3,
                                                   R1,
                                                   R2,
                                                   R3)
        titles = ["NumberResults",
                  "Obj",
                  "Elm",
                  "LoadCase",
                  "StepType",
                  "StepNum",
                  "U1",
                  "U2",
                  "U3",
                  "R1",
                  "R2",
                  "R3"]
        results = [NumberResults,
                   Obj,
                   Elm,
                   LoadCase,
                   StepType,
                   StepNum,
                   U1,
                   U2,
                   U3,
                   R1,
                   R2,
                   R3]
        outputs.append(results)

    output_writer(sheet_name, titles, outputs, nth_iteration, k)
def AssembledJointMass(nth_iteration, k):
    sheet_name = 'AssembledJointMass'
    outputs = []
    for j in range(0, len(cases_output)):
        combo = cases_output[j]
        ret = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
        ret = SapModel.Results.Setup.SetCaseSelectedForOutput(combo)
        Name = "ALL"
        ItemTypeElm = 2
        NumberResults = 0
        PointElm = []
        U1 = []
        U2 = []
        U3 = []
        R1 = []
        R2 = []
        R3 = []

        [NumberResults,
         PointElm,
         U1,
         U2,
         U3,
         R1,
         R2,
         R3, ret] = SapModel.Results.AssembledJointMass(Name,
                                                        ItemTypeElm,
                                                        NumberResults,
                                                        PointElm,
                                                        U1,
                                                        U2,
                                                        U3,
                                                        R1,
                                                        R2,
                                                        R3)
        titles = ["NumberResults",
                  "PointElm",
                  "U1",
                  "U2",
                  "U3",
                  "R1",
                  "R2",
                  "R3"]
        results = [NumberResults,
                   PointElm,
                   U1,
                   U2,
                   U3,
                   R1,
                   R2,
                   R3]
        outputs.append(results)
    for j in range(0, len(combos_output)):
        combo = combos_output[j]
        ret = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
        ret = SapModel.Results.Setup.SetComboSelectedForOutput(combo)
        Name = "ALL"
        ItemTypeElm = 2
        NumberResults = 0
        PointElm = []
        U1 = []
        U2 = []
        U3 = []
        R1 = []
        R2 = []
        R3 = []

        [NumberResults,
         PointElm,
         U1,
         U2,
         U3,
         R1,
         R2,
         R3, ret] = SapModel.Results.AssembledJointMass(Name,
                                                        ItemTypeElm,
                                                        NumberResults,
                                                        PointElm,
                                                        U1,
                                                        U2,
                                                        U3,
                                                        R1,
                                                        R2,
                                                        R3)
        titles = ["NumberResults",
                  "PointElm",
                  "U1",
                  "U2",
                  "U3",
                  "R1",
                  "R2",
                  "R3"]
        results = [NumberResults,
                   PointElm,
                   U1,
                   U2,
                   U3,
                   R1,
                   R2,
                   R3]
        outputs.append(results)

    output_writer(sheet_name, titles, outputs, nth_iteration, k)
def GeneralizedDispl(nth_iteration, k):
    sheet_name = 'GeneralizedDispl'
    outputs = []
    for i in range(0, len(joints)):
        for j in range(0, len(cases_output)):
            combo = cases_output[j]
            ret = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
            ret = SapModel.Results.Setup.SetCaseSelectedForOutput(combo)

            ret = SapModel.GDispl.Add("GD1", 1)
            # SF = [None for _ in range(6)]
            # SF[0] = 0.5
            SF = [0.5]
            ret = SapModel.GDispl.SetPoint(f"GD {i + 1}", joints[i], SF)

            Name = f"GD {i + 1}"
            NumberResults = 0
            GD = []
            LoadCase = []
            StepType = []
            StepNum = []
            DType = []
            Value = []

            [NumberResults,
             GD,
             LoadCase,
             StepType,
             StepNum,
             DType,
             Value, ret] = SapModel.Results.GeneralizedDispl(Name,
                                                             NumberResults,
                                                             GD,
                                                             LoadCase,
                                                             StepType,
                                                             StepNum,
                                                             DType,
                                                             Value)
            titles = ["NumberResults",
                      "GD",
                      "LoadCase",
                      "StepType",
                      "StepNum",
                      'DType',
                      "Value"]
            results = [NumberResults,
                       GD,
                       LoadCase,
                       StepType,
                       StepNum,
                       DType,
                       Value]
            outputs.append(results)
    for i in range(0, len(joints)):
        for j in range(0, len(combos_output)):
            combo = combos_output[j]
            ret = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
            ret = SapModel.Results.Setup.SetComboSelectedForOutput(combo)

            ret = SapModel.GDispl.Add("GD1", 1)
            # SF = [None for _ in range(6)]
            # SF[0] = 0.5
            SF = [0.5]
            ret = SapModel.GDispl.SetPoint(f"GD {i + 1}", joints[i], SF)

            Name = f"GD {i + 1}"
            NumberResults = 0
            GD = []
            LoadCase = []
            StepType = []
            StepNum = []
            DType = []
            Value = []

            [NumberResults,
             GD,
             LoadCase,
             StepType,
             StepNum,
             DType,
             Value, ret] = SapModel.Results.GeneralizedDispl(Name,
                                                             NumberResults,
                                                             GD,
                                                             LoadCase,
                                                             StepType,
                                                             StepNum,
                                                             DType,
                                                             Value)
            titles = ["NumberResults",
                      "GD",
                      "LoadCase",
                      "StepType",
                      "StepNum",
                      'DType',
                      "Value"]
            results = [NumberResults,
                       GD,
                       LoadCase,
                       StepType,
                       StepNum,
                       DType,
                       Value]
            outputs.append(results)

    output_writer(sheet_name, titles, outputs, nth_iteration, k)
def GetTableForDisplayArray(nth_iteration, k):
    def tables_writer(sheet_name, titles, outputs, nth_iteration=1, k=0, special_case=0, h_index = 0, heading= ""):
        global f_result
        global wb
        global ws
        os.chdir(exceldst)

        if h_index == 0:
            try:
                wb = load_workbook(f_result)
            except:
                wb = Workbook()
        if special_case == 0:
            try:
                ws = wb[sheet_name]
            except:
                ws = wb.create_sheet("Sheet_1")
                ws.title = sheet_name
                ws.merge_cells('F2:M3')
                ws.cell(row=2, column=6).value = heading
                ws_h2 = ws['F2']
                ws_h2.font = Font(size=23, underline='single', color='FFBB00', bold=True, italic=False)

            row = ws.max_row
            # Titles Writer
            for i in range(0, len(titles)):
                ws.cell(row=5, column=i + 1).value = titles[i]
                ws.cell(row=5, column=i + 1).fill = PatternFill(start_color='32CD32',
                                                                 end_color='32CD32', fill_type="solid")

            # Results Writer
            for h in range(0, len(outputs)):
                for column in range(1, len(titles) + 1):
                    try:
                        cell_value = float(outputs[h][column - 1])
                    except:
                        cell_value = str(outputs[h][column - 1])
                    if cell_value == "None":
                        cell_value = ''

                    ws.cell(row=row + h + 3, column=column).value = cell_value

            ws.freeze_panes = ws['A6']
            wb.save(filename=f_result)
            if h_index == (len(Table_Keys) - 1):
                wb.close()

    ret = SapModel.SelectObj.Group("ALL")

    for h in range(0, len(Table_Keys)):
        sheet_name  = Table_Keys[h]
        TableKey = sheet_name
        FieldKeyList = []
        GroupName = "ALL"
        TableVersion = 0
        FieldsKeysIncluded = []
        NumberRecords = 0
        TableData = []
        [FieldKeyList, TableVersion, FieldsKeysIncluded, NumberRecords, TableData,
         ret] = SapModel.DatabaseTables.GetTableForDisplayArray(TableKey,
                                                                FieldKeyList, GroupName, TableVersion,
                                                                FieldsKeysIncluded, NumberRecords, TableData)
        titles = [x for x in FieldsKeysIncluded]
        title_length = len(FieldsKeysIncluded)
        titles.insert(0, "Model Number")
        titles.insert(1, "Angle Deviation")

        outputs = []
        data = []
        for i in range(1, len(TableData) + 1):
            data.append(TableData[i - 1])

            if i % title_length == 0:
                data.insert(0, nth_iteration)
                data.insert(1, angles[nth_iteration - 1])

                outputs.append(data)
                data = []

        heading = sheet_name
        while len(sheet_name) > 30:
            sheet_name = " ".join(sheet_name.split(" ")[0: - 1])

        tables_writer(sheet_name,titles, outputs,nth_iteration, k, h_index=h, heading = heading)
        print(f'{Table_Keys[h]} has successfully been written to document')

def Results_Handler(nth_iteration, angles, i = 0):
        # BaseReact(nth_iteration, angles[nth_iteration-1])
        # StoryDrifts(nth_iteration, angles[nth_iteration-1])
        # BaseReactWithCentroid(nth_iteration, angles[nth_iteration-1], i)
        # ModalLoadParticipationRatios(nth_iteration, angles[nth_iteration-1])
        # ModalParticipatingMassRatios(nth_iteration, angles[nth_iteration-1])
        # ModalPeriod(nth_iteration, angles[nth_iteration-1])
        # ModeShape(nth_iteration, angles[nth_iteration-1])
        # BucklingFactor(nth_iteration, angles[nth_iteration-1])
        # FrameForce(nth_iteration, angles[nth_iteration-1])
        # FrameJointForce(nth_iteration, angles[nth_iteration-1])
        # JointAcc(nth_iteration, angles[nth_iteration-1])
        # JointAccAbs(nth_iteration, angles[nth_iteration-1])
        # JointDrifts(nth_iteration, angles[nth_iteration-1])
        # JointReact(nth_iteration, angles[nth_iteration-1])
        # JointVel(nth_iteration, angles[nth_iteration-1])
        # JointVelAbs(nth_iteration, angles[nth_iteration-1])
        # JointDispl(nth_iteration, angles[nth_iteration-1])
        # JointDisplAbs (nth_iteration, angles[nth_iteration-1])
        # AssembledJointMass  (nth_iteration, angles[nth_iteration-1])
        # GeneralizedDispl(nth_iteration, angles[nth_iteration-1])
        # SectionCutAnalysis(nth_iteration, angles[nth_iteration-1])
        # SectionCutDesign(nth_iteration, angles[nth_iteration-1])

        GetTableForDisplayArray(nth_iteration, angles[nth_iteration-1])




#___________________________________________________________________________________________________________
def Definition_Center(i = 0, angle = 0):
    global wb
    global ws
    wb = load_workbook(f_name, data_only=True)
    ws = wb["Modelling"]

    model_initializer()
    if predefined_model == False:
        material_definition()
        section_definition()
        loadpattern_definition()
        loadcase_definition(i)
        loadcombo_definition()
    else:
        combos_input = []
        # General Combo
        for i in range(lcRow, lcRow + 20):
            if ws.cell(row=i, column=2).value != None:
                a = ws.cell(row=i, column=3).value
                combos_input.append(a)

        ####### Combos_input for the Codes
        if ws.cell(row=gdRow + 8, column=4).value == "IS":
            for i in range(ilcRow, ilcRow + 57):
                if ws.cell(row=i, column=2).value != None:
                    a = ws.cell(row=i, column=3).value
                    combos_input.append(a)
        else:
            for i in range(nlcRow, nlcRow + 22):
                if ws.cell(row=i, column=2).value != None:
                    a = ws.cell(row=i, column=3).value
                    combos_input.append(a)

    ResponseSpectrum_Modifier(angle)
    modelling_preparation()  ##### #Skipping This Steps Only halts all activity under this File
def ModellingAndResults_Center(nth_iteration, angles, i = 0):
        excelmodifier_modellinginitializor(angles[nth_iteration - 1])
        global ws
        wb = load_workbook(f_name, data_only=True)  # If data only is false it shows the formula of the cell
        ws = wb["Modelling"]

        SapModel.SetPresentUnits(modelling_unit)

        if CoOrdinatesFrom_Excel:
            beamcolumnslab_modelling()
        else:
            Co_Ordinate_Computer(angles[nth_iteration - 1])

        section_assigner()
        restraint_assignment()
        load_application()
        diaphragm_assignment()
        # run_analysis(angles[nth_iteration - 1])
        # Results_Handler(nth_iteration, angles, i)
system_adjuster()
#___________________________________________________________________________________________________________

# # #######________________Calling Zone_
# if Critical_Angle_Determination:
#     global Calculation_Option
#     global max_Angle
#
#     if Calculation_Option == 4:
#         for nth_iteration in range(1, len(angles) + 1):
#             for step in range(1, 4):
#                 max_Angle = 0
#                 Calculation_Option = step
#                 for i in range(0, 2):
#                     Definition_Center(i, angles[nth_iteration-1])
#                     ModellingAndResults_Center(nth_iteration, angles, i)
#     else:
#         for nth_iteration in range(1, len(angles) + 1):
#             max_Angle = 0
#             for i in range(0, 2):
#                 Definition_Center(i, angles[nth_iteration - 1])
#                 ModellingAndResults_Center(nth_iteration, angles, i)
#
#
#
# else:
#     for nth_iteration in range(1, len(angles) + 1):
#         Definition_Center(angle = angles[nth_iteration - 1])
#         ModellingAndResults_Center(nth_iteration, angles)


# Response_correction = False
# if Response_correction:
#     if len(Earthquake_Casesx) > 0 :
#         looper = len(Earthquake_Casesx)
#     elif len(Earthquake_Casesy) > 0 :
#         looper = len(Earthquake_Casesy)
#     else:
#         print("No Set of ESM Load case found for correction")
#
#     # try:
#     for i in range(0, looper):
#             for k in range(1, len(angles) + 1):
#                 #X SF Correction
#                 earthquake_casex = Earthquake_Casesx[i]
#                 response_casex = Response_Casesx[i]
#                 f_result = "Result.xlsx"
#                 wb = load_workbook(f_result, data_only=True)
#                 try:
#                     ws = wb["BaseReactWithCentroid"]
#                     indexer(wb, "BaseReactWithCentroid", "FX")
#
#                 except:
#                     ws = wb["Base Reactions"]
#                     indexer(wb, "Base Reactions", "FX")
#
#                 angle = angles[k - 1]
#                 earthquake_valuex = 0
#                 response_valuex = 0
#
#                 for l in range(block_start_row_i[k - 1], block_end_row_i[k - 1]):
#
#                     if ws.cell(row=l, column=Pre_Processings.load_case_col).value == earthquake_casex:
#                         earthquake_valuex = float(
#                             ws.cell(row=l, column=Pre_Processings.start_column_i).value)
#                     if ws.cell(row=l, column=Pre_Processings.load_case_col).value == response_casex:
#                         response_valuex = float(ws.cell(row=l, column=Pre_Processings.start_column_i).value)
#
#                 if earthquake_valuex != 0 and response_valuex != 0 :
#                     Resp_Spectrum_SFx = [abs((earthquake_valuex / response_valuex) * original_RS_SFx)]
#                 else:
#                     Resp_Spectrum_SFx = original_RS_SFx
#
#
#                 #Y SF Correction
#                 earthquake_casey = Earthquake_Casesy[i]
#                 response_casey = Response_Casesy[i]
#                 earthquake_valuey = 0
#                 response_valuey = 0
#                 try:
#                     indexer(wb, "BaseReactWithCentroid", "FY")
#
#                 except:
#                     indexer(wb, "Base Reactions", "FY")
#                 for l in range(block_start_row_i[k - 1], block_end_row_i[k - 1]):
#                     if ws.cell(row=l, column=Pre_Processings.load_case_col).value == earthquake_casey:
#                         earthquake_valuey = float(
#                             ws.cell(row=l, column=Pre_Processings.start_column_i).value)
#                     if ws.cell(row=l, column=Pre_Processings.load_case_col).value == response_casey:
#                         response_valuey = float(ws.cell(row=l, column=Pre_Processings.start_column_i).value)
#
#                 if earthquake_valuey != 0 and response_valuey != 0 :
#                     Resp_Spectrum_SFy = [abs((earthquake_valuey / response_valuey) * original_RS_SFy)]
#                 else:
#                     Resp_Spectrum_SFy = original_RS_SFy
#
#
#                 wb.close()
#
#                 print(earthquake_valuex, response_valuex, Resp_Spectrum_SFx, "Resp Spec", k)
#                 print(earthquake_valuey, response_valuey, Resp_Spectrum_SFy, "Resp Spec", k)
#
#                 # # #Calling Zone
#                 f_result = f'Result {earthquake_casex}-{response_casex}.xlsx'
#                 Definition_Center()
#                 ModellingAndResults_Center(k, angles)
#                 print(f"Angle {angles[k - 1]} writing has successfully been completed, Thank You")
#     # except:
#     #     print("TRY Statement hasnot been executed due to internal error, Check Either CASEs, Scale Factor Provided, Deviation Angles")
#
# print(f'ETABS.py is running')
