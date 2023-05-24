import math
import os
import shutil
import sys
import comtypes.client
import openpyxl.chart
from openpyxl import *
from openpyxl.styles import PatternFill  # Connect cell styles
from openpyxl.styles import Font, Fill, Color  # Connect styles for text
from openpyxl.styles import colors
import xlwings as xw
def bool_converter(input):
    if input == "True" or "TRUE" or "true":
        return (True)
    else:
        return (False)

#****************************************************
Titles = ["P", "V2", "V3", "T", "M2", "M3"] #Donot modify
elements = [1,2] #One for the Absoulute Value and One for percentage difference type Column group
data_column = {"S.N.": 1, "Element": 2, "Combo": 3, "Angle": 4, "Parameter1": 5,
               "Start": 5, "SPD": 6, "Centre": 7, "CPD": 8, "End": 9, "EPD": 10, "8": "P",
               "9": "V2", "10": "V3", "11": "T", "12": "M2", "13": "M3"}  #
data_column_BaseReaction = {"S.N.": 1, "Combo": 2, "Angle": 3, "Number of Storey": 4, "Result's Point Location":5, "Fx": 6,
               "Fy": 7, "Fz": 8, "Mx": 9, "My": 10, "Mz": 11}
properties_ = ["Fx", "Fy", "Fz", "Mx", "My", "Mz"]
########### Extract the required data from here ["P", "V2", "V3", "T", "M2", "M3"] 8,9,10,11,12,13
create_chart = True
parameter = [0]  # Provide only one value inside it in this phase


#****************************************************"""Data Extraction For the Initialization of Code"""
os.chdir(os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop'))
wb = load_workbook("Design Sheet.xlsx",data_only=True) #If data only is false it shows the formula of the cell
ws = wb["Modelling"]
giRow = ws.cell(row=3, column=12).value
gdRow = ws.cell(row=4, column=12).value
mdRow = ws.cell(row=5, column=12).value
fsdRow = ws.cell(row=6, column=12).value
lpRow = ws.cell(row=7, column=12).value
lcRow = ws.cell(row=8, column=12).value
ilcRow = ws.cell(row=9, column=12).value
nlcRow = ws.cell(row=10, column=12).value
fjRow = ws.cell(row=11, column=12).value
laRow = ws.cell(row=12, column=12).value
instance = bool_converter(ws.cell(row=3, column=10).value)
lockmodel = bool_converter(ws.cell(row=5, column=10).value)
defining_unit = ws.cell(row=gdRow + 6, column=5).value
modelling_unit = ws.cell(row=gdRow + 7, column=5).value
starting_angle = ws.cell(row=gdRow + 5, column=6).value
increment_angle = ws.cell(row=gdRow + 5, column=8).value
ending_angle = ws.cell(row=gdRow + 5, column=10).value
user_name = ws.cell(row=giRow, column=4).value
chart_position = bool_converter(ws.cell(row=7, column=10).value)
number_of_stories = ws.cell(row=giRow + 10, column=4).value
#****************************************************
# #----------------------------------------------Creating Directory and exporting Excel---------------------------------
main_folder = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Documents', 'ETABS Design Report')
ResearchFolder = os.path.join(main_folder,user_name, "ETABS", "Research Folder")
excelscr = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop','Design Sheet.xlsx')
exceldst = main_folder + "\\" + user_name
etabspath = os.path.join(exceldst, "ETABS")
if os.path.exists(ResearchFolder) is False:
    os.makedirs(ResearchFolder)

shutil.copy(excelscr, exceldst)  #Never Open excel sheet while running else this code doesnot run
ModelPath = etabspath + os.sep + 'Design File.edb'
ExcelPath = exceldst + os.sep + "Design Sheet.xlsx"
FilePath = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop',"ETABS File","Model - Seismic.EDB")
os.chdir(exceldst)
# # # --------------------------------------------Initializing the ETABS-----------------------------------------------
# #______________________________________________________________________________________________________
#Use this code after ETABS is run and new model is created in metric.
#clean up variables
SapModel = None
EtabsObject = None
AttachToInstance = 1
if AttachToInstance:
    try:
        myETABSObject = comtypes.client.GetActiveObject("CSI.ETABS.API.ETABSObject")
    except (OSError, comtypes.COMError):
        print("No running instance of the program found or failed to attach.")
        sys.exit(-1)
SapModel = myETABSObject.SapModel
ret = SapModel.File.OpenFile(FilePath)  #******** Just Open Etabs And No need to create NEW Model for opening file
# ------------------------------------------------------Materials Definition------------------------------------------
# Set present Units
ret = SapModel.SetPresentUnits(defining_unit)
# Concrete Material Definition
row = mdRow - 3
col = 4
if ws.cell(row=row, column=col).value == "Auto Defining":
    a = ws.cell(row=mdRow, column=3).value
    b = int(ws.cell(row=mdRow, column=4).value)
    c = ws.cell(row=mdRow, column=5).value
    d = ws.cell(row=mdRow, column=6).value
    e = ws.cell(row=mdRow, column=7).value
    ret = SapModel.PropMaterial.AddMaterial(a, b, c, d, e)
else:
    a = ws.cell(row=mdRow + 5, column=3).value
    b = int(ws.cell(row=mdRow + 5, column=4).value)
    c = int(ws.cell(row=mdRow + 5, column=5).value)
    d = bool_converter(ws.cell(row=mdRow + 5, column=6).value)
    e = int(ws.cell(row=mdRow + 5, column=7).value)
    f = int(ws.cell(row=mdRow + 5, column=8).value)
    g = int(ws.cell(row=mdRow + 5, column=9).value)
    h = int(ws.cell(row=mdRow + 5, column=10).value)
    i = int(ws.cell(row=mdRow + 5, column=11).value)
    j = int(ws.cell(row=mdRow + 5, column=12).value)
    k = int(ws.cell(row=mdRow + 5, column=13).value)
    l = int(ws.cell(row=mdRow + 5, column=14).value)
    ret = SapModel.PropMaterial.SetMaterial(a, b)
    ret = SapModel.PropMaterial.SetOConcrete_1(a, c, d, e, f, g, h, i, j, k, l)
# Rebar Material Definition
row = mdRow + 7
col = 4
if ws.cell(row=row, column=col).value == "Auto Defining":
    a = ws.cell(row=mdRow + 10, column=3).value
    b = int(ws.cell(row=mdRow + 10, column=4).value)
    c = ws.cell(row=mdRow + 10, column=5).value
    d = ws.cell(row=mdRow + 10, column=6).value
    e = ws.cell(row=mdRow + 10, column=7).value
    ret = SapModel.PropMaterial.AddMaterial(a, b, c, d, e)
else:
    a = ws.cell(row=mdRow + 15, column=3).value
    b = int(ws.cell(row=mdRow + 15, column=4).value)
    c = int(ws.cell(row=mdRow + 15, column=5).value)
    d = int(ws.cell(row=mdRow + 15, column=6).value)
    e = int(ws.cell(row=mdRow + 15, column=7).value)
    f = int(ws.cell(row=mdRow + 15, column=8).value)
    g = int(ws.cell(row=mdRow + 15, column=9).value)
    h = int(ws.cell(row=mdRow + 15, column=10).value)
    i = int(ws.cell(row=mdRow + 15, column=11).value)
    j = int(ws.cell(row=mdRow + 15, column=12).value)
    k = bool_converter(ws.cell(row=mdRow + 15, column=13).value)
    l = int(ws.cell(row=mdRow + 15, column=14).value)
    ret = SapModel.PropMaterial.SetMaterial(a, b)
    ret = SapModel.PropMaterial.SetORebar(a, c, d, e, f, g, h, i, j, k, l)
# -------------------------------------------------Section Definition------------------------------------------
# Column Section Definition
if ws.cell(row=fsdRow - 3, column=5).value == "Rectangular":
    a = ws.cell(row=fsdRow, column=3).value
    b = ws.cell(row=fsdRow, column=4).value
    c = int(ws.cell(row=fsdRow, column=5).value)
    d = int(ws.cell(row=fsdRow, column=6).value)
    e = ws.cell(row=fsdRow, column=7).value
    f = ws.cell(row=fsdRow, column=8).value
    g = int(ws.cell(row=fsdRow, column=9).value)
    h = int(ws.cell(row=fsdRow, column=10).value)
    i = int(ws.cell(row=fsdRow, column=11).value)
    j = int(ws.cell(row=fsdRow, column=12).value)
    k = int(ws.cell(row=fsdRow, column=13).value)
    l = int(ws.cell(row=fsdRow, column=14).value)
    m = str(ws.cell(row=fsdRow, column=15).value)
    n = str(ws.cell(row=fsdRow, column=16).value)
    o = int(ws.cell(row=fsdRow, column=17).value)
    p = int(ws.cell(row=fsdRow, column=18).value)
    q = int(ws.cell(row=fsdRow, column=19).value)
    r = bool_converter(int(ws.cell(row=fsdRow, column=20).value))
    ret = SapModel.PropFrame.SetRectangle(a, b, c, d)
    ret = SapModel.PropFrame.SetRebarColumn(a, e, f, g, h, i, j, k, l, m, n, o, p, q, r)
# Beam Section Definition
if ws.cell(row=fsdRow + 2, column=5).value == "Rectangular":
    a = ws.cell(row=fsdRow + 5, column=3).value
    b = ws.cell(row=fsdRow + 5, column=4).value
    c = int(ws.cell(row=fsdRow + 5, column=5).value)
    d = int(ws.cell(row=fsdRow + 5, column=6).value)
    e = ws.cell(row=fsdRow + 5, column=7).value
    f = ws.cell(row=fsdRow + 5, column=8).value
    g = int(ws.cell(row=fsdRow + 5, column=9).value)
    h = int(ws.cell(row=fsdRow + 5, column=10).value)
    i = int(ws.cell(row=fsdRow + 5, column=11).value)
    j = int(ws.cell(row=fsdRow + 5, column=12).value)
    k = int(ws.cell(row=fsdRow + 5, column=13).value)
    l = int(ws.cell(row=fsdRow + 5, column=14).value)
    ret = SapModel.PropFrame.SetRectangle(a, b, c, d)
    ret = SapModel.PropFrame.SetRebarBeam(a, e, f, g, h, i, j, k, l)
    # Slab Section Definition
a = ws.cell(row=fsdRow + 10, column=3).value
b = int(ws.cell(row=fsdRow + 10, column=4).value)
c = int(ws.cell(row=fsdRow + 10, column=5).value)
d = ws.cell(row=fsdRow + 10, column=6).value
e = int(ws.cell(row=fsdRow + 10, column=7).value)
ret = SapModel.PropArea.SetSlab(a, b, c, d, e)

# -------------------------------------------------Load Patterns Definition------------------------------------------
# Load Patterns Cases and Combos are predefined in the model itself.
# Check for the previous phases documents for Extracting the code if needed  (Here the code only fills the list of combos)
#### Creating Cases for the Results
cases = ["EQx ULS", "EQx SLS", "EQy ULS", "EQy SLS"]

combos = []
# General Combo
for i in range(lcRow, lcRow + 20):
    if ws.cell(row=i, column=2).value != None:
        a = ws.cell(row=i, column=3).value
        combos.append(a)

# Combos for the Codes
if ws.cell(row=gdRow + 8, column=4).value == "IS":
    for i in range(ilcRow, ilcRow + 57):
        if ws.cell(row=i, column=2).value != None:
            a = ws.cell(row=i, column=3).value
            combos.append(a)
else:
    for i in range(nlcRow, nlcRow + 22):
        if ws.cell(row=i, column=2).value != None:
            a = ws.cell(row=i, column=3).value
            combos.append(a)

# -------------------------------------------------Saving Model -------------------------------------------------------
ret = SapModel.File.Save(ModelPath)
wb.close()
#Preparation for the iterations
SapModel = None
EtabsObject = None
angles = []
for k in range(starting_angle, ending_angle, increment_angle):
    angles.append(k)
number_of_iterations = int(len(angles))
nth_iteration = 1
spacing_required = int(number_of_iterations + 1)
elm_col_width = 5 + (6 * (len(parameter) + 1))
if create_chart:
    if chart_position is True:
        elm_col_width = 5 + (6 * (len(parameter) + 1))
        spacing_required = int(number_of_iterations + 1) + 15
    else:
        elm_col_width = 5 + (6 * (len(parameter) + 1)) + 9
        if spacing_required < 15:
            spacing_required = 15
if len(parameter) != 1:
    elm_col_width = 9*len(parameter)

##########&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&
# # #*****************************************************Iteration Starts*****************************************
for k in range(starting_angle, ending_angle, increment_angle):
    #************************** Creating Name for the files of model
    experiment_file = ResearchFolder + os.sep + f'Design File {k}.edb'

    # Modify Angle cell and save it by activating related formula
    f_name = "Design Sheet.xlsx"
    wb = xw.Book(f_name)
    ws = wb.sheets["Modelling"]
    ws.range('F30').value = k
    wb.save()
    wb.close()

    # Opens the Values from the modified Cells
    wb = load_workbook(f_name, data_only=True)  # If data only is false it shows the formula of the cell
    ws = wb["Modelling"]

    # clean up variables and Restart the Sapmodel for the file opening
    SapModel = None
    EtabsObject = None
    AttachToInstance = 1
    if AttachToInstance:
        try:
            myETABSObject = comtypes.client.GetActiveObject("CSI.ETABS.API.ETABSObject")
        except (OSError, comtypes.COMError):
            print("No running instance of the program found or failed to attach.")
            sys.exit(-1)
    SapModel = myETABSObject.SapModel
    ret = SapModel.File.OpenFile(ModelPath)
    ret = SapModel.File.Save(experiment_file)
    ret = SapModel.File.OpenFile(experiment_file)

    # #-------------------------------------------------FrameWorks Modelling------------------------------------------
    # Set present Units
    ret = SapModel.SetPresentUnits(modelling_unit)

    baysX = ws.cell(row=gdRow, column=4).value
    baysY = ws.cell(row=gdRow + 1, column=4).value
    storey = ws.cell(row=giRow + 10, column=4).value
    col_section = ws.cell(row=fsdRow, column=3).value
    beam_section = ws.cell(row=fsdRow + 5, column=3).value
    slab_section = ws.cell(row=fsdRow + 10, column=3).value

    # material = ws.cell(row=fsdRow+5, column=4).value
    id_unique = {}  # All kept in managed order of storey as key -> (Columns , BeamsX, BeamsY) as value for storey and key for unique id
    id_fname = {}  # All seperated with element Like beam columns as keys and their frame name (..+"to"+..) as list
    id_uname = {}  # All seperated with element Like beam columns as keys and their unique name (".."+"..") as list
    ids = {}  # All kept in random order
    slab_uname = []  # All kept unique name in list form ... First value indicates storey number
    roofSlab_uname = []
    floorSlab_uname = []
    beams_outer = []
    beams_inner = []
    columns_inner = []
    columns_outer = []
    alphabet = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]
    number = ["1", "2", "3", "4", "5", "6", "7", "8", "9", "10"]
    for h in range(13, (int(storey)) * 13 + 1, 13):
        fjcRow = fjRow + h
        columns_unique = []
        columns_fname = []
        columns_uname = []
        beamsX_unique = []
        beamsX_fname = []
        beamsX_uname = []
        beamsY_unique = []
        beamsY_fname = []
        beamsY_uname = []

        # Columns Along Z Modelling
        for i in range(0, baysY + 1):
            for j in range(3, 4 * baysX + 4, 4):
                p1 = ws.cell(row=fjcRow - 13 + i, column=j).value
                p2 = ws.cell(row=fjcRow, column=j).value
                x1 = (ws.cell(row=fjcRow - 13 + i, column=j + 1).value)
                y1 = (ws.cell(row=fjcRow - 13 + i, column=j + 2).value)
                z1 = (ws.cell(row=fjcRow - 13 + i, column=j + 3).value)
                x2 = (ws.cell(row=fjcRow + i, column=j + 1).value)
                y2 = (ws.cell(row=fjcRow + i, column=j + 2).value)
                z2 = (ws.cell(row=fjcRow + i, column=j + 3).value)
                frame_key = str(p1) + "to" + str(p2)
                unique_namez = str(p1) + str(p2)  # Creates the unique name to the each of the frame element
                columns_unique.append(unique_namez)
                columns_fname.append(frame_key)
                columns_uname.append(unique_namez)
                ids[frame_key] = unique_namez
                [frame_key, ret] = SapModel.FrameObj.AddByCoord(x1, y1, z1, x2, y2, z2, frame_key, col_section,
                                                                unique_namez)
                if ((i == 0 or j == 3) or (i == baysY or j == 4 * baysX + 4 - 1)):
                    columns_outer.append(unique_namez)
                else:
                    columns_inner.append(unique_namez)

        # Beams Along X Modelling
        for i in range(0, baysY + 1):
            for j in range(3, 4 * baysX, 4):
                p1 = ws.cell(row=fjcRow + i, column=j).value
                p2 = ws.cell(row=fjcRow + i, column=j + 4).value
                x1 = (ws.cell(row=fjcRow + i, column=j + 1).value)
                y1 = (ws.cell(row=fjcRow + i, column=j + 2).value)
                z1 = (ws.cell(row=fjcRow + i, column=j + 3).value)
                x2 = (ws.cell(row=fjcRow + i, column=j + 5).value)
                y2 = (ws.cell(row=fjcRow + i, column=j + 6).value)
                z2 = (ws.cell(row=fjcRow + i, column=j + 7).value)
                frame_key = str(p1) + "to" + str(p2)
                unique_namex = str(p1) + str(p2)
                beamsX_unique.append(unique_namex)
                beamsX_fname.append(frame_key)
                beamsX_uname.append(unique_namex)
                ids[frame_key] = unique_namex
                [frame_key, ret] = SapModel.FrameObj.AddByCoord(x1, y1, z1, x2, y2, z2, frame_key, beam_section,
                                                                unique_namex)
                if (i == 0 or i == baysY):
                    beams_outer.append(unique_namex)
                else:
                    beams_inner.append(unique_namex)

        # Beams Along Y Modelling
        for j in range(3, 4 * baysX + 4, 4):
            for i in range(0, baysY):
                p1 = ws.cell(row=fjcRow + i, column=j).value
                p2 = ws.cell(row=fjcRow + i + 1, column=j).value
                x1 = ws.cell(row=fjcRow + i, column=j + 1).value
                y1 = ws.cell(row=fjcRow + i, column=j + 2).value
                z1 = ws.cell(row=fjcRow + i, column=j + 3).value
                x2 = ws.cell(row=fjcRow + i + 1, column=j + 1).value
                y2 = ws.cell(row=fjcRow + i + 1, column=j + 2).value
                z2 = ws.cell(row=fjcRow + i + 1, column=j + 3).value
                frame_key = str(p1) + "to" + str(p2)
                unique_namey = str(p1) + str(p2)
                beamsY_unique.append(unique_namey)
                beamsY_fname.append(frame_key)
                beamsY_uname.append(unique_namey)

                ids[frame_key] = unique_namey
                [frame_key, ret] = SapModel.FrameObj.AddByCoord(x1, y1, z1, x2, y2, z2, frame_key, beam_section,
                                                                unique_namey)
                if (j == 3 or j == 4 * baysX + 4 - 1):
                    beams_outer.append(unique_namey)
                else:
                    beams_inner.append(unique_namey)
        id_unique[f"{int(h / 13)} Columns"] = columns_unique
        id_unique[f"{int(h / 13)} BeamsY"] = beamsX_unique
        id_unique[f"{int(h / 13)} BeamsX"] = beamsY_unique
        id_fname[f"{int(h / 13)} Columns"] = columns_fname
        id_fname[f"{int(h / 13)} BeamsX"] = beamsX_fname
        id_fname[f"{int(h / 13)} BeamsY"] = beamsY_fname
        id_uname[f"{int(h / 13)} Columns"] = columns_uname
        id_uname[f"{int(h / 13)} BeamsX"] = beamsX_uname
        id_uname[f"{int(h / 13)} BeamsY"] = beamsY_uname

        #   #Slab Modelling
        for i in range(0, baysY):  # Along Y Modelling
            for j in range(3, 4 * baysX, 4):  # Along X Modelling
                p1 = (ws.cell(row=fjcRow + i, column=j).value)
                x1 = (ws.cell(row=fjcRow + i, column=j + 1).value)
                y1 = (ws.cell(row=fjcRow + i, column=j + 2).value)
                z1 = (ws.cell(row=fjcRow + i, column=j + 3).value)
                p4 = (ws.cell(row=fjcRow + i, column=j + 4).value)
                x4 = (ws.cell(row=fjcRow + i, column=j + 5).value)
                y4 = (ws.cell(row=fjcRow + i, column=j + 6).value)
                z4 = (ws.cell(row=fjcRow + i, column=j + 7).value)
                p2 = (ws.cell(row=fjcRow + i + 1, column=j).value)
                x2 = (ws.cell(row=fjcRow + i + 1, column=j + 1).value)
                y2 = (ws.cell(row=fjcRow + i + 1, column=j + 2).value)
                z2 = (ws.cell(row=fjcRow + i + 1, column=j + 3).value)
                p3 = (ws.cell(row=fjcRow + i + 1, column=j + 4).value)
                x3 = (ws.cell(row=fjcRow + i + 1, column=j + 5).value)
                y3 = (ws.cell(row=fjcRow + i + 1, column=j + 6).value)
                z3 = (ws.cell(row=fjcRow + i + 1, column=j + 7).value)
                x = [x1, x2, x3, x4]
                y = [y1, y2, y3, y4]
                z = [z1, z2, z3, z4]
                unique_name = str(str(int(h / 13)) + alphabet[int((j - 3) / 4)] + number[i])
                slab_fname = str(p1) + str(p2) + str(p3) + str(p4)
                slab_uname.append(unique_name)
                ret = SapModel.AreaObj.AddByCoord(4, x, y, z, slab_fname, slab_section, unique_name)
                if int(h / 13) != storey:
                    floorSlab_uname.append(unique_name)
                else:
                    roofSlab_uname.append(unique_name)

    # #-------------------------------------------------Restraint Assignment------------------------------------------
    fjcRow = fjRow + 13
    for i in range(0, baysY + 1):
        for j in range(3, 4 * baysX + 4, 4):
            point1 = " "
            point2 = " "
            restraint = [True, True, True, True, True, True]
            p1 = ws.cell(row=fjcRow - 13 + i, column=j).value
            p2 = ws.cell(row=fjcRow, column=j).value
            unique_namez = str(p1) + str(p2)
            [point1, point2, ret] = SapModel.FrameObj.GetPoints(unique_namez, point1, point2)
            ret = SapModel.PointObj.SetRestraint(point1, restraint)



    # -------------------------------------------------Load Application------------------------------------------
    loads = {}
    for i in range(0, 11):
        loads[ws.cell(row=laRow + i, column=3).value] = ws.cell(row=laRow + i, column=4).value

    # Outer Wall Load
    for i in beams_outer:
        ret = SapModel.FrameObj.SetLoadDistributed(i, list(loads.keys())[0], 1, 2, 0, 1,
                                                   -(loads[list(loads.keys())[0]]),
                                                   -(loads[list(loads.keys())[0]]), "Local")
    #
    # #Inner Wall Load
    for i in beams_inner:
        #     ret = SapModel.FrameObj.SetLoadDistributed(list(ids.keys())[list(ids.values()).index(i)],list(loads.keys())[1], 1, 10, 0, 1,
        #                                                (loads[list(loads.keys())[1]]), (loads[list(loads.keys())[1]]))
        ret = SapModel.FrameObj.SetLoadDistributed(i, list(loads.keys())[1], 1, 2,
                                                   0, 1, -(loads[list(loads.keys())[1]]),
                                                   -(loads[list(loads.keys())[1]]), "Local")
    for i in slab_uname:
        Load_name = ws.cell(row=laRow + 3, column=3).value
        Load_value = ws.cell(row=laRow + 3, column=4).value
        ret = SapModel.AreaObj.SetLoadUniform(i, Load_name, (float(Load_value)), 3, True, "Local")
    for i in floorSlab_uname:
        Load_name = ws.cell(row=laRow + 5, column=3).value
        Load_value = ws.cell(row=laRow + 5, column=4).value
        ret = SapModel.AreaObj.SetLoadUniform(i, Load_name, (float(Load_value)), 3, True, "Local")
    for i in floorSlab_uname:
        Load_name = ws.cell(row=laRow + 6, column=3).value
        Load_value = ws.cell(row=laRow + 6, column=4).value
        ret = SapModel.AreaObj.SetLoadUniform(i, Load_name, (float(Load_value)), 3, True, "Local")
    for i in roofSlab_uname:
        Load_name = ws.cell(row=laRow + 7, column=3).value
        Load_value = ws.cell(row=laRow + 7, column=4).value
        ret = SapModel.AreaObj.SetLoadUniform(i, Load_name, (float(Load_value)), 3, True, "Local")

    #Closing the excel of design file as all data are succesfully entered for modelling
    number_of_stories = ws.cell(row=giRow + 10, column=4).value
    wb.close()
#
#
#     # -------------------------------------------------Diaphragm Assignment--------------------------------------------------
    ret = SapModel.Diaphragm.SetDiaphragm("MyDiaphragm1A", SemiRigid := False)
    # -------------------------------------------------Running Analysis--------------------------------------------------
    ret = SapModel.Analyze.RunAnalysis()

    # -------------------------------------------------Displaying Result and Writing in Excel File------------------------------------------
    ####FOR FORCES WRITING +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    # ____________Initializing the Excel
    f_result = "Result.xlsx"
    forces_sheet = "Forces"
    os.chdir(exceldst)
    ResultPath = os.path.join(exceldst, f_result)
    rows_reqn = (len(combos) *(spacing_required)) + 6
    cols_reqn = elm_col_width*(len(elements))

    try:
        wb = load_workbook(f_result)
        ws = wb[forces_sheet]

    except:
        wb = Workbook()
        ws = wb.active
        ws.title = forces_sheet
        items = 0 #Items Loop for data extraction on other column group values
        properties = 0 # Used for extracting various properties
        # ________Occupying the required number of Cells As None
        for m in range(1, int(rows_reqn) + 1):
            for n in range(1, int(cols_reqn) + 1):
                ws.cell(row=m, column=n).value = None
        for items in range(0, len(elements)):

            ws.cell(row=5, column=(items * elm_col_width) + data_column_BaseReaction["S.N."]).value = "S.N."
            ws.cell(row=5, column=(items * elm_col_width) + data_column_BaseReaction["Combo"]).value = "Combo"
            ws.cell(row=5, column=(items * elm_col_width) + data_column_BaseReaction["Angle"]).value = "Angle"
            ws.cell(row=5, column=(items * elm_col_width) + data_column_BaseReaction["Number of Storey"]).value = "Number of Storey"


            ws.cell(row=5, column=(items * elm_col_width) + 6*properties  + data_column_BaseReaction["Fx"]).value = \
            "Base Reactions"
            ws.cell(row=6, column=(items * elm_col_width) + 6*properties + data_column_BaseReaction["Result's Point Location"]).value = "Result's Point Location"
            ws.cell(row=6, column=(items * elm_col_width) + 6*properties + data_column_BaseReaction["Fx"]).value = "Fx"
            ws.cell(row=6, column=(items * elm_col_width) + 6*properties + data_column_BaseReaction["Fy"]).value = "Fy"
            ws.cell(row=6, column=(items * elm_col_width) + 6*properties + data_column_BaseReaction["Fz"]).value = "Fz"
            ws.cell(row=6, column=(items * elm_col_width) + 6*properties + data_column_BaseReaction["Mx"]).value = "Mx"
            ws.cell(row=6, column=(items * elm_col_width) + 6*properties + data_column_BaseReaction["My"]).value = "My"
            ws.cell(row=6, column=(items * elm_col_width) + 6*properties + data_column_BaseReaction["Mz"]).value = "Mz"
        ws.merge_cells('H2:K3')
        ws.cell(row=2, column=8).value = "Results: Base Reactions"
        ws_h2 = ws['H2']  # Created a variable that contains cell A1 with the existing text
        ws_h2.font = Font(size=23, underline='single', color='FFBB00', bold=True,
                          italic=False)  # We apply the following parameters to the text: size - 23, underline, color = FFBB00 (text color is specified in RGB), bold, oblique. If we do not need a bold font, we use the construction: bold = False. We act similarly if we do not need an oblique font: italic = False.
        # ws_h2.font = Font(size=23, underline='single', color = colors.RED, bold=True, italic=True) #what color = colors.RED — color prescribed in styles
        # ws_h2.fill = PatternFill(fill_type='solid', start_color='ff8327', end_color='ff8327')#This code allows you to do design color cells

    # ______________For Forces And Stresses
    for i in range(0, len(elements)):
        for j in range(0,len(cases)):
            combo = cases[j]
            ret = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
            ret = SapModel.Results.Setup.SetCaseSelectedForOutput(combo)

            NumberResults = 0
            LoadCase = []
            StepType = []
            StepNum = []
            Fx = []
            Fy = []
            Fz = []
            Mx= []
            ParamMy = []
            Mz = []
            gx = 0.00
            gy = 0.00
            gz = 0.00
            # ret = SapModel.Results.FrameForce(elements[i], ObjectElm, NumberResults, Obj, Objsta, Elm, Elmsta,
            #                                   LoadCase, StepType, StepNum, P, V2, V3, T, M2, M3)
            ret = SapModel.Results.BaseReact(NumberResults, LoadCase, StepType, StepNum, Fx, Fy, Fz, Mx, ParamMy, Mz,
                                             gx, gy, gz)
            # -------------------Writing Results to Excel
            ws.cell(row=6+nth_iteration+(j*spacing_required), column=(i * elm_col_width) + data_column_BaseReaction["S.N."] ).value = nth_iteration
            ws.cell(row=6+nth_iteration+(j*spacing_required), column=(i * elm_col_width) + data_column_BaseReaction["Combo"]).value = combo
            ws.cell(row=6+nth_iteration+(j*spacing_required), column=(i * elm_col_width) + data_column_BaseReaction["Angle"]).value = k
            ws.cell(row=6+nth_iteration+(j*spacing_required), column=(i * elm_col_width) + data_column_BaseReaction["Number of Storey"]).value = number_of_stories
            point_location = str(gx) + ", " + str(gy) + ", " + str(gz)
            ws.cell(row=6+nth_iteration+(j*spacing_required), column=data_column_BaseReaction["Result's Point Location"]).value = point_location
            if i ==0:
                ws.cell(row=6+nth_iteration+(j*spacing_required), column=(i * elm_col_width) + data_column_BaseReaction["Fx"]).value = ret[4][0]
                ws.cell(row=6+nth_iteration+(j*spacing_required), column=(i * elm_col_width) + data_column_BaseReaction["Fy"]).value = ret[5][0]
                ws.cell(row=6+nth_iteration+(j*spacing_required), column=(i * elm_col_width) + data_column_BaseReaction["Fz"]).value = ret[6][0]
                ws.cell(row=6+nth_iteration+(j*spacing_required), column=(i * elm_col_width) + data_column_BaseReaction["Mx"]).value = ret[7][0]
                ws.cell(row=6+nth_iteration+(j*spacing_required), column=(i * elm_col_width) + data_column_BaseReaction["My"]).value = ret[8][0]
                ws.cell(row=6+nth_iteration+(j*spacing_required), column=(i * elm_col_width) + data_column_BaseReaction["Mz"]).value = ret[9][0]

      # ws.freeze_panes = 'Z7'

    nth_iteration += 1
    wb.save(filename=f_result)
    wb.close()
###########&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&

# # #____________--------Excel File Post Processing------------____________________________________________________-_-
os.chdir(exceldst)
f_result = "Result.xlsx"
forces_sheet  = "Forces"
try:
    wb = load_workbook(f_result)
    ws = wb[forces_sheet]
except:
    print("Excel File Post Processing Error Occured. Check the code")
num_alphabet = {1:"A",2:"B",3:"C",4:"D",5:"E",6:"F",7:"G",8:"H",9:"I",10:"J",11:"K",12:"L",13:"M",14:"N",15:"O",16:"P",17:"Q",18:"R",19:"S",20:"T",21:"U",22:"V",23:"W",24:"X",25:"Y",26:"Z",27:"AA",28:"AB",29:"AC",30:"AD",31:"AE",32:"AF",33:"AG",34:"AH",35:"AI",36:"AJ",37:"AK",38:"AL",39:"AM",40:"AN",41:"AO",42:"AP",43:"AQ",44:"AR",45:"AS",46:"AT",47:"AU",48:"AV",49:"AW",50:"AX",51:"AY",52:"AZ",53:"BA",54:"BB",55:"BC",56:"BD",57:"BE",58:"BF",59:"BG",60:"BH",61:"BI",62:"BJ",63:"BK",64:"BL",65:"BM",66:"BN",67:"BO",68:"BP",69:"BQ",70:"BR",71:"BS",72:"BT",73:"BU",74:"BV",75:"BW",76:"BX",77:"BY",78:"BZ",79:"CA",80:"CB",81:"CC",82:"CD",83:"CE",84:"CF",85:"CG",86:"CH",87:"CI",88:"CJ",89:"CK",90:"CL",91:"CM",92:"CN",93:"CO",94:"CP",95:"CQ",96:"CR",97:"CS",98:"CT",99:"CU",100:"CV",101:"CW",102:"CX",103:"CY",104:"CZ",105:"DA",106:"DB",107:"DC",108:"DD",109:"DE",110:"DF",111:"DG",112:"DH",113:"DI",114:"DJ",115:"DK",116:"DL",117:"DM",118:"DN",119:"DO",120:"DP",121:"DQ",122:"DR",123:"DS",124:"DT",125:"DU",126:"DV",127:"DW",128:"DX",129:"DY",130:"DZ",131:"EA",132:"EB",133:"EC",134:"ED",135:"EE",136:"EF",137:"EG",138:"EH",139:"EI",140:"EJ",141:"EK",142:"EL",143:"EM",144:"EN",145:"EO",146:"EP",147:"EQ",148:"ER",149:"ES",150:"ET",151:"EU",152:"EV",153:"EW",154:"EX",155:"EY",156:"EZ",157:"FA",158:"FB",159:"FC",160:"FD",161:"FE",162:"FF",163:"FG",164:"FH",165:"FI",166:"FJ",167:"FK",168:"FL",169:"FM",170:"FN",171:"FO",172:"FP",173:"FQ",174:"FR",175:"FS",176:"FT",177:"FU",178:"FV",179:"FW",180:"FX",181:"FY",182:"FZ",183:"GA",184:"GB",185:"GC",186:"GD",187:"GE",188:"GF",189:"GG",190:"GH",191:"GI",192:"GJ",193:"GK",194:"GL",195:"GM",196:"GN",197:"GO",198:"GP",199:"GQ",200:"GR",201:"GS",202:"GT",203:"GU",204:"GV",205:"GW",206:"GX",207:"GY",208:"GZ",209:"HA",210:"HB",211:"HC",212:"HD",213:"HE",214:"HF",215:"HG",216:"HH",217:"HI",218:"HJ",219:"HK",220:"HL",221:"HM",222:"HN",223:"HO",224:"HP",225:"HQ",226:"HR",227:"HS",228:"HT",229:"HU",230:"HV",231:"HW",232:"HX",233:"HY",234:"HZ",235:"IA",236:"IB",237:"IC",238:"ID",239:"IE",240:"IF",241:"IG",242:"IH",243:"II",244:"IJ",245:"IK",246:"IL",247:"IM",248:"IN",249:"IO",250:"IP",251:"IQ",252:"IR",253:"IS",254:"IT",255:"IU",256:"IV",257:"IW",258:"IX",259:"IY",260:"IZ",261:"JA",262:"JB",263:"JC",264:"JD",265:"JE",266:"JF",267:"JG",268:"JH",269:"JI",270:"JJ",271:"JK",272:"JL",273:"JM",274:"JN",275:"JO",276:"JP",277:"JQ",278:"JR",279:"JS",280:"JT",281:"JU",282:"JV",283:"JW",284:"JX",285:"JY",286:"JZ",287:"KA",288:"KB",289:"KC",290:"KD",291:"KE",292:"KF",293:"KG",294:"KH",295:"KI",296:"KJ",297:"KK",298:"KL",299:"KM",300:"KN",301:"KO",302:"KP",303:"KQ",304:"KR",305:"KS",306:"KT",307:"KU",308:"KV",309:"KW",310:"KX",311:"KY",312:"KZ",313:"LA",314:"LB",315:"LC",316:"LD",317:"LE",318:"LF",319:"LG",320:"LH",321:"LI",322:"LJ",323:"LK",324:"LL",325:"LM",326:"LN",327:"LO",328:"LP",329:"LQ",330:"LR",331:"LS",332:"LT",333:"LU",334:"LV",335:"LW",336:"LX",337:"LY",338:"LZ",339:"MA",340:"MB",341:"MC",342:"MD",343:"ME",344:"MF",345:"MG",346:"MH",347:"MI",348:"MJ",349:"MK",350:"ML",351:"MM",352:"MN",353:"MO",354:"MP",355:"MQ",356:"MR",357:"MS",358:"MT",359:"MU",360:"MV",361:"MW",362:"MX",363:"MY",364:"MZ",364:"NA",364:"NB",364:"NC",364:"ND",364:"NE",364:"NF",364:"NG",364:"NH",364:"NI",364:"NJ",364:"NK",364:"NL",364:"NM",364:"NN",364:"NO",364:"NP",364:"NQ",364:"NR",364:"NS",364:"NT",364:"NU",364:"NV",364:"NW",364:"NX",364:"NY",364:"NZ"}

############&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&
#*************************************************Table and chart for COMBO data**************************************
# # # ##Percentage Difference Calculations
h = 1 #for h in range(0, len(elements)): For placing the Chart below in second column group
for i in range(0, len(cases)):
    for j in range(1, number_of_iterations):
        for k in range (data_column_BaseReaction["Fx"],data_column_BaseReaction["Mz"] + 1):
            a = ws.cell(row=6 + (i * spacing_required) + j, column= k).value   #Column last value is used for the percentage
            b = ws.cell(row=6 + (i * spacing_required) + j+1, column= k).value       #difference in the cell
            ws.cell(row=6 + (i * spacing_required) + j+1, column=(h *    elm_col_width) + k).value = (b-a) * (100/a)


# #

# # # ###Scatter Line charts Formation (Each Chart plotted for each combo) ---For Absoulute Value
for c in range(0, len(cases)):
    items = 0 #for items in range(0, len(elements)):
    properties = 0  # for properties in range(len(parameter)):
    chart = openpyxl.chart.ScatterChart()
    chart.title = "Base Reactions variation of " + str(number_of_stories) + " story building due to " + str(cases[c])# element i to be defined
    chart.style = 13
    chart.x_axis.title = 'Angle Deviation from Main Orthogonal Axis (In degrees)'
    chart.y_axis.title = 'Percentage Variation'
    # Inputting data of X and Y of Start Centre and End to series for plotting
    xvalues=None
    values = None
    series= None
    xvalues = openpyxl.chart.Reference(ws, min_col=(items * elm_col_width) + data_column_BaseReaction["Angle"], min_row=7 + c*spacing_required, max_row=6 + ((c + 1)*spacing_required) - 1)
    for i in range(data_column_BaseReaction["Fx"], data_column_BaseReaction["Mz"] + 1):
        values = openpyxl.chart.Reference(ws, min_col=((items * elm_col_width) + i + (properties*6)), min_row=7 + c*spacing_required, max_row=6 + ((c + 1)*spacing_required) - 1)
        series = openpyxl.chart.Series(values, xvalues, title_from_data=False)
        chart.series.append(series)
    def anchor_value(items): # for further increase in the element data the value here must be amended based on column Alphabet
        if chart_position is True:
            row_number = str(7+number_of_iterations + c * spacing_required)
            col_number = ((items * elm_col_width) + properties * 9) + 1 #1 is added as A is started from 1 index
        else:
            row_number = str(7 + c*spacing_required +(15*properties))
            col_number = (items * elm_col_width) + (elm_col_width-9)
        value = num_alphabet[col_number] + row_number
        return(value)
    anchor = anchor_value(items)
    ws.add_chart(chart, anchor)
#
# # # # ###Scatter Line charts Formation (Each Chart plotted for each combo) ---For Percentage Variation
for c in range(0, len(cases)):
    items = 1 #for items in range(0, len(elements)):
    properties = 0  # for properties in range(len(parameter)):
    chart = openpyxl.chart.ScatterChart()
    chart.title = "Base Reactions variation of " + str(number_of_stories) + " story building due to " + str(
        cases[c])  # element i to be defined
    chart.style = 13
    chart.x_axis.title = 'Angle Deviation from Main Orthogonal Axis (In degrees)'
    chart.y_axis.title = 'Percentage Variation'
    # Inputting data of X and Y of Start Centre and End to series for plotting
    xvalues = None
    values = None
    series = None
    xvalues = openpyxl.chart.Reference(ws, min_col=(items * elm_col_width) + data_column_BaseReaction["Angle"],
                                       min_row=8 + c * spacing_required,
                                       max_row=6 + ((c + 1) * spacing_required) - 1)
    for i in range(data_column_BaseReaction["Fx"], data_column_BaseReaction["Mz"] + 1):
        values = openpyxl.chart.Reference(ws, min_col=((items * elm_col_width) + i + (properties * 6)),
                                          min_row=8 + c * spacing_required,
                                          max_row=6 + ((c + 1) * spacing_required) - 1)
        series = openpyxl.chart.Series(values, xvalues, title_from_data=False)
        chart.series.append(series)
    def anchor_value(
            items):  # for further increase in the element data the value here must be amended based on column Alphabet
        if chart_position is True:
            row_number = str(7 + number_of_iterations + c * spacing_required)
            col_number = ((items * elm_col_width) + properties * 9) + 1  # 1 is added as A is started from 1 index
        else:
            row_number = str(7 + c * spacing_required + (15 * properties))
            col_number = (items * elm_col_width) + (elm_col_width - 9)
        value = num_alphabet[col_number] + row_number
        return (value)
    anchor = anchor_value(items)
    ws.add_chart(chart, anchor)
##########&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&

#*************************************************Table and chart for combined data**************************************
# ##Defining titles
last_row = 16+ 6 + (len(cases) * spacing_required) # OR starting Row of the last combined table
for items in range(0, len(elements)):
    ws.cell(row=((items-1) * spacing_required) + 15 + last_row + 1, column= data_column_BaseReaction["S.N."]).value = "S.N."
    ws.cell(row=((items-1) * spacing_required) + 15 + last_row + 1, column= data_column_BaseReaction["Combo"]).value = "Combo"
    ws.cell(row=((items-1) * spacing_required) + 15 + last_row + 1, column= data_column_BaseReaction["Angle"]).value = "Angle"
    ws.cell(row=((items-1) * spacing_required) + 15 + last_row + 1, column= data_column_BaseReaction["Number of Storey"]).value = "Number of Storey"
    ws.cell(row=((items-1) * spacing_required) + 15 + last_row + 1, column= data_column_BaseReaction["Result's Point Location"]).value = "Point Location"
    for block in range(0, 6):
        ws.cell(row=((items - 1) * spacing_required) + 15 + last_row + 1,
                column=6 + (len(cases)*block)).value = properties_[block]
####Defining the iterative COnstant Values(Not Data)
i = 1 #for i in range(0, len(elements)):
for items in range(0, len(elements)):
    for nth_iteration in range(1, number_of_iterations+1):
        combo = cases[0]
        # -------------------Writing Results to Excel
        nth_iteration = ws.cell(row=6+nth_iteration+(j*spacing_required), column=(i * elm_col_width) + data_column_BaseReaction["S.N."] ).value
        combo = ws.cell(row=6+nth_iteration+(j*spacing_required), column=(i * elm_col_width) + data_column_BaseReaction["Combo"]).value
        k = ws.cell(row=6+nth_iteration+(j*spacing_required), column=(i * elm_col_width) + data_column_BaseReaction["Angle"]).value
        number_of_stories = ws.cell(row=6+nth_iteration+(j*spacing_required), column=(i * elm_col_width) + data_column_BaseReaction["Number of Storey"]).value
        point_location = ws.cell(row=6 + nth_iteration + (j * spacing_required),
                                 column=data_column_BaseReaction["Result's Point Location"]).value
        ws.cell(row=((items - 1) * spacing_required) + 16 + last_row + nth_iteration,
                column=data_column_BaseReaction["S.N."]).value = nth_iteration
        ws.cell(row=((items - 1) * spacing_required) + 16 + last_row + nth_iteration,
                column=data_column_BaseReaction["Combo"]).value = combo
        ws.cell(row=((items - 1) * spacing_required) + 16 + last_row + nth_iteration,
                column=data_column_BaseReaction["Angle"]).value = k
        ws.cell(row=((items - 1) * spacing_required) + 16 + last_row + nth_iteration,
                column=data_column_BaseReaction["Number of Storey"]).value = number_of_stories
        ws.cell(row=((items - 1) * spacing_required) + 16 + last_row + nth_iteration,
                column=data_column_BaseReaction["Result's Point Location"]).value = point_location

##### Writing the The data in seperate Position
fx_col = data_column_BaseReaction["Fx"]
for i in range(0, len(elements)):
    for j in range(0,len(cases)):
        for nth_iteration in range(1, number_of_iterations + 1):
            combo = cases[j]
            # -------------------Writing Results to Excel
            combo = ws.cell(row=6+nth_iteration+(j*spacing_required), column=(i * elm_col_width) + data_column_BaseReaction["Combo"]).value
            Fx = ws.cell(row=6+nth_iteration+(j*spacing_required), column=(i * elm_col_width) + data_column_BaseReaction["Fx"]).value
            Fy = ws.cell(row=6+nth_iteration+(j*spacing_required), column=(i * elm_col_width) + data_column_BaseReaction["Fy"]).value
            Fz = ws.cell(row=6+nth_iteration+(j*spacing_required), column=(i * elm_col_width) + data_column_BaseReaction["Fz"]).value
            Mx = ws.cell(row=6+nth_iteration+(j*spacing_required), column=(i * elm_col_width) + data_column_BaseReaction["Mx"]).value
            My = ws.cell(row=6+nth_iteration+(j*spacing_required), column=(i * elm_col_width) + data_column_BaseReaction["My"]).value
            Mz = ws.cell(row=6+nth_iteration+(j*spacing_required), column=(i * elm_col_width) + data_column_BaseReaction["Mz"]).value

            ws.cell(row=((i - 1) * spacing_required) + 16 + last_row + nth_iteration,
                    column= fx_col + j + len(cases) * (data_column_BaseReaction["Fx"] - fx_col)).value = Fx
            ws.cell(row=((i - 1) * spacing_required) + 16 + last_row + nth_iteration,
                    column=fx_col + j + len(cases) * (data_column_BaseReaction["Fy"] - fx_col)).value = Fy
            ws.cell(row=((i - 1) * spacing_required) + 16 + last_row + nth_iteration,
                    column=fx_col + j + len(cases) * (data_column_BaseReaction["Fz"] - fx_col)).value = Fz
            ws.cell(row=((i - 1) * spacing_required) + 16 + last_row + nth_iteration,
                    column=fx_col + j + len(cases) * (data_column_BaseReaction["Mx"] - fx_col)).value = Mx
            ws.cell(row=((i - 1) * spacing_required) + 16 + last_row + nth_iteration,
                    column=fx_col + j + len(cases) * (data_column_BaseReaction["My"] - fx_col)).value = My
            ws.cell(row=((i - 1) * spacing_required) + 16 + last_row + nth_iteration,
                    column=fx_col + j + len(cases) * (data_column_BaseReaction["Mz"] - fx_col)).value = Mz

###### Creating Charts for seperate data
for i in range(0, len(elements)):
    for j in range(0, len(properties_)):
        for m in range(0, int(len(cases)/2)):
            chart = openpyxl.chart.ScatterChart()
            chart.style = 5
            chart.x_axis.title = 'Angle Variation'
            chart.y_axis.title = 'Percentage Variation'
            xvalues = None
            values = None
            series = None
            xvalues = openpyxl.chart.Reference(ws, min_col=data_column_BaseReaction["Angle"], min_row=((i - 1) * spacing_required) + 16 + last_row + 1,
                                               max_row=((i - 1) * spacing_required) + 16 + last_row + number_of_iterations)
            chart.title = str(properties_[j]) + " variation due to" + cases[(m+1)*2-2] + " and " + cases[(m+1)*2-1]

            for k in range(0, int(len(cases) / 2)):
                values = openpyxl.chart.Reference(ws, min_col=fx_col + (j * len(cases)) + (m*len(cases)/2) + k, min_row=((i - 1) * spacing_required) + 16 + last_row + 1,
                                               max_row=((i - 1) * spacing_required) + 16 + last_row + number_of_iterations)
                series = openpyxl.chart.Series(values, xvalues, title_from_data=False)
                chart.series.append(series)


            def anchor_value(items):  # for further increase in the element data the value here must be amended based on column Alphabet
                row_number = str(((i - 1) * spacing_required) + 16 + last_row + number_of_iterations + 1)
                col_number = ((((j*2) + m) * 9 ) + 1)  # 1 is added as A is started from 1 index
                value = num_alphabet[col_number] + row_number
                return (value)


            anchor = anchor_value(items)
            ws.add_chart(chart, anchor)

wb.save(filename=f_result)
wb.close()
#*****************************************************************************************************************
